from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import JsonResponse
from django.db import models, IntegrityError
from django.db.models import Sum, Count, Q, F, Avg
from django.db.models.functions import Extract
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from django.core.cache import cache
import os
import json
import logging
from .models import (
    Cliente, Proyecto, Colaborador, Factura, Pago, 
    Gasto, CategoriaGasto, GastoFijoMensual, LogActividad, Anticipo, AplicacionAnticipo, ArchivoProyecto,
    NotificacionSistema, ConfiguracionNotificaciones, HistorialNotificaciones, IngresoProyecto, Cotizacion,
    ItemInventario, CategoriaInventario, AsignacionInventario,
    Rol, PerfilUsuario, Modulo, Permiso, RolPermiso, AnticipoProyecto,
    CarpetaProyecto, ConfiguracionSistema, EventoCalendario,
    TrabajadorDiario, RegistroTrabajo, AnticipoTrabajadorDiario, PlanillaLiquidada, PlanillaTrabajadoresDiarios,
    ItemCotizacion, ItemReutilizable, ConfiguracionPlanilla,
    ServicioTorrero, RegistroDiasTrabajados, PagoServicioTorrero, Torrero, AsignacionTorrero,
    Subproyecto, NotaPostit, CajaMenuda
)
from .forms_simple import (
    ClienteForm, ProyectoForm, ColaboradorForm, FacturaForm, 
    GastoForm, CategoriaGastoForm, UsuarioForm, RolForm, 
    PerfilUsuarioForm, ArchivoProyectoForm, CarpetaProyectoForm,
    AnticipoForm, PagoForm, EventoCalendarioForm,
    CategoriaInventarioForm, ItemInventarioForm,
    AsignacionInventarioForm, TrabajadorDiarioForm, RegistroTrabajoForm,
    AnticipoTrabajadorDiarioForm, PlanillaTrabajadoresDiariosForm, IngresoProyectoForm, CotizacionForm,
    ConfiguracionPlanillaForm, ServicioTorreroForm, RegistroDiasTrabajarForm, PagoServicioTorreroForm,
    TorreroForm, SubproyectoForm, CajaMenudaForm
)
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from .services import NotificacionService, DashboardService, ProyectoService
from .query_utils import QueryOptimizer, DashboardQueries
from .decorators import api_view, secure_view, cache_view
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
import os
from django.conf import settings

# Configurar logger
logger = logging.getLogger(__name__)

def login_view(request):
    """Vista de login"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            # Registrar actividad
            LogActividad.objects.create(
                usuario=user,
                accion='Login',
                modulo='Sistema',
                descripcion=f'Usuario {username} inició sesión',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return redirect('dashboard')
        else:
            messages.error(request, 'Credenciales incorrectas')
    
    return render(request, 'core/login.html')


@login_required
def logout_view(request):
    """Vista de logout"""
    # Registrar actividad
    LogActividad.objects.create(
        usuario=request.user,
        accion='Logout',
        modulo='Sistema',
        descripcion=f'Usuario {request.user.username} cerró sesión',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    logout(request)
    return redirect('login')


@login_required
def dashboard(request):
    """
    Vista del dashboard principal SUPER SIMPLE Y FUNCIONAL
    """
    try:
        # Datos básicos del sistema
        total_clientes = Cliente.objects.filter(activo=True).count()
        total_proyectos = Proyecto.objects.filter(activo=True).count()
        total_facturado = Factura.objects.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        # Total cobrado = Facturas pagadas + Anticipos aplicados al proyecto
        total_facturas_pagadas = Factura.objects.filter(estado='pagada').aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        total_anticipos_aplicados = Anticipo.objects.filter(aplicado_al_proyecto=True).aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
        total_cobrado = total_facturas_pagadas + total_anticipos_aplicados
        
        # ============================================================================
        # DATOS DE RENTABILIDAD REAL
        # ============================================================================
        
        # Calcular rentabilidad del mes actual
        hoy = timezone.now()
        mes_actual = hoy.month
        año_actual = hoy.year
        
        # Ingresos del mes = Facturas pagadas + Anticipos aplicados al proyecto
        facturas_pagadas_mes = Factura.objects.filter(
            fecha_emision__month=mes_actual,
            fecha_emision__year=año_actual,
            estado='pagada'  # Solo facturas pagadas
        ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        anticipos_aplicados_mes = Anticipo.objects.filter(
            fecha_aplicacion__month=mes_actual,
            fecha_aplicacion__year=año_actual,
            aplicado_al_proyecto=True
        ).aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
        
        ingresos_mes = facturas_pagadas_mes + anticipos_aplicados_mes
        
        # DEBUG: Mostrar diferencia en dashboard
        total_facturado_mes = Factura.objects.filter(
            fecha_emision__month=mes_actual,
            fecha_emision__year=año_actual
        ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        # Log información del dashboard para debugging
        logger.debug(f"Dashboard - Mes: {mes_actual}/{año_actual}, Facturado: ${total_facturado_mes}, Cobrado: ${ingresos_mes}")
        
        # Gastos del mes (gastos aprobados)
        gastos_mes_raw = Gasto.objects.filter(
            fecha_gasto__month=mes_actual,
            fecha_gasto__year=año_actual,
            aprobado=True
        ).aggregate(total=Sum('monto'))['total'] or 0
        gastos_mes = Decimal(str(gastos_mes_raw))
        
        # Calcular rentabilidad
        rentabilidad_mes = ingresos_mes - gastos_mes
        margen_rentabilidad = (rentabilidad_mes / ingresos_mes * 100) if ingresos_mes > 0 else Decimal('0.00')
        
        # Gastos por categoría del mes
        gastos_categoria_mes = Gasto.objects.filter(
            fecha_gasto__month=mes_actual,
            fecha_gasto__year=año_actual,
            aprobado=True
        ).values('categoria__nombre').annotate(
            total=Sum('monto')
        ).order_by('-total')[:5]
        
        # Proyectos más rentables del mes
        proyectos_rentables = []
        proyectos_activos = Proyecto.objects.filter(activo=True)
        
        for proyecto in proyectos_activos:
            # Ingresos del proyecto = Facturas pagadas + Anticipos aplicados (TODOS LOS TIEMPOS)
            facturas_pagadas_proyecto = Factura.objects.filter(
                proyecto=proyecto,
                monto_pagado__gt=0
            ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
            
            anticipos_aplicados_proyecto = Anticipo.objects.filter(
                proyecto=proyecto,
                aplicado_al_proyecto=True
            ).aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
            
            ingresos_proyecto = facturas_pagadas_proyecto + anticipos_aplicados_proyecto
            
            gastos_proyecto_raw = Gasto.objects.filter(
                proyecto=proyecto,
                aprobado=True
            ).aggregate(total=Sum('monto'))['total'] or 0
            gastos_proyecto = Decimal(str(gastos_proyecto_raw))
            
            rentabilidad_proyecto = ingresos_proyecto - gastos_proyecto
            
            # Mostrar todos los proyectos, no solo rentables
            if ingresos_proyecto > 0 or gastos_proyecto > 0:  # Proyectos con actividad
                proyectos_rentables.append({
                    'nombre': proyecto.nombre,
                    'rentabilidad': rentabilidad_proyecto,
                    'ingresos': ingresos_proyecto,
                    'gastos': gastos_proyecto
                })
        
        # Ordenar por rentabilidad
        proyectos_rentables.sort(key=lambda x: x['rentabilidad'], reverse=True)
        proyectos_rentables = proyectos_rentables[:5]  # Top 5
        
        # Convertir Decimals a float para JSON (y calcular margen)
        proyectos_rentables_json = []
        for p in proyectos_rentables:
            ingresos = float(p['ingresos'])
            gastos = float(p['gastos'])
            rentabilidad = float(p['rentabilidad'])
            margen = (rentabilidad / ingresos * 100) if ingresos > 0 else 0
            
            proyectos_rentables_json.append({
                'nombre': p['nombre'],
                'rentabilidad': rentabilidad,
                'ingresos': ingresos,
                'gastos': gastos,
                'margen': round(margen, 2)
            })
        
        # Eventos del calendario (para la agenda)
        eventos_agenda = EventoCalendario.objects.filter(
            creado_por=request.user,
            fecha_inicio__gte=timezone.now().date()
        ).order_by('fecha_inicio')[:10]
        
        # Calendario con eventos básicos (para FullCalendar si se usa)
        eventos_calendario = []
        
        # Eventos de facturas
        facturas = Factura.objects.filter(fecha_vencimiento__isnull=False)[:5]
        for factura in facturas:
            eventos_calendario.append({
                'id': f'factura_{factura.id}',
                'title': f'Vencimiento: {factura.numero_factura}',
                'start': factura.fecha_vencimiento.isoformat(),
                'end': factura.fecha_vencimiento.isoformat(),
                'className': 'evento-factura',
                'backgroundColor': '#dc3545',
                'borderColor': '#dc3545',
                'extendedProps': {
                    'tipo': 'vencimiento',
                    'descripcion': f'Vencimiento de factura {factura.numero_factura}',
                    'todo_el_dia': True
                }
            })
        
        # Eventos de proyectos
        proyectos = Proyecto.objects.filter(fecha_inicio__isnull=False)[:5]
        for proyecto in proyectos:
            eventos_calendario.append({
                'id': f'proyecto_{proyecto.id}',
                'title': f'Inicio: {proyecto.nombre}',
                'start': proyecto.fecha_inicio.isoformat(),
                'end': proyecto.fecha_inicio.isoformat(),
                'className': 'evento-proyecto',
                'backgroundColor': '#28a745',
                'borderColor': '#28a745',
                'extendedProps': {
                    'tipo': 'proyecto',
                    'descripcion': f'Inicio del proyecto {proyecto.nombre}',
                    'todo_el_dia': True
                }
            })
        
        # Eventos del calendario personalizados
        eventos_personalizados = EventoCalendario.objects.filter(creado_por=request.user)
        for evento in eventos_personalizados:
            eventos_calendario.append(evento.to_calendar_event())
        
        # Datos para gráfico de proyectos por estado
        proyectos_por_estado = {
            'Planificación': Proyecto.objects.filter(activo=True, estado='planificacion').count(),
            'En Progreso': Proyecto.objects.filter(activo=True, estado='en_progreso').count(),
            'En Pausa': Proyecto.objects.filter(activo=True, estado='en_pausa').count(),
            'Completado': Proyecto.objects.filter(activo=True, estado='completado').count(),
        }
        
        # Datos para gráfico de gastos por categoría (top 8) - TODOS LOS TIEMPOS
        # Cambiado para mostrar todos los gastos, no solo del mes actual
        gastos_por_categoria = Gasto.objects.filter(
            aprobado=True
        ).values('categoria__nombre').annotate(
            total=Sum('monto')
        ).order_by('-total')[:8]
        
        # Convertir a JSON para el template
        import json
        eventos_calendario_json = json.dumps(eventos_calendario, default=str)
        
        # Datos de proyectos por estado
        estados_proyectos_json = json.dumps(list(proyectos_por_estado.keys()), ensure_ascii=False)
        cantidad_proyectos_json = json.dumps(list(proyectos_por_estado.values()))
        
        # Datos de gastos por categoría - asegurar que sean listas válidas
        categorias_gastos_list = [item['categoria__nombre'] if item['categoria__nombre'] else 'Sin Categoría' for item in gastos_por_categoria]
        montos_gastos_list = [float(item['total']) for item in gastos_por_categoria]
        categorias_gastos_json = json.dumps(categorias_gastos_list, ensure_ascii=False)
        montos_gastos_json = json.dumps(montos_gastos_list)
        
        # Log eventos del calendario para debugging
        logger.debug(f"Eventos calendario: {len(eventos_calendario)} eventos generados")
        
        # ============================================================================
        # DATOS REALES PARA GRÁFICOS - INGRESOS VS GASTOS
        # ============================================================================
        
        # Obtener período seleccionado (por defecto 6 meses)
        periodo = request.GET.get('periodo', '6')
        
        # Calcular datos reales para el gráfico según el período
        
        from datetime import datetime, timedelta
        hoy = timezone.now().date()
        
        if periodo == '1':
            # Mes actual
            mes_actual = hoy.month
            año_actual = hoy.year
            meses_grafico = [hoy.strftime('%b')]
            
            # Ingresos reales del mes: facturas pagadas + anticipos aplicados
            facturas_pagadas_mes = Factura.objects.filter(
                estado='pagada',
                fecha_emision__month=mes_actual,
                fecha_emision__year=año_actual
            ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
            
            anticipos_aplicados_mes = Anticipo.objects.filter(
                aplicado_al_proyecto=True,
                fecha_aplicacion__month=mes_actual,
                fecha_aplicacion__year=año_actual
            ).aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
            
            ingresos_mes_real = facturas_pagadas_mes + anticipos_aplicados_mes
            ingresos_mensuales = [float(ingresos_mes_real)]
            
            # Gastos reales del mes: todos los gastos aprobados
            gastos_mes_real_raw = Gasto.objects.filter(
                aprobado=True,
                fecha_gasto__month=mes_actual,
                fecha_gasto__year=año_actual
            ).aggregate(total=Sum('monto'))['total'] or 0
            gastos_mes_real = Decimal(str(gastos_mes_real_raw))
            
            gastos_mensuales = [float(gastos_mes_real)]
            evolucion_proyectos = [total_proyectos]
            
        elif periodo == '3':
            # 3 meses
            meses_grafico = []
            ingresos_mensuales = []
            gastos_mensuales = []
            evolucion_proyectos = []
            for i in range(3):
                fecha = hoy - timedelta(days=30*i)
                meses_grafico.insert(0, fecha.strftime('%b'))
                
                mes = fecha.month
                año = fecha.year
                
                # Ingresos reales del mes
                facturas_pagadas_mes = Factura.objects.filter(
                    estado='pagada',
                    fecha_emision__month=mes,
                    fecha_emision__year=año
                ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
                
                anticipos_aplicados_mes = Anticipo.objects.filter(
                    aplicado_al_proyecto=True,
                    fecha_aplicacion__month=mes,
                    fecha_aplicacion__year=año
                ).aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
                
                ingresos_mes_real = facturas_pagadas_mes + anticipos_aplicados_mes
                ingresos_mensuales.insert(0, float(ingresos_mes_real))
                
                # Gastos reales del mes
                gastos_mes_real_raw = Gasto.objects.filter(
                    aprobado=True,
                    fecha_gasto__month=mes,
                    fecha_gasto__year=año
                ).aggregate(total=Sum('monto'))['total'] or 0
                gastos_mes_real = Decimal(str(gastos_mes_real_raw))
                
                gastos_mensuales.insert(0, float(gastos_mes_real))
                
                # Proyectos activos en ese mes (aproximación)
                proyectos_mes = Proyecto.objects.filter(
                    activo=True,
                    creado_en__lte=fecha
                ).count()
                evolucion_proyectos.insert(0, proyectos_mes)
        else:
            # 6 meses (por defecto)
            meses_grafico = []
            ingresos_mensuales = []
            gastos_mensuales = []
            evolucion_proyectos = []
            for i in range(6):
                fecha = hoy - timedelta(days=30*i)
                meses_grafico.insert(0, fecha.strftime('%b'))
                
                mes = fecha.month
                año = fecha.year
                
                # Ingresos reales del mes
                facturas_pagadas_mes = Factura.objects.filter(
                    estado='pagada',
                    fecha_emision__month=mes,
                    fecha_emision__year=año
                ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
                
                anticipos_aplicados_mes = Anticipo.objects.filter(
                    aplicado_al_proyecto=True,
                    fecha_aplicacion__month=mes,
                    fecha_aplicacion__year=año
                ).aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
                
                ingresos_mes_real = facturas_pagadas_mes + anticipos_aplicados_mes
                ingresos_mensuales.insert(0, float(ingresos_mes_real))
                
                # Gastos reales del mes
                gastos_mes_real_raw = Gasto.objects.filter(
                    aprobado=True,
                    fecha_gasto__month=mes,
                    fecha_gasto__year=año
                ).aggregate(total=Sum('monto'))['total'] or 0
                gastos_mes_real = Decimal(str(gastos_mes_real_raw))
                
                gastos_mensuales.insert(0, float(gastos_mes_real))
                
                # Proyectos activos en ese mes (aproximación)
                proyectos_mes = Proyecto.objects.filter(
                    activo=True,
                    creado_en__lte=fecha
                ).count()
                evolucion_proyectos.insert(0, proyectos_mes)
        
        
        # Datos adicionales para el dashboard
        total_colaboradores = Colaborador.objects.filter(activo=True).count()
        proyectos_completados = Proyecto.objects.filter(activo=True, estado='completado').count()
        
        # Contexto simplificado
        context = {
            'total_clientes': total_clientes,
            'total_proyectos': total_proyectos,
            'total_facturado': total_facturado,
            'total_cobrado': total_cobrado,
            'total_facturas_pagadas': total_facturas_pagadas,
            'total_anticipos_aplicados': total_anticipos_aplicados,
            'total_colaboradores': total_colaboradores,
            'proyectos_completados': proyectos_completados,
            'eventos_calendario': eventos_agenda,  # Eventos reales del modelo
            'eventos_calendario_json': eventos_calendario_json,
            'estados_proyectos': estados_proyectos_json,
            'cantidad_proyectos': cantidad_proyectos_json,
            'categorias_gastos': categorias_gastos_json,
            'montos_gastos': montos_gastos_json,
            'ingresos_mensuales': json.dumps(ingresos_mensuales),
            'gastos_mensuales': json.dumps(gastos_mensuales),
            'meses_grafico': json.dumps(meses_grafico, ensure_ascii=False),
            'periodo_actual': periodo,
            # ============================================================================
            # DATOS DE RENTABILIDAD REAL PARA EL DASHBOARD
            # ============================================================================
            'ingresos_mes': ingresos_mes,
            'gastos_mes': gastos_mes,
            'rentabilidad_mes': rentabilidad_mes,
            'margen_rentabilidad': margen_rentabilidad,
            'gastos_categoria_mes': gastos_categoria_mes,
            'proyectos_rentables': json.dumps(proyectos_rentables_json),
        }
        
        # Log información del contexto para debugging
        logger.debug(f"Contexto dashboard generado con {len(context)} variables")
        # Verificar tipos de datos en el contexto
        for key, value in context.items():
            if isinstance(value, Decimal):
                logger.debug(f"Variable {key} es Decimal: {value}")
        
        return render(request, 'core/dashboard.html', context)
        
    except Exception as e:
        logger.error(f"Error en dashboard: {str(e)}")
        # Contexto de emergencia
        logger.warning("Usando contexto de emergencia para dashboard")
        context = {
            'total_clientes': 0,
            'total_proyectos': 0,
            'total_facturado': 0,
            'total_cobrado': 0,
            'total_facturas_pagadas': 0,
            'total_anticipos_aplicados': 0,
            'total_colaboradores': 0,
            'proyectos_completados': 0,
            'eventos_calendario': [],
            'eventos_calendario_json': '[]',
            'evolucion_proyectos': '[]',
            'categorias_gastos': '[]',
            'montos_gastos': '[]',
            'ingresos_mensuales': json.dumps([]),
            'gastos_mensuales': json.dumps([]),
            'meses_grafico': json.dumps([]),
            'meses_grafico': [],
        }
        
        return render(request, 'core/dashboard.html', context)


# ===== CRUD CLIENTES =====
@login_required
def clientes_list(request):
    """Lista de clientes"""
    # Obtener filtro de estado desde la URL
    estado_filtro = request.GET.get('estado', 'activos')
    
    # Debug: Log para verificar el filtro
    logger.info(f"Filtro aplicado: {estado_filtro}")
    
    # Aplicar filtro según el parámetro
    if estado_filtro == 'inactivos':
        clientes = Cliente.objects.filter(activo=False).order_by('razon_social')
        logger.info(f"Clientes inactivos encontrados: {clientes.count()}")
    elif estado_filtro == 'todos':
        clientes = Cliente.objects.all().order_by('razon_social')
        logger.info(f"Todos los clientes encontrados: {clientes.count()}")
    else:  # activos por defecto
        clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
        logger.info(f"Clientes activos encontrados: {clientes.count()}")
    
    # Calcular estadísticas correctas
    total_clientes = Cliente.objects.count()
    clientes_activos = Cliente.objects.filter(activo=True).count()
    clientes_inactivos = Cliente.objects.filter(activo=False).count()
    
    context = {
        'clientes': clientes,
        'total_clientes': total_clientes,
        'clientes_activos': clientes_activos,
        'clientes_inactivos': clientes_inactivos,
        'estado_filtro': estado_filtro,
    }
    
    return render(request, 'core/clientes/list.html', context)


@login_required
def cliente_create(request):
    """Crear cliente"""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Clientes',
                descripcion=f'Cliente creado: {cliente.razon_social}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Cliente "{cliente.razon_social}" creado exitosamente')
            return redirect('clientes_list')
    else:
        form = ClienteForm()
    
    return render(request, 'core/clientes/create.html', {'form': form})


@login_required
def cliente_edit(request, cliente_id):
    """Editar cliente"""
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            cliente = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Clientes',
                descripcion=f'Cliente editado: {cliente.razon_social}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Cliente "{cliente.razon_social}" actualizado exitosamente')
            return redirect('clientes_list')
    else:
        form = ClienteForm(instance=cliente)
    
    return render(request, 'core/clientes/edit.html', {'form': form, 'cliente': cliente})


@login_required
def cliente_delete(request, cliente_id):
    """Eliminar cliente (desactivar)"""
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    if request.method == 'POST':
        cliente.activo = False
        cliente.save()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Clientes',
            descripcion=f'Cliente desactivado: {cliente.razon_social}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'Cliente "{cliente.razon_social}" eliminado exitosamente')
        return redirect('clientes_list')
    
    return render(request, 'core/clientes/delete.html', {'cliente': cliente})


@login_required
def cliente_detail(request, cliente_id):
    """Detalle de cliente"""
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    # Obtener proyectos del cliente
    proyectos = Proyecto.objects.filter(cliente=cliente, activo=True).order_by('-creado_en')
    
    # Obtener anticipos del cliente
    anticipos = Anticipo.objects.filter(cliente=cliente).order_by('-fecha_anticipo')
    
    # Obtener facturas del cliente
    facturas = Factura.objects.filter(cliente=cliente).order_by('-fecha_factura')
    
    # Calcular estadísticas
    total_proyectos = proyectos.count()
    total_anticipos = anticipos.count()
    total_facturas = facturas.count()
    
    # Calcular montos
    monto_total_anticipos = anticipos.aggregate(total=Sum('monto'))['total'] or 0
    monto_total_facturas = facturas.aggregate(total=Sum('monto'))['total'] or 0
    
    context = {
        'cliente': cliente,
        'proyectos': proyectos[:10],  # Últimos 10 proyectos
        'anticipos': anticipos[:10],  # Últimos 10 anticipos
        'facturas': facturas[:10],    # Últimas 10 facturas
        'total_proyectos': total_proyectos,
        'total_anticipos': total_anticipos,
        'total_facturas': total_facturas,
        'monto_total_anticipos': monto_total_anticipos,
        'monto_total_facturas': monto_total_facturas,
    }
    
    return render(request, 'core/clientes/detail.html', context)


# ===== CRUD PROYECTOS =====
@login_required
def proyectos_list(request):
    """Lista de proyectos con paginación"""
    
    # Obtener parámetros de paginación
    per_page = int(request.GET.get('per_page', 25))
    page = request.GET.get('page', 1)
    
    # Aplicar filtros si existen
    filters = {}
    estado = request.GET.get('estado')
    cliente_id = request.GET.get('cliente')
    
    if estado:
        filters['estado'] = estado
    if cliente_id:
        filters['cliente_id'] = cliente_id
    
    # Query base optimizada
    proyectos = Proyecto.objects.filter(activo=True).select_related(
        'cliente'
    ).order_by('-creado_en')
    
    # Aplicar filtros
    for field, value in filters.items():
        if value:
            proyectos = proyectos.filter(**{field: value})
    
    # Paginar resultados
    from django.core.paginator import Paginator
    paginator = Paginator(proyectos, per_page)
    
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)
    
    # Obtener opciones de filtro
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    estados = Proyecto.ESTADO_CHOICES
    
    # Obtener estadísticas totales (sin filtros)
    total_proyectos = Proyecto.objects.filter(activo=True).count()
    proyectos_en_progreso = Proyecto.objects.filter(activo=True, estado='en_progreso').count()
    proyectos_completados = Proyecto.objects.filter(activo=True, estado='completado').count()
    proyectos_pendientes = Proyecto.objects.filter(activo=True, estado='pendiente').count()
    proyectos_cancelados = Proyecto.objects.filter(activo=True, estado='cancelado').count()
    
    context = {
        'proyectos': page_obj,
        'clientes': clientes,
        'estados': estados,
        'total_proyectos': total_proyectos,
        'proyectos_en_progreso': proyectos_en_progreso,
        'proyectos_completados': proyectos_completados,
        'proyectos_pendientes': proyectos_pendientes,
        'proyectos_cancelados': proyectos_cancelados,
        'filtros_activos': {
            'estado': request.GET.get('estado'),
            'cliente': request.GET.get('cliente'),
        },
        'paginator': paginator,
        'page_obj': page_obj,
    }
    
    return render(request, 'core/proyectos/list.html', context)


@login_required
def proyecto_create(request):
    """Crear proyecto"""
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            proyecto = form.save(commit=False)
            # Asegurar que el proyecto se cree como activo
            proyecto.activo = True
            proyecto.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Proyectos',
                descripcion=f'Proyecto creado: {proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Proyecto creado exitosamente')
            return redirect('proyectos_list')
    else:
        form = ProyectoForm()
    
    # Obtener clientes activos para el dropdown
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    
    # Obtener estadísticas para el sidebar
    proyectos_activos = Proyecto.objects.filter(activo=True).count()
    # CÁLCULO DE PRESUPUESTO PROMEDIO ELIMINADO - YA NO SE USA
    
    context = {
        'form': form,
        'clientes': clientes,
        'proyectos_activos': proyectos_activos,
        # 'presupuesto_promedio' ELIMINADO - YA NO SE USA
    }
    
    return render(request, 'core/proyectos/create.html', context)


@login_required
def proyecto_edit(request, proyecto_id):
    """Editar proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            proyecto = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Proyectos',
                descripcion=f'Proyecto editado: {proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Proyecto actualizado exitosamente')
            return redirect('proyectos_list')
    else:
        form = ProyectoForm(instance=proyecto)
    
    # Obtener clientes activos para el dropdown
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    
    # Obtener estadísticas para el sidebar
    proyectos_activos = Proyecto.objects.filter(activo=True).count()
    # CÁLCULO DE PRESUPUESTO PROMEDIO ELIMINADO - YA NO SE USA
    
    context = {
        'form': form,
        'proyecto': proyecto,
        'clientes': clientes,
        'proyectos_activos': proyectos_activos,
        # 'presupuesto_promedio' ELIMINADO - YA NO SE USA
    }
    
    return render(request, 'core/proyectos/edit.html', context)


@login_required
def proyecto_delete(request, proyecto_id):
    """Eliminar proyecto (desactivar)"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        proyecto.activo = False
        proyecto.save()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Proyectos',
            descripcion=f'Proyecto desactivado: {proyecto.nombre}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, 'Proyecto eliminado exitosamente')
        return redirect('proyectos_list')
    
    return render(request, 'core/proyectos/delete.html', {'proyecto': proyecto})


# ===== CRUD COLABORADORES =====
@login_required
def colaboradores_list(request):
    """Lista de colaboradores"""
    # Obtener todos los colaboradores (activos e inactivos) para la lista
    colaboradores = Colaborador.objects.all().order_by('nombre')
    
    # Calcular estadísticas correctas
    total_colaboradores = Colaborador.objects.count()
    colaboradores_activos = Colaborador.objects.filter(activo=True).count()
    colaboradores_inactivos = Colaborador.objects.filter(activo=False).count()
    
    context = {
        'colaboradores': colaboradores,
        'total_colaboradores': total_colaboradores,
        'colaboradores_activos': colaboradores_activos,
        'colaboradores_inactivos': colaboradores_inactivos,
    }
    
    return render(request, 'core/colaboradores/list.html', context)


@login_required
def colaborador_detail(request, colaborador_id):
    """Detalle de colaborador"""
    colaborador = get_object_or_404(Colaborador, id=colaborador_id)
    
    # Obtener proyectos del colaborador usando la relación ManyToManyField
    proyectos = colaborador.proyectos.all()
    
    context = {
        'colaborador': colaborador,
        'proyectos': proyectos
    }
    
    return render(request, 'core/colaboradores/detail.html', context)


@login_required
def asignar_colaboradores_proyecto(request, proyecto_id):
    """Asignar colaboradores a un proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        colaboradores_ids = request.POST.getlist('colaboradores')
        
        # Limpiar asignaciones existentes
        proyecto.colaboradores.clear()
        
        # Asignar nuevos colaboradores
        if colaboradores_ids:
            colaboradores = Colaborador.objects.filter(id__in=colaboradores_ids, activo=True)
            proyecto.colaboradores.add(*colaboradores)
            
            messages.success(request, f'Colaboradores asignados exitosamente al proyecto {proyecto.nombre}')
        else:
            messages.info(request, f'No se asignaron colaboradores al proyecto {proyecto.nombre}')
        
        return redirect('proyectos_list')
    
    # Obtener todos los colaboradores activos
    colaboradores_disponibles = Colaborador.objects.filter(activo=True).order_by('nombre')
    colaboradores_asignados = proyecto.colaboradores.all()
    
    context = {
        'proyecto': proyecto,
        'colaboradores_disponibles': colaboradores_disponibles,
        'colaboradores_asignados': colaboradores_asignados,
    }
    
    return render(request, 'core/proyectos/asignar_colaboradores.html', context)


@login_required
def colaborador_create(request):
    """Crear colaborador"""
    if request.method == 'POST':
        form = ColaboradorForm(request.POST)
        if form.is_valid():
            colaborador = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Colaboradores',
                descripcion=f'Colaborador creado: {colaborador.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Colaborador creado exitosamente')
            return redirect('colaboradores_list')
    else:
        form = ColaboradorForm()
    
    return render(request, 'core/colaboradores/create.html', {'form': form})


@login_required
def colaborador_edit(request, colaborador_id):
    """Editar colaborador"""
    colaborador = get_object_or_404(Colaborador, id=colaborador_id)
    
    if request.method == 'POST':
        form = ColaboradorForm(request.POST, instance=colaborador)
        if form.is_valid():
            colaborador = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Colaboradores',
                descripcion=f'Colaborador editado: {colaborador.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Colaborador actualizado exitosamente')
            return redirect('colaboradores_list')
    else:
        form = ColaboradorForm(instance=colaborador)
    
    return render(request, 'core/colaboradores/edit.html', {'form': form, 'colaborador': colaborador})


@login_required
def colaborador_delete(request, colaborador_id):
    """Eliminar colaborador"""
    colaborador = get_object_or_404(Colaborador, id=colaborador_id)
    
    if request.method == 'POST':
        colaborador.delete()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Colaboradores',
            descripcion=f'Colaborador eliminado: {colaborador.nombre}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, 'Colaborador eliminado exitosamente')
        return redirect('colaboradores_list')
    
    return render(request, 'core/colaboradores/delete.html', {'colaborador': colaborador})


# ===== CRUD FACTURAS =====
@login_required
def facturas_list(request):
    """Lista de facturas con paginación"""
    
    # Obtener parámetros de paginación
    per_page = int(request.GET.get('per_page', 25))
    page = request.GET.get('page', 1)
    
    # Aplicar filtros si existen
    filters = {}
    estado = request.GET.get('estado')
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if estado:
        filters['estado'] = estado
    if cliente_id:
        filters['cliente_id'] = cliente_id
    if proyecto_id:
        filters['proyecto_id'] = proyecto_id
    
    # Query base optimizada
    facturas = Factura.objects.select_related(
        'cliente', 'proyecto'
    ).prefetch_related(
        'pagos'
    ).order_by('-fecha_emision')
    
    # Aplicar filtros de fecha si existen
    if fecha_desde:
        facturas = facturas.filter(fecha_emision__gte=fecha_desde)
    if fecha_hasta:
        facturas = facturas.filter(fecha_emision__lte=fecha_hasta)
    
    # Aplicar otros filtros
    for field, value in filters.items():
        if value:
            facturas = facturas.filter(**{field: value})
    
    # Paginar resultados
    from django.core.paginator import Paginator
    paginator = Paginator(facturas, per_page)
    
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)
    
    # Obtener opciones de filtro
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    estados = Factura.ESTADO_CHOICES
    
    # Organizar proyectos por cliente para el filtro dinámico
    proyectos_por_cliente = {}
    for cliente in clientes:
        proyectos_por_cliente[cliente.id] = [
            {'id': p.id, 'nombre': p.nombre} 
            for p in proyectos.filter(cliente=cliente)
        ]
    
    # Obtener estadísticas totales
    total_facturas = Factura.objects.count()
    total_facturado = Factura.objects.aggregate(total=Sum('monto_total'))['total'] or 0.00
    
    # Calcular total cobrado sumando solo las facturas pagadas
    facturas_pagadas = Factura.objects.filter(estado='pagada')
    total_cobrado = facturas_pagadas.aggregate(total=Sum('monto_total'))['total'] or 0.00
    
    facturas_emitidas = Factura.objects.filter(estado='emitida').count()
    facturas_cobradas = Factura.objects.filter(estado='pagada').count()
    
    context = {
        'facturas': page_obj,
        'clientes': clientes,
        'proyectos': proyectos,
        'estados': estados,
        'proyectos_por_cliente': proyectos_por_cliente,
        'total_facturas': total_facturas,
        'total_facturado': total_facturado,
        'total_cobrado': total_cobrado,
        'facturas_emitidas': facturas_emitidas,
        'facturas_pagadas': facturas_cobradas,
        'filtros_activos': {
            'estado': request.GET.get('estado'),
            'cliente': request.GET.get('cliente'),
            'proyecto': request.GET.get('proyecto'),
            'fecha_desde': request.GET.get('fecha_desde'),
            'fecha_hasta': request.GET.get('fecha_hasta'),
        },
        'paginator': paginator,
        'page_obj': page_obj,
    }
    
    return render(request, 'core/facturas/list.html', context)


@login_required
def factura_detail(request, factura_id):
    """Detalle de la factura"""
    factura = get_object_or_404(Factura, id=factura_id)
    
    # Obtener pagos relacionados
    pagos = Pago.objects.filter(factura=factura)
    
    # Calcular balance del proyecto
    proyecto = factura.proyecto
    total_ingresos_proyecto = proyecto.ingresos.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    total_gastos_proyecto = proyecto.get_total_gastos_aprobados()
    proyecto_balance = total_ingresos_proyecto - total_gastos_proyecto
    
    context = {
        'factura': factura,
        'pagos': pagos,
        'proyecto_balance': proyecto_balance
    }
    
    return render(request, 'core/facturas/detail.html', context)


@login_required
def factura_create(request):
    """Crear factura"""
    # Inicializar form por defecto para GET requests
    form = FacturaForm()
    
    # Establecer valores por defecto
    if not request.method == 'POST':
        form.fields['fecha_emision'].initial = timezone.now().date()
        form.fields['fecha_vencimiento'].initial = timezone.now().date() + timedelta(days=30)  # 30 días después
    
    if request.method == 'POST':
        logger.debug("Creando nueva factura")
        form = FacturaForm(request.POST)
        
        if form.is_valid():
            logger.debug("Formulario válido, guardando factura")
            
            factura = form.save(commit=False)
            factura.usuario_creacion = request.user
            factura.estado = 'emitida'
            
            factura.save()
            logger.info(f"Factura {factura.numero_factura} creada exitosamente por {request.user}")
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Facturas',
                descripcion=f'Factura creada: {factura.numero_factura}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Factura creada exitosamente')
            return redirect('facturas_list')
        else:
            logger.warning(f"Formulario inválido: {form.errors}")
            logger.warning(f"Datos del POST: {request.POST}")
            messages.error(request, f'Error en el formulario: {form.errors}')
    
    # Obtener clientes y proyectos activos para los dropdowns
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    # Organizar proyectos por cliente para el filtro dinámico
    proyectos_por_cliente = {}
    for cliente in clientes:
        proyectos_por_cliente[cliente.id] = [
            {'id': p.id, 'nombre': p.nombre} 
            for p in proyectos.filter(cliente=cliente)
        ]
    
    # Obtener estadísticas para el sidebar
    total_facturas = Factura.objects.count()
    total_facturado = Factura.objects.aggregate(total=Sum('monto_total'))['total'] or 0.00
    
    context = {
        'form': form,
        'clientes': clientes,
        'proyectos': proyectos,
        'proyectos_por_cliente': proyectos_por_cliente,
        'total_facturas': total_facturas,
        'total_facturado': total_facturado
    }
    
    return render(request, 'core/facturas/create.html', context)


@login_required
def factura_edit(request, factura_id):
    """Editar factura"""
    factura = get_object_or_404(Factura, id=factura_id)
    
    if request.method == 'POST':
        form = FacturaForm(request.POST, instance=factura)
        if form.is_valid():
            factura = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Facturas',
                descripcion=f'Factura editada: {factura.numero_factura}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Factura actualizada exitosamente')
            return redirect('facturas_list')
    else:
        form = FacturaForm(instance=factura)
    
    # Obtener clientes y proyectos activos para los dropdowns
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    # Obtener estadísticas para el sidebar
    total_facturas = Factura.objects.count()
    total_facturado = Factura.objects.aggregate(total=Sum('monto_total'))['total'] or 0.00
    
    context = {
        'form': form,
        'factura': factura,
        'clientes': clientes,
        'proyectos': proyectos,
        'total_facturas': total_facturas,
        'total_facturado': total_facturado
    }
    
    return render(request, 'core/facturas/edit.html', context)


@login_required
def factura_delete(request, factura_id):
    """Eliminar factura"""
    factura = get_object_or_404(Factura, id=factura_id)
    
    if request.method == 'POST':
        # Registrar actividad antes de eliminar
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Facturas',
            descripcion=f'Factura eliminada: {factura.numero_factura}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        factura.delete()
        messages.success(request, 'Factura eliminada exitosamente')
        return redirect('facturas_list')
    
    return render(request, 'core/facturas/delete.html', {'factura': factura})


@login_required
def factura_marcar_pagada(request, factura_id):
    """Marcar factura como pagada"""
    if request.method == 'POST':
        try:
            factura = get_object_or_404(Factura, id=factura_id)
            
            # Verificar que la factura no esté ya pagada o cancelada
            if factura.estado == 'pagada':
                return JsonResponse({
                    'success': False,
                    'error': 'La factura ya está marcada como pagada'
                })
            
            if factura.estado == 'cancelada':
                return JsonResponse({
                    'success': False,
                    'error': 'No se puede marcar como pagada una factura cancelada'
                })
            
            # Cambiar estado a pagada y establecer fecha de pago
            factura.estado = 'pagada'
            factura.fecha_pago = timezone.now().date()
            # Actualizar monto pagado al monto total de la factura
            factura.monto_pagado = factura.monto_total
            factura.monto_pendiente = factura.monto_total - factura.monto_pagado
            factura.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Marcar como Pagada',
                modulo='Facturas',
                descripcion=f'Factura marcada como pagada: {factura.numero_factura}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Factura {factura.numero_factura} marcada como pagada exitosamente',
                'factura_id': factura.id,
                'nuevo_estado': factura.estado,
                'fecha_pago': factura.fecha_pago.strftime('%d/%m/%Y') if factura.fecha_pago else None
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al procesar la solicitud: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido'
    })


# ===== CRUD GASTOS =====
@login_required
def gastos_list(request):
    """Lista completa de gastos"""
    try:
        # Obtener todos los gastos con información relacionada
        gastos = Gasto.objects.select_related(
            'proyecto', 'categoria', 'aprobado_por'
        ).order_by('-fecha_gasto')
        
        # Filtros
        filtro_estado = request.GET.get('estado', 'todos')
        filtro_categoria = request.GET.get('categoria', '')
        filtro_proyecto = request.GET.get('proyecto', '')
        filtro_fecha_desde = request.GET.get('fecha_desde', '')
        filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
        
        # Aplicar filtros
        if filtro_estado == 'aprobados':
            gastos = gastos.filter(aprobado=True)
        elif filtro_estado == 'pendientes':
            gastos = gastos.filter(aprobado=False)
        
        if filtro_categoria:
            gastos = gastos.filter(categoria_id=filtro_categoria)
        
        if filtro_proyecto:
            gastos = gastos.filter(proyecto_id=filtro_proyecto)
        
        if filtro_fecha_desde:
            gastos = gastos.filter(fecha_gasto__gte=filtro_fecha_desde)
        
        if filtro_fecha_hasta:
            gastos = gastos.filter(fecha_gasto__lte=filtro_fecha_hasta)
        
        # Paginación
        from django.core.paginator import Paginator
        paginator = Paginator(gastos, 20)  # 20 gastos por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Obtener opciones para filtros
        categorias = CategoriaGasto.objects.all().order_by('nombre')
        proyectos = Proyecto.objects.all().order_by('nombre')
        
        # Estadísticas para la página
        total_gastos = gastos.count()
        total_monto = gastos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        gastos_aprobados = gastos.filter(aprobado=True).count()
        gastos_pendientes = gastos.filter(aprobado=False).count()
        
        context = {
            'page_obj': page_obj,
            'egresos': page_obj,  # Cambiado de 'gastos' a 'egresos' para coincidir con el template
            'categorias': categorias,
            'proyectos': proyectos,
            'filtro_estado': filtro_estado,
            'filtro_categoria': filtro_categoria,
            'filtro_proyecto': filtro_proyecto,
            'filtro_fecha_desde': filtro_fecha_desde,
            'filtro_fecha_hasta': filtro_fecha_hasta,
            'total_gastos': total_gastos,
            'total_monto': total_monto,
            'egresos_aprobados': gastos_aprobados,  # Cambiado para consistencia
            'egresos_pendientes': gastos_pendientes,  # Cambiado para consistencia
        }
        
        return render(request, 'core/egresos/list.html', context)  # Cambiado de list_moderno.html a list.html
        
    except Exception as e:
        logger.error(f'Error en gastos_list: {e}')
        messages.error(request, 'Error al cargar la lista de gastos')
        return redirect('egresos_dashboard')


@login_required
def gastos_dashboard(request):
    """Dashboard del módulo de gastos"""
    try:
        # Estadísticas generales
        total_gastos = Gasto.objects.count()
        total_monto = Gasto.objects.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        gastos_aprobados = Gasto.objects.filter(aprobado=True).count()
        gastos_pendientes = Gasto.objects.filter(aprobado=False).count()
        
        # Gastos por categoría
        gastos_por_categoria = Gasto.objects.values('categoria__nombre', 'categoria__color', 'categoria__icono').annotate(
            total=Sum('monto'),
            cantidad=Count('id')
        ).order_by('-total')
        
        # Gastos recientes
        gastos_recientes = Gasto.objects.select_related('categoria', 'proyecto').order_by('-fecha_gasto')[:10]
        
        # Gastos por mes (últimos 6 meses)
        from datetime import datetime, timedelta
        from django.db.models.functions import TruncMonth
        
        fecha_inicio = datetime.now() - timedelta(days=180)
        gastos_por_mes = Gasto.objects.filter(
            fecha_gasto__gte=fecha_inicio
        ).annotate(
            mes=TruncMonth('fecha_gasto')
        ).values('mes').annotate(
            total=Sum('monto')
        ).order_by('mes')
        
        context = {
            'total_gastos': total_gastos,
            'total_monto': total_monto,
            'gastos_aprobados': gastos_aprobados,
            'gastos_pendientes': gastos_pendientes,
            'gastos_por_categoria': gastos_por_categoria,
            'gastos_recientes': gastos_recientes,
            'gastos_por_mes': gastos_por_mes,
        }
        
        return render(request, 'core/egresos/dashboard.html', context)
        
    except Exception as e:
        logger.error(f'Error en gastos_dashboard: {e}')
        messages.error(request, 'Error al cargar el dashboard de gastos')
        return redirect('egresos_list')


@login_required
def gasto_create(request):
    """Crear gasto"""
    if request.method == 'POST':
        logger.info(f"📝 POST recibido para crear gasto")
        logger.info(f"📝 Datos POST: {request.POST}")
        logger.info(f"📝 Archivos FILES: {request.FILES}")
        
        form = GastoForm(request.POST, request.FILES)
        logger.info(f"📝 Formulario creado: {form}")
        
        if form.is_valid():
            logger.info("✅ Formulario es válido, guardando gasto...")
            gasto = form.save()
            logger.info(f"✅ Gasto guardado con ID: {gasto.id}")
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Gastos',
                descripcion=f'Gasto creado: {gasto.descripcion} - ${gasto.monto}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Gasto creado exitosamente')
            return redirect('egresos_dashboard')
        else:
            # Log de errores para debugging
            logger.error(f'❌ Errores en formulario de gasto: {form.errors}')
            logger.error(f'❌ Datos del formulario: {form.data}')
            logger.error(f'❌ Clases de datos: {[(k, type(v)) for k, v in form.data.items()]}')
            messages.error(request, f'Error al crear el gasto. Por favor verifica los campos.')
    else:
        form = GastoForm()
    
    context = {
        'form': form,
        'categorias': CategoriaGasto.objects.all(),
        'proyectos': Proyecto.objects.filter(activo=True)
    }
    
    return render(request, 'core/egresos/create_moderno.html', context)


@login_required
def gasto_aprobar(request, gasto_id):
    """Aprobar o desaprobar un gasto (toggle)"""
    try:
        gasto = Gasto.objects.get(id=gasto_id)
        
        # Toggle del estado de aprobación
        if gasto.aprobado:
            # DESAPROBAR
            proyecto = gasto.proyecto
            if proyecto:
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Desaprobar',
                    modulo='Gastos',
                    descripcion=f'Gasto desaprobado para el proyecto {proyecto.nombre}: {gasto.descripcion} - ${gasto.monto}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            
            gasto.aprobado = False
            gasto.aprobado_por = None
            gasto.save()
            messages.success(request, f'Gasto "{gasto.descripcion}" desaprobado exitosamente')
        else:
            # APROBAR
            gasto.aprobado = True
            gasto.aprobado_por = request.user
            gasto.save()
            
            # Registrar actividad (sin modificar presupuesto)
            proyecto = gasto.proyecto
            if proyecto:
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Aprobar',
                    modulo='Gastos',
                    descripcion=f'Gasto aprobado para el proyecto {proyecto.nombre}: {gasto.descripcion} - ${gasto.monto}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                # Mostrar información del impacto en el proyecto (sin modificar presupuesto)
                # presupuesto_disponible ELIMINADO - YA NO SE USA
                total_gastos = proyecto.get_total_gastos_aprobados()
                print(f"✅ Total gastos aprobados del proyecto: ${total_gastos}")
                print(f"✅ Presupuesto eliminado del sistema - ya no se usa")
            
            messages.success(request, f'Gasto "{gasto.descripcion}" aprobado y aplicado al proyecto exitosamente')
        
        # Retornar respuesta JSON para AJAX o redirigir si es petición normal
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
            from django.http import JsonResponse
            return JsonResponse({'success': True, 'aprobado': gasto.aprobado})
        
        return redirect('egresos_list')
        
    except Gasto.DoesNotExist:
        messages.error(request, 'Gasto no encontrado')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Gasto no encontrado'}, status=404)
        return redirect('egresos_list')
    except Exception as e:
        logger.error(f'Error aprobando/desaprobando gasto {gasto_id}: {e}')
        messages.error(request, 'Error al procesar el gasto')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
        return redirect('egresos_list')


@login_required
def gasto_desaprobar(request, gasto_id):
    """Desaprobar un gasto"""
    try:
        gasto = Gasto.objects.get(id=gasto_id)
        
        if not gasto.aprobado:
            messages.warning(request, 'Este gasto ya está pendiente')
        else:
            # Registrar actividad (sin modificar presupuesto)
            proyecto = gasto.proyecto
            if proyecto:
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Desaprobar',
                    modulo='Gastos',
                    descripcion=f'Gasto desaprobado para el proyecto {proyecto.nombre}: {gasto.descripcion} - ${gasto.monto}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                # Mostrar información del impacto en el proyecto (sin modificar presupuesto)
                # presupuesto_disponible ELIMINADO - YA NO SE USA
                total_gastos = proyecto.get_total_gastos_aprobados()
                print(f"✅ Total gastos aprobados del proyecto: ${total_gastos}")
                print(f"✅ Presupuesto eliminado del sistema - ya no se usa")
                print(f"✅ Presupuesto eliminado del sistema - ya no se usa")
            
            gasto.aprobado = False
            gasto.aprobado_por = None
            gasto.save()
            messages.success(request, f'Gasto "{gasto.descripcion}" desaprobado y revertido del proyecto exitosamente')
        
        return redirect('egresos_list')
        
    except Gasto.DoesNotExist:
        messages.error(request, 'Gasto no encontrado')
        return redirect('egresos_list')
    except Exception as e:
        logger.error(f'Error desaprobando gasto {gasto_id}: {e}')
        messages.error(request, 'Error al desaprobar el gasto')
        return redirect('egresos_list')


@login_required
def gasto_detail(request, gasto_id):
    """Ver detalles de un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'core/egresos/detail.html', context)


@login_required
def gasto_edit(request, gasto_id):
    """Editar gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if request.method == 'POST':
        form = GastoForm(request.POST, instance=gasto)
        if form.is_valid():
            gasto = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Gastos',
                descripcion=f'Gasto editado: {gasto.descripcion} - ${gasto.monto}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Gasto actualizado exitosamente')
            return redirect('egresos_list')
    else:
        form = GastoForm(instance=gasto)
    
    return render(request, 'core/egresos/edit.html', {'form': form, 'gasto': gasto})


@login_required
def gasto_delete(request, gasto_id):
    """Eliminar gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if request.method == 'POST':
        # Registrar actividad antes de eliminar
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Gastos',
            descripcion=f'Gasto eliminado: {gasto.descripcion} - ${gasto.monto}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente')
        return redirect('egresos_list')
    
    return render(request, 'core/egresos/delete.html', {'gasto': gasto})


@login_required
# FUNCIÓN DUPLICADA ELIMINADA - Se usa la de la línea 1409


# ===== CRUD PAGOS =====
@login_required
def pagos_list(request):
    """Lista de pagos"""
    pagos = Pago.objects.all().order_by('-fecha_pago')
    return render(request, 'core/pagos/list.html', {'pagos': pagos})


@login_required
def pago_create(request):
    """Crear pago"""
    if request.method == 'POST':
        form = PagoForm(request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.registrado_por = request.user
            pago.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Pagos',
                descripcion=f'Pago registrado: ${pago.monto} para factura {pago.factura.numero_factura}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Pago registrado exitosamente')
            return redirect('pagos_list')
    else:
        form = PagoForm()
    
    return render(request, 'core/pagos/create.html', {'form': form})


@login_required
def pago_edit(request, pago_id):
    """Editar pago"""
    pago = get_object_or_404(Pago, id=pago_id)
    
    if request.method == 'POST':
        form = PagoForm(request.POST, instance=pago)
        if form.is_valid():
            pago = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Pagos',
                descripcion=f'Pago editado: ${pago.monto} para factura {pago.factura.numero_factura}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Pago actualizado exitosamente')
            return redirect('pagos_list')
    else:
        form = PagoForm(instance=pago)
    
    return render(request, 'core/pagos/edit.html', {'form': form, 'pago': pago})


@login_required
def pago_delete(request, pago_id):
    """Eliminar pago"""
    pago = get_object_or_404(Pago, id=pago_id)
    
    if request.method == 'POST':
        # Registrar actividad antes de eliminar
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Pagos',
            descripcion=f'Pago eliminado: ${pago.monto} para factura {pago.factura.numero_factura}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        pago.delete()
        messages.success(request, 'Pago eliminado exitosamente')
        return redirect('pagos_list')
    
    return render(request, 'core/pagos/delete.html', {'pago': pago})


# ===== CRUD CATEGORÍAS DE GASTO =====
@login_required
def categorias_gasto_list(request):
    """Lista de categorías de gasto"""
    categorias = CategoriaGasto.objects.all().order_by('nombre')
    
    # Calcular estadísticas
    categorias_activas = categorias.count()
    
    context = {
        'categorias': categorias,
        'categorias_activas': categorias_activas,
    }
    
    return render(request, 'core/categorias-egreso/list.html', context)



def categoria_gasto_create(request):
    """Crear categoría de gasto"""
    if request.method == 'POST':
        form = CategoriaGastoForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Categorías de Gasto',
                descripcion=f'Categoría creada: {categoria.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Categoría creada exitosamente')
            return redirect('categoria_egreso_list')
    else:
        form = CategoriaGastoForm()
    
    return render(request, 'core/categorias-egreso/create.html', {'form': form})


@login_required
def categoria_gasto_edit(request, categoria_id):
    """Editar categoría de gasto"""
    categoria = get_object_or_404(CategoriaGasto, id=categoria_id)
    
    if request.method == 'POST':
        form = CategoriaGastoForm(request.POST, instance=categoria)
        if form.is_valid():
            categoria = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Categorías de Gasto',
                descripcion=f'Categoría editada: {categoria.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Categoría actualizada exitosamente')
            return redirect('categoria_egreso_list')
    else:
        form = CategoriaGastoForm(instance=categoria)
    
    return render(request, 'core/categorias-egreso/edit.html', {'form': form, 'categoria': categoria})


@login_required
def categoria_gasto_delete(request, categoria_id):
    """Eliminar categoría de gasto"""
    categoria = get_object_or_404(CategoriaGasto, id=categoria_id)
    
    if request.method == 'POST':
        # Verificar si hay gastos usando esta categoría
        if Gasto.objects.filter(categoria=categoria).exists():
            messages.error(request, 'No se puede eliminar la categoría porque tiene gastos asociados')
            return redirect('categoria_egreso_list')
        
        # Registrar actividad antes de eliminar
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Categorías de Gasto',
            descripcion=f'Categoría eliminada: {categoria.nombre}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente')
        return redirect('categoria_egreso_list')
    
    return render(request, 'core/categorias-egreso/delete.html', {'categoria': categoria})


# ==================== VISTAS DE ANTICIPOS ====================

@login_required
def anticipos_list(request):
    """Lista de anticipos del sistema"""
    anticipos = Anticipo.objects.select_related('cliente', 'proyecto').all()
    
    # Filtros
    estado = request.GET.get('estado')
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')
    
    if estado:
        anticipos = anticipos.filter(estado=estado)
    if cliente_id:
        anticipos = anticipos.filter(cliente_id=cliente_id)
    if proyecto_id:
        anticipos = anticipos.filter(proyecto_id=proyecto_id)
    
    # Estadísticas
    total_anticipos = anticipos.count()
    monto_total = anticipos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    monto_aplicado = anticipos.aggregate(total=Sum('monto_aplicado'))['total'] or Decimal('0.00')
    monto_disponible = anticipos.aggregate(total=Sum('monto_disponible'))['total'] or Decimal('0.00')
    
    # Contar por estado
    anticipos_aplicados = anticipos.filter(estado='aplicado').count()
    anticipos_pendientes = anticipos.filter(estado='pendiente').count()
    
    context = {
        'anticipos': anticipos,
        'total_anticipos': total_anticipos,
        'monto_total': monto_total,
        'monto_aplicado': monto_aplicado,
        'monto_disponible': monto_disponible,
        'anticipos_aplicados': anticipos_aplicados,
        'anticipos_pendientes': anticipos_pendientes,
        'estados': Anticipo.ESTADO_CHOICES,
        'tipos': Anticipo.TIPO_CHOICES,
        'clientes': Cliente.objects.filter(activo=True),
        'proyectos': Proyecto.objects.filter(activo=True),
    }
    
    return render(request, 'core/anticipos/list.html', context)


# ============================================================================
# VISTAS PARA EVENTOS DEL CALENDARIO
# ============================================================================

@login_required
def eventos_calendario_list(request):
    """Lista de eventos del calendario"""
    eventos = EventoCalendario.objects.filter(creado_por=request.user).order_by('-fecha_inicio')
    return render(request, 'core/eventos/list.html', {'eventos': eventos})


@login_required
def evento_calendario_create(request):
    """Crear evento del calendario"""
    if request.method == 'POST':
        form = EventoCalendarioForm(request.POST)
        if form.is_valid():
            evento = form.save(commit=False)
            evento.creado_por = request.user
            evento.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Eventos del Calendario',
                descripcion=f'Evento creado: {evento.titulo}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Evento creado exitosamente')
            return redirect('eventos_calendario_list')
    else:
        form = EventoCalendarioForm()
    
    return render(request, 'core/eventos/create.html', {'form': form})


@login_required
def evento_calendario_edit(request, evento_id):
    """Editar evento del calendario"""
    evento = get_object_or_404(EventoCalendario, id=evento_id, creado_por=request.user)
    
    if request.method == 'POST':
        form = EventoCalendarioForm(request.POST, instance=evento)
        if form.is_valid():
            evento = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Eventos del Calendario',
                descripcion=f'Evento editado: {evento.titulo}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Evento actualizado exitosamente')
            return redirect('eventos_calendario_list')
    else:
        form = EventoCalendarioForm(instance=evento)
    
    return render(request, 'core/eventos/edit.html', {'form': form, 'evento': evento})


@login_required
def evento_calendario_delete(request, evento_id):
    """Eliminar evento del calendario"""
    evento = get_object_or_404(EventoCalendario, id=evento_id, creado_por=request.user)
    
    if request.method == 'POST':
        # Registrar actividad antes de eliminar
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Eventos del Calendario',
            descripcion=f'Evento eliminado: {evento.titulo}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        evento.delete()
        messages.success(request, 'Evento eliminado exitosamente')
        return redirect('eventos_calendario_list')
    
    return render(request, 'core/eventos/delete.html', {'evento': evento})


@login_required
def eventos_calendario_json(request):
    """API para obtener eventos del calendario en formato JSON"""
    eventos = EventoCalendario.objects.filter(creado_por=request.user)
    eventos_data = [evento.to_calendar_event() for evento in eventos]
    
    return JsonResponse(eventos_data, safe=False)


@login_required
def nota_postit_create(request):
    """Crear nota post-it para un evento"""
    if request.method == 'POST':
        try:
            evento_id = request.POST.get('evento_id')
            contenido = request.POST.get('contenido')
            color = request.POST.get('color', '#fef3c7')
            
            if not evento_id or not contenido:
                return JsonResponse({
                    'success': False,
                    'message': 'Evento y contenido son requeridos'
                }, status=400)
            
            evento = get_object_or_404(EventoCalendario, id=evento_id)
            
            nota = NotaPostit.objects.create(
                evento=evento,
                contenido=contenido,
                color=color,
                creado_por=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Post-it creado exitosamente',
                'nota': {
                    'id': nota.id,
                    'contenido': nota.contenido,
                    'color': nota.color
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear post-it: {str(e)}'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

@login_required
def nota_postit_delete(request, nota_id):
    """Eliminar nota post-it"""
    if request.method == 'POST':
        try:
            nota = get_object_or_404(NotaPostit, id=nota_id, creado_por=request.user)
            nota.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Post-it eliminado'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

@login_required
def evento_calendario_create_ajax(request):
    """Crear evento del calendario via AJAX"""
    if request.method == 'POST':
        try:
            # Leer datos desde request.POST (FormData) en lugar de request.body (JSON)
            titulo = request.POST.get('titulo')
            fecha_inicio = request.POST.get('fecha_inicio')
            
            # Validar datos requeridos
            if not titulo or not fecha_inicio:
                return JsonResponse({
                    'success': False,
                    'message': 'Título y fecha de inicio son requeridos'
                }, status=400)
            
            # Procesar fechas desde request.POST
            from datetime import datetime, date
            
            fecha_inicio_str = request.POST.get('fecha_inicio')
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            
            fecha_fin_str = request.POST.get('fecha_fin')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else None
            
            # Procesar horas (opcionales)
            hora_inicio_str = request.POST.get('hora_inicio')
            hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time() if hora_inicio_str else None
            
            hora_fin_str = request.POST.get('hora_fin')
            hora_fin = datetime.strptime(hora_fin_str, '%H:%M').time() if hora_fin_str else None
            
            # Obtener otros datos del formulario
            tipo = request.POST.get('tipo', 'otro')
            descripcion = request.POST.get('descripcion', '')
            color = request.POST.get('color', '#3b82f6')
            
            # Crear evento
            evento = EventoCalendario.objects.create(
                titulo=titulo,
                descripcion=descripcion,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                hora_inicio=hora_inicio,
                hora_fin=hora_fin,
                tipo=tipo,
                color=color,
                todo_el_dia=True,
                creado_por=request.user
            )
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Eventos del Calendario',
                descripcion=f'Evento creado: {evento.titulo}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Evento creado exitosamente',
                'evento': evento.to_calendar_event()
            })
            
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'message': f'Error al crear evento: {str(e)}',
                'debug': traceback.format_exc()
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@require_http_methods(["PUT"])
def evento_calendario_update_ajax(request, evento_id):
    """Actualizar evento del calendario via AJAX"""
    try:
        data = json.loads(request.body)
        
        # Obtener evento
        try:
            evento = EventoCalendario.objects.get(id=evento_id)
        except EventoCalendario.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Evento no encontrado'
            }, status=404)
        
        # Validar datos requeridos
        if not data.get('titulo') or not data.get('fecha_inicio'):
            return JsonResponse({
                'success': False,
                'message': 'Título y fecha de inicio son requeridos'
            }, status=400)
        
        # Procesar fechas
        from datetime import datetime, date
        
        fecha_inicio = data.get('fecha_inicio')
        if isinstance(fecha_inicio, str):
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        
        fecha_fin = data.get('fecha_fin')
        if fecha_fin and isinstance(fecha_fin, str):
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Procesar horas
        hora_inicio = data.get('hora_inicio')
        if hora_inicio and isinstance(hora_inicio, str):
            hora_inicio = datetime.strptime(hora_inicio, '%H:%M').time()
        
        hora_fin = data.get('hora_fin')
        if hora_fin and isinstance(hora_fin, str):
            hora_fin = datetime.strptime(hora_fin, '%H:%M').time()
        
        # Actualizar evento
        evento.titulo = data.get('titulo')
        evento.descripcion = data.get('descripcion', '')
        evento.fecha_inicio = fecha_inicio
        evento.fecha_fin = fecha_fin
        evento.hora_inicio = hora_inicio
        evento.hora_fin = hora_fin
        evento.tipo = data.get('tipo', 'otro')
        evento.color = data.get('color', '#667eea')
        evento.todo_el_dia = data.get('todo_el_dia', True)
        evento.proyecto_id = data.get('proyecto_id') if data.get('proyecto_id') else None
        evento.factura_id = data.get('factura_id') if data.get('factura_id') else None
        evento.save()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Actualizar',
            modulo='Eventos del Calendario',
            descripcion=f'Evento actualizado: {evento.titulo}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Evento actualizado exitosamente',
            'evento': evento.to_calendar_event()
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al actualizar evento: {str(e)}',
            'debug': traceback.format_exc()
        }, status=400)


@require_http_methods(["DELETE"])
def evento_calendario_delete_ajax(request, evento_id):
    """Eliminar evento del calendario via AJAX"""
    try:
        # Obtener evento
        try:
            evento = EventoCalendario.objects.get(id=evento_id)
            titulo_evento = evento.titulo  # Guardar título para el log
        except EventoCalendario.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Evento no encontrado'
            }, status=404)
        
        # Eliminar evento
        evento.delete()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Eventos del Calendario',
            descripcion=f'Evento eliminado: {titulo_evento}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Evento eliminado exitosamente'
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error al eliminar evento: {str(e)}',
            'debug': traceback.format_exc()
        }, status=400)


@login_required
def anticipo_create(request):
    """Crear nuevo anticipo"""
    if request.method == 'POST':
        logger.info(f"📝 POST recibido para crear anticipo")
        logger.info(f"📝 Datos POST: {request.POST}")
        
        form = AnticipoForm(request.POST)
        logger.info(f"📝 Formulario creado: {form}")
        
        if form.is_valid():
            logger.info("✅ Formulario es válido, guardando anticipo...")
            anticipo = form.save(commit=False)
            anticipo.creado_por = request.user
            anticipo.save()
            logger.info(f"✅ Anticipo guardado con ID: {anticipo.id}")
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Anticipos',
                descripcion=f'Anticipo creado: {anticipo.numero_anticipo} - ${anticipo.monto}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 
                f'✅ <strong>Anticipo creado exitosamente</strong><br>'
                f'💰 Monto: <strong>${anticipo.monto:,.2f}</strong><br>'
                f'🏗️ Proyecto: <strong>{anticipo.proyecto.nombre}</strong><br>'
                f'👤 Cliente: <strong>{anticipo.cliente.razon_social}</strong>',
                extra_tags='html'
            )
            logger.info("✅ Redirigiendo a anticipos_list")
            return redirect('anticipos_list')
        else:
            logger.error(f"❌ Formulario no es válido. Errores: {form.errors}")
            logger.error(f"❌ Datos del formulario: {form.cleaned_data}")
            # Agregar los errores del formulario a los mensajes
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = AnticipoForm()
    
    # Obtener clientes y proyectos activos para los dropdowns
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    # Organizar proyectos por cliente para el filtro dinámico
    proyectos_por_cliente = {}
    for cliente in clientes:
        proyectos_del_cliente = proyectos.filter(cliente=cliente)
        proyectos_por_cliente[cliente.id] = [
            {'id': p.id, 'nombre': p.nombre} 
            for p in proyectos_del_cliente
        ]
        # Debug: Log para verificar datos
        logger.info(f"Cliente {cliente.razon_social} (ID: {cliente.id}) tiene {len(proyectos_del_cliente)} proyectos")
    
    # Debug adicional
    logger.info(f"Proyectos por cliente: {proyectos_por_cliente}")
    
    context = {
        'form': form,
        'clientes': clientes,
        'proyectos': proyectos,
        'proyectos_por_cliente': proyectos_por_cliente,
    }
    
    return render(request, 'core/anticipos/create.html', context)


@login_required
def anticipo_detail(request, anticipo_id):
    """Detalle de un anticipo"""
    anticipo = get_object_or_404(Anticipo, id=anticipo_id)
    
    # Obtener facturas donde se puede aplicar este anticipo
    facturas_aplicables = Factura.objects.filter(
        cliente=anticipo.cliente,
        proyecto=anticipo.proyecto,
        estado__in=['emitida', 'enviada', 'vencida']
    ).exclude(
        monto_anticipos__gte=F('monto_total')
    )
    
    context = {
        'anticipo': anticipo,
        'facturas_aplicables': facturas_aplicables,
    }
    
    return render(request, 'core/anticipos/detail.html', context)


@login_required
def anticipo_edit(request, anticipo_id):
    """Editar anticipo"""
    anticipo = get_object_or_404(Anticipo, id=anticipo_id)
    
    if request.method == 'POST':
        form = AnticipoForm(request.POST, instance=anticipo)
        if form.is_valid():
            anticipo = form.save(commit=False)
            anticipo.modificado_por = request.user
            anticipo.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Anticipos',
                descripcion=f'Anticipo editado: {anticipo.numero_anticipo}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 
                f'✅ <strong>Anticipo actualizado exitosamente</strong><br>'
                f'💰 Monto: <strong>${anticipo.monto:,.2f}</strong><br>'
                f'🏗️ Proyecto: <strong>{anticipo.proyecto.nombre}</strong><br>'
                f'👤 Cliente: <strong>{anticipo.cliente.razon_social}</strong>',
                extra_tags='html'
            )
            return redirect('anticipos_list')
        else:
            # Mostrar errores de validación
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = AnticipoForm(instance=anticipo)
    
    context = {
        'form': form,
        'anticipo': anticipo,
    }
    
    return render(request, 'core/anticipos/edit.html', context)


@login_required
def anticipo_delete(request, anticipo_id):
    """Eliminar anticipo"""
    anticipo = get_object_or_404(Anticipo, id=anticipo_id)
    
    if request.method == 'POST':
        numero_anticipo = anticipo.numero_anticipo
        anticipo.delete()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Anticipos',
            descripcion=f'Anticipo eliminado: {numero_anticipo}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, 
            f'🗑️ <strong>Anticipo eliminado exitosamente</strong><br>'
            f'💰 Monto: <strong>${anticipo.monto:,.2f}</strong><br>'
            f'🏗️ Proyecto: <strong>{anticipo.proyecto.nombre}</strong><br>'
            f'👤 Cliente: <strong>{anticipo.cliente.razon_social}</strong>',
            extra_tags='html'
        )
        return redirect('anticipos_list')
    
    return render(request, 'core/anticipos/delete.html', {'anticipo': anticipo})


@login_required
def aplicar_anticipo(request, anticipo_id):
    """Aplicar anticipo a facturas o al proyecto"""
    anticipo = get_object_or_404(Anticipo, id=anticipo_id)
    
    if request.method == 'POST':
        tipo_aplicacion = request.POST.get('tipo_aplicacion')
        monto_aplicar = Decimal(request.POST.get('monto_aplicar'))
        
        try:
            if tipo_aplicacion == 'factura':
                factura_id = request.POST.get('factura')
                factura = Factura.objects.get(id=factura_id)
                anticipo.aplicar_a_factura(factura, monto_aplicar)
                
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Aplicar Anticipo',
                    modulo='Anticipos',
                    descripcion=f'Anticipo {anticipo.numero_anticipo} aplicado a factura {factura.numero_factura} por ${monto_aplicar}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, 
                    f'✅ <strong>Anticipo aplicado a factura</strong><br>'
                    f'💰 Monto: <strong>${monto_aplicar:,.2f}</strong><br>'
                    f'📄 Factura: <strong>{factura.numero_factura}</strong><br>'
                    f'🏗️ Proyecto: <strong>{anticipo.proyecto.nombre}</strong>',
                    extra_tags='html'
                )
                
            elif tipo_aplicacion == 'proyecto':
                anticipo.aplicar_al_proyecto(monto_aplicar)
                
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Aplicar Anticipo',
                    modulo='Anticipos',
                    descripcion=f'Anticipo {anticipo.numero_anticipo} aplicado al proyecto {anticipo.proyecto.nombre} por ${monto_aplicar}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, 
                    f'✅ <strong>Anticipo aplicado exitosamente</strong><br>'
                    f'💰 Monto: <strong>${monto_aplicar:,.2f}</strong><br>'
                    f'🏗️ Proyecto: <strong>{anticipo.proyecto.nombre}</strong><br>'
                    f'👤 Cliente: <strong>{anticipo.cliente.razon_social}</strong>',
                    extra_tags='html'
                )
            
            return redirect('anticipo_detail', anticipo_id=anticipo.id)
            
        except (Factura.DoesNotExist, ValueError) as e:
            messages.error(request, f'Error al aplicar anticipo: {str(e)}')
    
    # Obtener facturas aplicables
    facturas_aplicables = Factura.objects.filter(
        cliente=anticipo.cliente,
        proyecto=anticipo.proyecto,
        estado__in=['emitida', 'enviada', 'vencida']
    ).exclude(
        monto_anticipos__gte=F('monto_total')
    )
    
    context = {
        'anticipo': anticipo,
        'facturas_aplicables': facturas_aplicables,
    }
    
    return render(request, 'core/anticipos/aplicar.html', context)


# ==================== VISTAS DE ARCHIVOS DE PROYECTOS ====================

@login_required
def proyecto_dashboard(request, proyecto_id=None):
    """Dashboard específico de un proyecto con selector"""
    if proyecto_id:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id, activo=True)
    else:
        # Si no se especifica proyecto, mostrar el primero activo o redirigir
        proyectos = Proyecto.objects.filter(activo=True)
        if proyectos.exists():
            proyecto = proyectos.first()
        else:
            messages.warning(request, 'No hay proyectos activos para mostrar')
            return redirect('proyectos_list')
    
    # Obtener todos los proyectos para el selector
    todos_proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    # Estadísticas del proyecto específico
    facturas_proyecto = Factura.objects.filter(proyecto=proyecto)
    gastos_proyecto = Gasto.objects.filter(proyecto=proyecto)
    gastos_aprobados = Gasto.objects.filter(proyecto=proyecto, aprobado=True)
    anticipos_proyecto = Anticipo.objects.filter(proyecto=proyecto)
    
    # Totales financieros del proyecto
    total_facturado = facturas_proyecto.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    
    # Total gastos: usar los métodos del modelo para consistencia
    total_gastos = proyecto.get_total_gastos_aprobados()  # Solo gastos aprobados
    total_gastos_aprobados = proyecto.get_total_gastos_aprobados()
    total_gastos_pendientes = proyecto.get_total_gastos_pendientes()
    
    # Estadísticas de anticipos
    total_anticipos = anticipos_proyecto.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    total_anticipos_aplicados = anticipos_proyecto.aggregate(total=Sum('monto_aplicado'))['total'] or Decimal('0.00')
    total_anticipos_disponibles_base = anticipos_proyecto.aggregate(total=Sum('monto_disponible'))['total'] or Decimal('0.00')
    
    # Anticipos aplicados directamente al proyecto
    anticipos_aplicados_proyecto = anticipos_proyecto.filter(aplicado_al_proyecto=True)
    total_anticipos_aplicados_proyecto = anticipos_aplicados_proyecto.aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
    
    # Fondos disponibles: anticipos aplicados al proyecto - gastos aprobados del proyecto
    total_anticipos_disponibles = max(total_anticipos_aplicados_proyecto - total_gastos, Decimal('0.00'))
    
    # Fondos pendientes: monto disponible de anticipos que aún no se ha aplicado
    total_anticipos_pendientes = total_anticipos_disponibles_base
    
    # Facturas pagadas (solo para mostrar en estadísticas)
    facturas_pagadas = facturas_proyecto.filter(estado='pagada')
    total_facturas_pagadas = facturas_pagadas.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    
    # Total cobrado REAL: facturas pagadas + anticipos aplicados a facturas + anticipos aplicados al proyecto
    total_cobrado = total_facturas_pagadas + total_anticipos_aplicados + total_anticipos_aplicados_proyecto
    
    # Anticipos recientes
    anticipos_recientes = anticipos_proyecto.order_by('-fecha_recepcion')[:5]
    
    # Archivos del proyecto
    archivos_proyecto = ArchivoProyecto.objects.filter(proyecto=proyecto, activo=True)
    
    # Facturas recientes del proyecto
    facturas_recientes = facturas_proyecto.order_by('-fecha_emision')[:5]
    
    # Gastos recientes del proyecto
    gastos_recientes = gastos_proyecto.order_by('-fecha_gasto')[:5]
    
    # Calcular histórico de nómina (Personal + Trabajadores Diarios)
    # 1. Histórico de Personal (planillas regulares - sin observaciones de trabajadores diarios)
    planillas_personal_liquidadas = PlanillaLiquidada.objects.filter(
        proyecto=proyecto
    ).exclude(observaciones__icontains='trabajadores diarios')
    total_historico_personal = planillas_personal_liquidadas.aggregate(total=Sum('total_planilla'))['total'] or Decimal('0.00')
    
    # 2. Histórico de Trabajadores Diarios (planillas de trabajadores diarios)
    planillas_trabajadores_diarios = PlanillaLiquidada.objects.filter(
        proyecto=proyecto,
        observaciones__icontains='trabajadores diarios'
    )
    total_historico_trabajadores_diarios = planillas_trabajadores_diarios.aggregate(total=Sum('total_planilla'))['total'] or Decimal('0.00')
    
    # 3. Total Histórico de Nómina Combinado
    total_historico_nomina = total_historico_personal + total_historico_trabajadores_diarios
    
    # Cálculo del Balance del Proyecto (Ingresos + Facturas - Gastos - Nómina Histórica)
    # IMPORTANTE: El total histórico de nómina SÍ afecta el balance porque son gastos reales del proyecto
    total_ingresos_proyecto = proyecto.ingresos.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    total_facturas_proyecto = proyecto.facturas.filter(estado='pagada').aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    total_ingresos_totales = total_ingresos_proyecto + total_facturas_proyecto
    
    # Total de gastos incluyendo nómina histórica (gastos aprobados + planillas liquidadas)
    total_gastos_completo = total_gastos + total_historico_nomina
    
    # Rentabilidad del proyecto: Total cobrado REAL - (Total gastos aprobados + Nómina Histórica)
    # IMPORTANTE: La nómina histórica SÍ afecta la rentabilidad porque son costos reales del proyecto
    rentabilidad_proyecto = total_cobrado - total_gastos_completo
    
    # Balance del proyecto: Ingresos - (Gastos + Nómina Histórica)
    proyecto_balance = total_ingresos_totales - total_gastos_completo
    
    context = {
        'proyecto': proyecto,
        'todos_proyectos': todos_proyectos,
        'total_facturado': total_facturado,
        'total_cobrado': total_cobrado,
        'total_facturas_pagadas': total_facturas_pagadas,
        'total_gastos': total_gastos,
        'total_gastos_aprobados': total_gastos_aprobados,
        'total_gastos_pendientes': total_gastos_pendientes,
        'total_gastos_completo': total_gastos_completo,  # Gastos + Nómina Histórica
        'total_anticipos': total_anticipos,
        'total_anticipos_aplicados': total_anticipos_aplicados,
        'total_anticipos_disponibles': total_anticipos_disponibles,
        'total_anticipos_disponibles_base': total_anticipos_disponibles_base,
        'total_anticipos_aplicados_proyecto': total_anticipos_aplicados_proyecto,
        'total_anticipos_pendientes': total_anticipos_pendientes,
        'rentabilidad_proyecto': rentabilidad_proyecto,
        'archivos_proyecto': archivos_proyecto,
        'facturas_recientes': facturas_recientes,
        'gastos_recientes': gastos_recientes,
        'anticipos_recientes': anticipos_recientes,
        'total_archivos': archivos_proyecto.count(),
        'total_facturas': facturas_proyecto.count(),
        'total_gastos_count': gastos_proyecto.count(),
        'facturas_pagadas_count': facturas_pagadas.count(),
        'total_anticipos_count': anticipos_proyecto.count(),
        'total_historico_personal': total_historico_personal,
        'total_historico_trabajadores_diarios': total_historico_trabajadores_diarios,
        'total_historico_nomina': total_historico_nomina,
        'total_ingresos_proyecto': total_ingresos_proyecto,
        'total_facturas_proyecto': total_facturas_proyecto,
        'total_ingresos_totales': total_ingresos_totales,
        'proyecto_balance': proyecto_balance,
    }
    
    return render(request, 'core/proyecto_dashboard.html', context)


@login_required
def archivos_proyectos_list(request):
    """Lista de todos los proyectos para gestión de archivos"""
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    # Estadísticas de archivos por proyecto
    for proyecto in proyectos:
        proyecto.total_archivos = ArchivoProyecto.objects.filter(
            proyecto=proyecto, 
            activo=True
        ).count()
    
    context = {
        'proyectos': proyectos,
        'total_proyectos': proyectos.count(),
    }
    
    return render(request, 'core/archivos/proyectos_list.html', context)


@login_required
def archivos_proyecto_list(request, proyecto_id):
    """Lista de archivos de un proyecto específico"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id, activo=True)
    
    # Obtener carpeta actual (si se especifica)
    carpeta_id = request.GET.get('carpeta')
    carpeta_actual = None
    
    if carpeta_id:
        try:
            carpeta_actual = CarpetaProyecto.objects.get(id=carpeta_id, proyecto=proyecto, activa=True)
        except CarpetaProyecto.DoesNotExist:
            pass
    
    # Obtener archivos
    if carpeta_actual:
        archivos = ArchivoProyecto.objects.filter(proyecto=proyecto, carpeta=carpeta_actual, activo=True)
        subcarpetas = carpeta_actual.get_subcarpetas_activas()
    else:
        # Carpeta raíz - archivos sin carpeta y carpetas raíz
        archivos = ArchivoProyecto.objects.filter(proyecto=proyecto, carpeta__isnull=True, activo=True)
        subcarpetas = CarpetaProyecto.objects.filter(proyecto=proyecto, carpeta_padre__isnull=True, activa=True)
    
    # Filtros
    tipo = request.GET.get('tipo')
    if tipo:
        archivos = archivos.filter(tipo=tipo)
    
    # Obtener todas las carpetas del proyecto para el breadcrumb
    todas_carpetas = CarpetaProyecto.objects.filter(proyecto=proyecto, activa=True).order_by('nombre')
    
    context = {
        'proyecto': proyecto,
        'carpeta_actual': carpeta_actual,
        'carpeta_filtrada': carpeta_actual,  # Para el template
        'archivos': archivos,
        'subcarpetas': subcarpetas,
        'todas_carpetas': todas_carpetas,
        'tipos': ArchivoProyecto.TIPO_CHOICES,
        'total_archivos': archivos.count(),
        'total_subcarpetas': subcarpetas.count(),
    }
    
    return render(request, 'core/archivos/list.html', context)


@login_required
def archivo_upload(request, proyecto_id):
    """Subir nuevo archivo a un proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id, activo=True)
    
    # Obtener carpeta_id de la URL si existe
    carpeta_id = request.GET.get('carpeta')
    carpeta_seleccionada = None
    logger.info(f"🔍 DEBUG carpeta_id: {carpeta_id}")
    logger.info(f"🔍 DEBUG URL completa: {request.get_full_path()}")
    if carpeta_id:
        try:
            carpeta_seleccionada = CarpetaProyecto.objects.get(id=carpeta_id, proyecto=proyecto, activa=True)
            logger.info(f"✅ Carpeta encontrada: {carpeta_seleccionada.nombre}")
        except CarpetaProyecto.DoesNotExist:
            logger.warning(f"⚠️ Carpeta {carpeta_id} no encontrada")
            pass
    
    if request.method == 'POST':
        logger.info(f"=" * 80)
        logger.info(f"📤 POST request recibido para subir archivo al proyecto {proyecto.id}")
        logger.info(f"📤 POST data: {request.POST}")
        logger.info(f"📤 FILES: {request.FILES}")
        
        form = ArchivoProyectoForm(request.POST, request.FILES, proyecto=proyecto)
        logger.info(f"📋 Formulario creado")
        logger.info(f"📋 Es válido: {form.is_valid()}")
        logger.info(f"📋 Datos del formulario: {form.cleaned_data if form.is_valid() else form.errors}")
        
        if form.is_valid():
            logger.info(f"✅ Formulario válido, intentando guardar...")
            try:
                archivo = form.save(commit=False)
                archivo.proyecto = proyecto
                archivo.subido_por = request.user
                
                # PRIORIDAD: Si se especificó carpeta en la URL, usar esa
                if carpeta_seleccionada:
                    archivo.carpeta = carpeta_seleccionada
                    logger.info(f"📁 Asignando carpeta desde URL: {carpeta_seleccionada.nombre}")
                # Si no hay carpeta en URL pero el formulario tiene una, usar esa
                elif form.cleaned_data.get('carpeta'):
                    archivo.carpeta = form.cleaned_data['carpeta']
                    logger.info(f"📁 Asignando carpeta desde formulario: {archivo.carpeta.nombre}")
                else:
                    logger.warning(f"⚠️ No hay carpeta seleccionada, archivo se guardará sin carpeta")
                logger.info(f"💾 Guardando archivo: {archivo.nombre}")
                logger.info(f"💾 Proyecto: {archivo.proyecto}")
                logger.info(f"💾 Carpeta: {archivo.carpeta}")
                logger.info(f"💾 Tipo: {archivo.tipo}")
                
                archivo.save()
                logger.info(f"✅ Archivo guardado exitosamente: ID={archivo.id}")
                
                # Generar thumbnail automáticamente
                try:
                    archivo.generar_thumbnail()
                except Exception as e:
                    logger.warning(f"⚠️ Error generando thumbnail: {e}")
                    pass
            
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Subir Archivo',
                    modulo='Archivos',
                    descripcion=f'Archivo subido: {archivo.nombre} al proyecto {proyecto.nombre}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, f'✅ Archivo "{archivo.nombre}" subido exitosamente')
                return redirect('archivos_proyecto_list', proyecto_id=proyecto.id)
            except Exception as e:
                logger.error(f"❌ Error guardando archivo: {e}")
                logger.error(f"❌ Tipo de error: {type(e)}")
                import traceback
                logger.error(f"❌ Traceback: {traceback.format_exc()}")
                messages.error(request, f'Error al subir archivo: {str(e)}')
        else:
            logger.error(f"❌ Formulario NO es válido")
            logger.error(f"❌ Errores del formulario: {form.errors}")
            logger.error(f"❌ Errores por campo:")
            for field, errors in form.errors.items():
                logger.error(f"   - {field}: {errors}")
            messages.error(request, f'Error en el formulario. Por favor verifica los datos.')
    else:
        form = ArchivoProyectoForm(proyecto=proyecto)
        # Pre-seleccionar la carpeta si se especificó en la URL
        if carpeta_seleccionada:
            form.fields['carpeta'].initial = carpeta_seleccionada
    
    context = {
        'form': form,
        'proyecto': proyecto,
        'carpeta_seleccionada': carpeta_seleccionada,
    }
    
    return render(request, 'core/archivos/upload.html', context)


@login_required
def archivo_upload_carpeta(request, carpeta_id):
    """Subir nuevo archivo directamente a una carpeta específica"""
    carpeta = get_object_or_404(CarpetaProyecto, id=carpeta_id, activa=True)
    proyecto = carpeta.proyecto
    
    logger.info(f"🔍 DEBUG Subiendo archivo a carpeta: {carpeta.nombre} (ID: {carpeta.id})")
    
    if request.method == 'POST':
        logger.info(f"=" * 80)
        logger.info(f"📤 POST request recibido para subir archivo a carpeta {carpeta.id}")
        logger.info(f"📤 POST data: {request.POST}")
        logger.info(f"📤 FILES: {request.FILES}")
        
        form = ArchivoProyectoForm(request.POST, request.FILES, proyecto=proyecto)
        logger.info(f"📋 Formulario creado")
        logger.info(f"📋 Es válido: {form.is_valid()}")
        logger.info(f"📋 Datos del formulario: {form.cleaned_data if form.is_valid() else form.errors}")
        
        if form.is_valid():
            logger.info(f"✅ Formulario válido, intentando guardar...")
            try:
                archivo = form.save(commit=False)
                archivo.proyecto = proyecto
                archivo.subido_por = request.user
                archivo.carpeta = carpeta  # FORZAR carpeta específica
                logger.info(f"📁 FORZANDO carpeta: {carpeta.nombre}")
                
                archivo.save()
                logger.info(f"✅ Archivo guardado exitosamente: ID={archivo.id}")
                
                # Generar thumbnail automáticamente
                try:
                    archivo.generar_thumbnail()
                except Exception as e:
                    logger.warning(f"⚠️ Error generando thumbnail: {e}")
                    pass
                
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Subir Archivo',
                    modulo='Archivos',
                    descripcion=f'Archivo subido: {archivo.nombre} a la carpeta {carpeta.nombre} del proyecto {proyecto.nombre}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, f'✅ Archivo "{archivo.nombre}" subido exitosamente a la carpeta "{carpeta.nombre}"')
                return redirect('carpeta_detail', carpeta_id=carpeta.id)
            except Exception as e:
                logger.error(f"❌ Error guardando archivo: {e}")
                logger.error(f"❌ Tipo de error: {type(e)}")
                import traceback
                logger.error(f"❌ Traceback: {traceback.format_exc()}")
                messages.error(request, f'Error al subir archivo: {str(e)}')
        else:
            logger.error(f"❌ Formulario NO es válido")
            logger.error(f"❌ Errores del formulario: {form.errors}")
            logger.error(f"❌ Errores por campo:")
            for field, errors in form.errors.items():
                logger.error(f"   - {field}: {errors}")
            messages.error(request, f'Error en el formulario. Por favor verifica los datos.')
    else:
        form = ArchivoProyectoForm(proyecto=proyecto)
        # Pre-seleccionar la carpeta específica
        form.fields['carpeta'].initial = carpeta
    
    context = {
        'form': form,
        'proyecto': proyecto,
        'carpeta': carpeta,
        'carpeta_seleccionada': carpeta,
    }
    
    return render(request, 'core/archivos/upload.html', context)


@login_required
def archivo_download(request, archivo_id):
    """Descargar un archivo"""
    archivo = get_object_or_404(ArchivoProyecto, id=archivo_id, activo=True)
    
    # Verificar permisos (opcional: solo usuarios del proyecto)
    if not request.user.is_superuser and archivo.proyecto.activo == False:
        messages.error(request, 'No tienes permisos para acceder a este archivo')
        return redirect('proyectos_list')
    
    # Registrar descarga
    LogActividad.objects.create(
        usuario=request.user,
        accion='Descargar Archivo',
        modulo='Archivos',
        descripcion=f'Archivo descargado: {archivo.nombre}',
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    # Retornar archivo para descarga
    from django.http import FileResponse
    
    # Verificar si el archivo tiene contenido físico
    if not archivo.archivo:
        # Para archivos generados (como planillas), crear contenido temporal
        if archivo.es_documento and archivo.get_extension() in ['txt', 'pdf']:
            from django.core.files.base import ContentFile
            if archivo.get_extension() == 'pdf':
                # Para PDFs, crear un mensaje de error más claro
                messages.error(request, 'El archivo PDF no está disponible para descarga')
                return redirect('archivos_proyecto_list', proyecto_id=archivo.proyecto.id)
            else:
                # Para archivos de texto, crear contenido temporal
                contenido = f"Archivo: {archivo.nombre}\nProyecto: {archivo.proyecto.nombre}\nDescripción: {archivo.descripcion}\nFecha: {archivo.fecha_subida.strftime('%d/%m/%Y %H:%M')}"
                buffer = ContentFile(contenido.encode('utf-8'))
                response = FileResponse(buffer, as_attachment=True)
                response['Content-Disposition'] = f'attachment; filename="{archivo.nombre}"'
                return response
        else:
            messages.error(request, 'No hay archivo asociado para descargar')
            return redirect('archivos_proyecto_list', proyecto_id=archivo.proyecto.id)
    
    try:
        file_path = archivo.archivo.path
        if os.path.exists(file_path):
            response = FileResponse(open(file_path, 'rb'))
            response['Content-Disposition'] = f'attachment; filename="{archivo.nombre}.{archivo.get_extension()}"'
            return response
        else:
            messages.error(request, 'El archivo no existe en el servidor')
            return redirect('archivos_proyecto_list', proyecto_id=archivo.proyecto.id)
    except ValueError as e:
        messages.error(request, f'Error al acceder al archivo: {str(e)}')
        return redirect('archivos_proyecto_list', proyecto_id=archivo.proyecto.id)


@login_required
def archivo_delete(request, archivo_id):
    """Eliminar un archivo"""
    archivo = get_object_or_404(ArchivoProyecto, id=archivo_id, activo=True)
    
    if request.method == 'POST':
        # Validar confirmación
        confirmacion = request.POST.get('confirmacion', '').strip().upper()
        
        print(f"🔍 Confirmación recibida: '{confirmacion}'")
        print(f"🔍 Método: {request.method}")
        print(f"🔍 POST data: {request.POST}")
        
        if confirmacion != 'ELIMINAR':
            print(f"❌ Confirmación inválida: '{confirmacion}' != 'ELIMINAR'")
            messages.error(request, 'Debes escribir "ELIMINAR" para confirmar la eliminación.')
            return render(request, 'core/archivos/delete.html', {'archivo': archivo})
        
        print(f"✅ Confirmación válida, procediendo con eliminación...")
        
        try:
            # Registrar actividad antes de eliminar
            LogActividad.objects.create(
                usuario=request.user,
                accion='Eliminar Archivo',
                modulo='Archivos',
                descripcion=f'Archivo eliminado: {archivo.nombre} del proyecto {archivo.proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
        
            # Obtener el ID del proyecto antes de eliminar
            proyecto_id = archivo.proyecto.id
            
            # Eliminar el archivo físico del servidor si existe
            if archivo.archivo and archivo.archivo.name:
                try:
                    if os.path.isfile(archivo.archivo.path):
                        os.remove(archivo.archivo.path)
                        logger.info(f"Archivo físico eliminado: {archivo.archivo.path}")
                except Exception as e:
                    logger.warning(f"Error eliminando archivo físico: {e}")
            
            # Eliminar el thumbnail si existe
            if archivo.thumbnail and archivo.thumbnail.name:
                try:
                    if os.path.isfile(archivo.thumbnail.path):
                        os.remove(archivo.thumbnail.path)
                        logger.info(f"Thumbnail eliminado: {archivo.thumbnail.path}")
                except Exception as e:
                    logger.warning(f"Error eliminando thumbnail: {e}")
            
            # Eliminar el registro de la base de datos
            archivo.delete()
            
            print(f"✅ Archivo eliminado exitosamente: {archivo.nombre}")
            messages.success(request, f'Archivo "{archivo.nombre}" eliminado exitosamente')
            return redirect('archivos_proyecto_list', proyecto_id=proyecto_id)
            
        except Exception as e:
            logger.error(f"Error eliminando archivo {archivo_id}: {e}")
            messages.error(request, f'Error al eliminar el archivo: {str(e)}')
            return render(request, 'core/archivos/delete.html', {'archivo': archivo})
    
    return render(request, 'core/archivos/delete.html', {'archivo': archivo})


@login_required
def archivo_preview(request, archivo_id):
    """Vista previa de un archivo (para imágenes y PDFs)"""
    archivo = get_object_or_404(ArchivoProyecto, id=archivo_id, activo=True)
    
    context = {
        'archivo': archivo,
    }
    
    return render(request, 'core/archivos/preview.html', context)
 









# ===== VISTAS DEL SISTEMA =====
@login_required
def sistema_view(request):
    """Vista principal del sistema"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    # Estadísticas del sistema
    total_usuarios = User.objects.count()
    total_logs = LogActividad.objects.count()
    logs_recientes = LogActividad.objects.all().order_by('-fecha_actividad')[:10]
    
    context = {
        'total_usuarios': total_usuarios,
        'total_logs': total_logs,
        'logs_recientes': logs_recientes,
    }
    
    return render(request, 'core/sistema/index.html', context)


@login_required
def sistema_configurar(request):
    """Configuración del sistema"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    config = ConfiguracionSistema.get_config()
    
    if request.method == 'POST':
        try:
            # Actualizar configuración general
            config.nombre_empresa = request.POST.get('nombre_empresa', config.nombre_empresa)
            config.moneda = request.POST.get('moneda', config.moneda)
            config.zona_horaria = request.POST.get('zona_horaria', config.zona_horaria)
            config.idioma = request.POST.get('idioma', config.idioma)
            
            # Configuración del sistema
            config.max_usuarios_simultaneos = int(request.POST.get('max_usuarios_simultaneos', config.max_usuarios_simultaneos))
            config.tiempo_sesion = int(request.POST.get('tiempo_sesion', config.tiempo_sesion))
            
            # Configuraciones avanzadas
            config.respaldo_automatico = 'respaldo_automatico' in request.POST
            config.notificaciones_email = 'notificaciones_email' in request.POST
            
            # Configuración de email
            config.email_host = request.POST.get('email_host', config.email_host)
            config.email_port = int(request.POST.get('email_port', config.email_port))
            config.email_username = request.POST.get('email_username', config.email_username)
            config.email_password = request.POST.get('email_password', config.email_password)
            config.email_use_tls = 'email_use_tls' in request.POST
            
            config.actualizado_por = request.user
            config.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Actualizar',
                modulo='Sistema',
                descripcion=f'Configuración del sistema actualizada por {request.user.username}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, '✅ Configuración actualizada exitosamente')
            return redirect('sistema_configurar')
            
        except Exception as e:
            logger.error(f'Error al actualizar configuración: {e}')
            messages.error(request, f'❌ Error al actualizar configuración: {str(e)}')
    
    context = {
        'config': config,
    }
    return render(request, 'core/sistema/configurar.html', context)


@login_required
def sistema_crear_respaldo(request):
    """Crear respaldo de la base de datos"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        from django.core.management import call_command
        from django.core.files.base import ContentFile
        import io
        import os
        from datetime import datetime
        
        # Crear respaldo
        output = io.StringIO()
        call_command('dumpdata', stdout=output, indent=2)
        output.seek(0)
        
        # Crear nombre de archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'respaldo_sistema_{timestamp}.json'
        
        # Crear directorio de respaldos si no existe
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'respaldos')
        os.makedirs(backup_dir, exist_ok=True)
        
        # Guardar archivo
        backup_path = os.path.join(backup_dir, filename)
        with open(backup_path, 'w', encoding='utf-8') as f:
            f.write(output.getvalue())
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Crear',
            modulo='Sistema',
            descripcion=f'Respaldo creado: {filename}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'✅ Respaldo creado exitosamente: {filename}')
        
    except Exception as e:
        logger.error(f'Error al crear respaldo: {e}')
        messages.error(request, f'❌ Error al crear respaldo: {str(e)}')
    
    return redirect('sistema_configurar')


@login_required
def sistema_limpiar_logs(request):
    """Limpiar logs antiguos"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        from datetime import timedelta
        
        # Eliminar logs más antiguos de 30 días
        fecha_limite = timezone.now() - timedelta(days=30)
        logs_eliminados = LogActividad.objects.filter(fecha__lt=fecha_limite).count()
        LogActividad.objects.filter(fecha__lt=fecha_limite).delete()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Limpiar',
            modulo='Sistema',
            descripcion=f'Logs antiguos eliminados: {logs_eliminados} registros',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'✅ Se eliminaron {logs_eliminados} logs antiguos')
        
    except Exception as e:
        logger.error(f'Error al limpiar logs: {e}')
        messages.error(request, f'❌ Error al limpiar logs: {str(e)}')
    
    return redirect('sistema_configurar')


@login_required
def sistema_exportar_config(request):
    """Exportar configuración del sistema"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        config = ConfiguracionSistema.get_config()
        
        # Crear respuesta JSON
        config_data = {
            'nombre_empresa': config.nombre_empresa,
            'moneda': config.moneda,
            'zona_horaria': config.zona_horaria,
            'idioma': config.idioma,
            'max_usuarios_simultaneos': config.max_usuarios_simultaneos,
            'tiempo_sesion': config.tiempo_sesion,
            'respaldo_automatico': config.respaldo_automatico,
            'notificaciones_email': config.notificaciones_email,
            'email_host': config.email_host,
            'email_port': config.email_port,
            'email_username': config.email_username,
            'email_use_tls': config.email_use_tls,
            'exportado_en': timezone.now().isoformat(),
            'exportado_por': request.user.username
        }
        
        response = HttpResponse(
            json.dumps(config_data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="configuracion_sistema_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json"'
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Exportar',
            modulo='Sistema',
            descripcion='Configuración del sistema exportada',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return response
        
    except Exception as e:
        logger.error(f'Error al exportar configuración: {e}')
        messages.error(request, f'❌ Error al exportar configuración: {str(e)}')
        return redirect('sistema_configurar')


@login_required
def sistema_ver_respaldos(request):
    """Ver lista de respaldos disponibles"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        import os
        from datetime import datetime
        
        # Directorio de respaldos
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'respaldos')
        
        # Crear directorio si no existe
        os.makedirs(backup_dir, exist_ok=True)
        
        # Obtener lista de archivos
        respaldos = []
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.json') and filename.startswith('respaldo_sistema_'):
                    file_path = os.path.join(backup_dir, filename)
                    file_stat = os.stat(file_path)
                    
                    # Extraer fecha del nombre del archivo
                    try:
                        date_str = filename.replace('respaldo_sistema_', '').replace('.json', '')
                        file_date = datetime.strptime(date_str, '%Y%m%d_%H%M%S')
                    except:
                        file_date = datetime.fromtimestamp(file_stat.st_mtime)
                    
                    respaldos.append({
                        'nombre': filename,
                        'fecha': file_date,
                        'tamaño': file_stat.st_size,
                        'ruta': file_path
                    })
            
            # Ordenar por fecha (más reciente primero)
            respaldos.sort(key=lambda x: x['fecha'], reverse=True)
        
        context = {
            'respaldos': respaldos,
            'backup_dir': backup_dir,
        }
        
        return render(request, 'core/sistema/respaldos.html', context)
        
    except Exception as e:
        logger.error(f'Error al listar respaldos: {e}')
        messages.error(request, f'❌ Error al listar respaldos: {str(e)}')
        return redirect('sistema_configurar')


@login_required
def sistema_restaurar_respaldo(request, filename):
    """Restaurar respaldo de la base de datos"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        from django.core.management import call_command
        import os
        from datetime import datetime
        
        # Verificar que el archivo existe
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'respaldos')
        backup_path = os.path.join(backup_dir, filename)
        
        if not os.path.exists(backup_path):
            messages.error(request, f'❌ Archivo de respaldo no encontrado: {filename}')
            return redirect('sistema_ver_respaldos')
        
        # Verificar que es un archivo de respaldo válido
        if not filename.endswith('.json') or not filename.startswith('respaldo_sistema_'):
            messages.error(request, '❌ Archivo de respaldo inválido')
            return redirect('sistema_ver_respaldos')
        
        # Crear respaldo de seguridad antes de restaurar
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safety_backup = f'respaldo_seguridad_{timestamp}.json'
        safety_path = os.path.join(backup_dir, safety_backup)
        
        # Crear respaldo de seguridad
        call_command('dumpdata', output=safety_path, indent=2)
        
        # Limpiar base de datos actual (solo datos, no estructura)
        from django.db import connection
        with connection.cursor() as cursor:
            # Obtener todas las tablas de la app core
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'core_%'")
            tables = [row[0] for row in cursor.fetchall()]
            
            # Deshabilitar foreign key checks temporalmente
            cursor.execute("PRAGMA foreign_keys = OFF")
            
            # Limpiar tablas (excepto migrations)
            for table in tables:
                if table != 'core_migrations':
                    cursor.execute(f"DELETE FROM {table}")
            
            # Rehabilitar foreign key checks
            cursor.execute("PRAGMA foreign_keys = ON")
        
        # Restaurar desde el respaldo
        call_command('loaddata', backup_path)
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Restaurar',
            modulo='Sistema',
            descripcion=f'Respaldo restaurado: {filename} (Respaldo de seguridad: {safety_backup})',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'✅ Respaldo restaurado exitosamente: {filename}')
        messages.info(request, f'ℹ️ Se creó un respaldo de seguridad: {safety_backup}')
        
    except Exception as e:
        logger.error(f'Error al restaurar respaldo: {e}')
        messages.error(request, f'❌ Error al restaurar respaldo: {str(e)}')
        messages.error(request, '⚠️ Si hay problemas, puedes restaurar desde el respaldo de seguridad')
    
    return redirect('sistema_ver_respaldos')


def offline_view(request):
    """Vista para modo offline"""
    return render(request, 'core/offline.html')


@login_required
def sistema_logs(request):
    """Ver logs del sistema"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    logs = LogActividad.objects.all().order_by('-fecha_actividad')
    
    # Paginación
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    
    return render(request, 'core/sistema/logs.html', context)


@login_required
def offline_page(request):
    """Página offline para PWA"""
    return render(request, 'offline.html')

def sistema_reset_app(request):
    """Reset completo de la aplicación - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            # Registrar la acción en logs
            LogActividad.objects.create(
                usuario=request.user,
                accion='RESET_APP',
                modulo='Sistema',
                descripcion=f'Usuario {request.user.username} inició RESET COMPLETO de la aplicación',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            # LIMPIEZA COMPLETA DE DATOS - En orden para evitar problemas de integridad referencial
            logger.info("Iniciando limpieza completa del sistema")
            
            try:
                # 1. Eliminar datos de facturación
                logger.info("Eliminando datos de facturación")
                Pago.objects.all().delete()
                Factura.objects.all().delete()
                logger.info("Facturación eliminada")
                
                # 2. Eliminar datos financieros
                logger.info("Eliminando datos financieros")
                Anticipo.objects.all().delete()
                Gasto.objects.all().delete()
                CategoriaGasto.objects.all().delete()
                GastoFijoMensual.objects.all().delete()
                logger.info("Datos financieros eliminados")
                
                # 3. Eliminar presupuestos (comentado - modelos no existen)
                # logger.info("Eliminando presupuestos")
                # PartidaPresupuesto.objects.all().delete()
                # Presupuesto.objects.all().delete()
                # logger.info("Presupuestos eliminados")
                
                # 4. Eliminar archivos
                logger.info("Eliminando archivos")
                ArchivoProyecto.objects.all().delete()
                logger.info("Archivos eliminados")
                
                # 5. Eliminar colaboradores
                logger.info("Eliminando colaboradores")
                Colaborador.objects.all().delete()
                logger.info("Colaboradores eliminados")
                
                # 6. Eliminar proyectos
                logger.info("Eliminando proyectos")
                Proyecto.objects.all().delete()
                logger.info("Proyectos eliminados")
                
                # 7. Eliminar clientes
                logger.info("Eliminando clientes")
                Cliente.objects.all().delete()
                logger.info("Clientes eliminados")
                
                # 8. Eliminar inventario
                logger.info("Eliminando inventario")
                AsignacionInventario.objects.all().delete()
                ItemInventario.objects.all().delete()
                CategoriaInventario.objects.all().delete()
                logger.info("Inventario eliminado")
                
                # 9. Eliminar notificaciones básicas
                logger.info("Eliminando notificaciones")
                NotificacionSistema.objects.all().delete()
                logger.info("Notificaciones eliminadas")
                
                # 10. Limpiar logs de actividad (mantener solo el log actual del reset)
                logger.info("Limpiando logs de actividad")
                LogActividad.objects.exclude(
                    accion__in=['RESET_APP', 'RESET_APP_SUCCESS', 'RESET_APP_ERROR']
                ).delete()
                logger.info("Logs limpiados")
                
                # 11. Limpiar perfiles de usuario (excepto el superusuario actual)
                logger.info("Limpiando perfiles de usuario")
                PerfilUsuario.objects.exclude(usuario=request.user).delete()
                logger.info("Perfiles limpiados")
                
                # 12. Limpiar usuarios (excepto superusuarios)
                logger.info("Limpiando usuarios")
                User.objects.exclude(is_superuser=True).delete()
                logger.info("Usuarios limpiados")
                
                # 13. Limpiar roles
                logger.info("Limpiando roles")
                Rol.objects.all().delete()
                logger.info("Roles eliminados")
                
                # 14. Limpiar caché del sistema
                logger.info("Limpiando caché")
                cache.clear()
                logger.info("Caché limpiado")
                
                # 15. Limpiar contadores de base de datos SQLite
                logger.info("Limpiando contadores de base de datos")
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute("DELETE FROM sqlite_sequence")
                logger.info("Contadores limpiados")
                
                logger.info("Sistema limpiado completamente")
                
            except Exception as e:
                logger.error(f"Error en paso específico del reset: {e}")
                raise e
            
            messages.success(request, '✅ RESET COMPLETO realizado exitosamente. Todos los datos han sido eliminados.')
            
            # Registrar éxito en logs
            LogActividad.objects.create(
                usuario=request.user,
                accion='RESET_APP_SUCCESS',
                modulo='Sistema',
                descripcion='Reset COMPLETO de aplicación realizado exitosamente - Todos los datos eliminados',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
        except Exception as e:
            logger.error(f"Error durante la limpieza del sistema: {e}")
            messages.error(request, f'❌ Error durante el reset: {str(e)}')
            
            # Registrar error en logs
            LogActividad.objects.create(
                usuario=request.user,
                accion='RESET_APP_ERROR',
                modulo='Sistema',
                descripcion=f'Error durante reset COMPLETO de aplicación: {str(e)}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
        
        return redirect('sistema')
    
    return render(request, 'core/sistema/reset_app.html')
 
# ===== VISTA DE RENTABILIDAD =====
def calcular_rentabilidad_proyecto(proyecto, fecha_inicio_dt=None, fecha_fin_dt=None):
    """Función auxiliar para calcular rentabilidad de un proyecto usando la misma lógica que el dashboard"""
    from decimal import Decimal
    
    # Construir filtros de fecha dinámicamente
    facturas_filter = {'proyecto': proyecto, 'estado': 'pagada'}
    anticipos_filter = {'proyecto': proyecto, 'monto_aplicado__gt': 0}
    anticipos_proyecto_filter = {'proyecto': proyecto, 'aplicado_al_proyecto': True}
    gastos_filter = {'proyecto': proyecto, 'aprobado': True}
    
    # Aplicar filtros de fecha solo si se proporcionan
    if fecha_inicio_dt and fecha_fin_dt:
        facturas_filter['fecha_emision__range'] = [fecha_inicio_dt, fecha_fin_dt]
        anticipos_filter['fecha_aplicacion__range'] = [fecha_inicio_dt, fecha_fin_dt]
        anticipos_proyecto_filter['fecha_aplicacion__range'] = [fecha_inicio_dt, fecha_fin_dt]
        gastos_filter['fecha_gasto__range'] = [fecha_inicio_dt, fecha_fin_dt]
    
    # Facturas pagadas del proyecto
    facturas_pagadas = Factura.objects.filter(**facturas_filter)
    total_facturas_pagadas = facturas_pagadas.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
    
    # Anticipos aplicados a facturas del proyecto
    anticipos_aplicados = Anticipo.objects.filter(**anticipos_filter)
    total_anticipos_aplicados = anticipos_aplicados.aggregate(total=Sum('monto_aplicado'))['total'] or Decimal('0.00')
    
    # Anticipos aplicados al proyecto
    anticipos_proyecto = Anticipo.objects.filter(**anticipos_proyecto_filter)
    total_anticipos_aplicados_proyecto = anticipos_proyecto.aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
    
    # Total de ingresos (misma lógica que dashboard)
    ingresos_proyecto = total_facturas_pagadas + total_anticipos_aplicados + total_anticipos_aplicados_proyecto
    
    # Gastos del proyecto
    gastos_proyecto_raw = Gasto.objects.filter(**gastos_filter).aggregate(total=Sum('monto'))['total'] or 0
    gastos_proyecto = Decimal(str(gastos_proyecto_raw))
    
    rentabilidad_proyecto = ingresos_proyecto - gastos_proyecto
    margen_proyecto = (rentabilidad_proyecto / ingresos_proyecto * 100) if ingresos_proyecto > 0 else Decimal('0.00')
    
    return {
        'ingresos': ingresos_proyecto,
        'gastos': gastos_proyecto,
        'rentabilidad': rentabilidad_proyecto,
        'margen': margen_proyecto
    }

@login_required
def rentabilidad_view(request):
    """Vista de rentabilidad y análisis financiero"""
    try:
        # Obtener parámetros de filtro
        periodo = request.GET.get('periodo', 'todos')  # todos, mes, trimestre, año
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Inicializar fechas como None (sin filtro)
        fecha_inicio_dt = None
        fecha_fin_dt = None
        
        # Solo aplicar filtros de fecha si NO es 'todos'
        if periodo != 'todos':
            hoy = timezone.now()
            if periodo == 'mes':
                fecha_inicio = fecha_inicio or (hoy - timedelta(days=30)).strftime('%Y-%m-%d')
                fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
            elif periodo == 'trimestre':
                fecha_inicio = fecha_inicio or (hoy - timedelta(days=90)).strftime('%Y-%m-%d')
                fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
            elif periodo == 'año':
                fecha_inicio = fecha_inicio or (hoy - timedelta(days=365)).strftime('%Y-%m-%d')
                fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
            
            # Convertir fechas a datetime solo si se especificaron
            if fecha_inicio and fecha_fin:
                fecha_inicio_dt = timezone.make_aware(datetime.strptime(fecha_inicio, '%Y-%m-%d'))
                fecha_fin_dt = timezone.make_aware(datetime.strptime(fecha_fin, '%Y-%m-%d'))
        else:
            # Cuando periodo es 'todos', forzar fechas a None
            fecha_inicio = None
            fecha_fin = None
            fecha_inicio_dt = None
            fecha_fin_dt = None
    
        # Calcular ingresos (usando la misma lógica que el dashboard)
        # Construir filtros de fecha dinámicamente
        facturas_filter = {'estado': 'pagada'}
        anticipos_filter = {'monto_aplicado__gt': 0}
        anticipos_proyecto_filter = {'aplicado_al_proyecto': True}
        
        # Aplicar filtros de fecha solo si se proporcionan
        if fecha_inicio_dt and fecha_fin_dt:
            facturas_filter['fecha_emision__range'] = [fecha_inicio_dt, fecha_fin_dt]
            anticipos_filter['fecha_aplicacion__range'] = [fecha_inicio_dt, fecha_fin_dt]
            anticipos_proyecto_filter['fecha_aplicacion__range'] = [fecha_inicio_dt, fecha_fin_dt]
        
        # Facturas pagadas
        facturas_pagadas = Factura.objects.filter(**facturas_filter)
        total_facturas_pagadas = facturas_pagadas.aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        # Anticipos aplicados a facturas
        anticipos_aplicados = Anticipo.objects.filter(**anticipos_filter)
        total_anticipos_aplicados = anticipos_aplicados.aggregate(total=Sum('monto_aplicado'))['total'] or Decimal('0.00')
        
        # Anticipos aplicados al proyecto
        anticipos_proyecto = Anticipo.objects.filter(**anticipos_proyecto_filter)
        total_anticipos_aplicados_proyecto = anticipos_proyecto.aggregate(total=Sum('monto_aplicado_proyecto'))['total'] or Decimal('0.00')
        
        # Total de ingresos
        ingresos = total_facturas_pagadas + total_anticipos_aplicados + total_anticipos_aplicados_proyecto
        
        # Calcular totales de facturación e ingresos
        facturado_filter = {}
        if fecha_inicio_dt and fecha_fin_dt:
            facturado_filter['fecha_emision__range'] = [fecha_inicio_dt, fecha_fin_dt]
        
        total_facturado = Factura.objects.filter(**facturado_filter).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        # Log información de rentabilidad
        periodo_info = f"{fecha_inicio} a {fecha_fin}" if fecha_inicio and fecha_fin else "Todos los datos"
        logger.debug(f"Período: {periodo_info}, Facturado: ${total_facturado}, Cobrado: ${ingresos}")
        
        # Calcular gastos
        gastos_filter = {'aprobado': True}
        if fecha_inicio_dt and fecha_fin_dt:
            gastos_filter['fecha_gasto__range'] = [fecha_inicio_dt, fecha_fin_dt]
        
        gastos_raw = Gasto.objects.filter(**gastos_filter).aggregate(total=Sum('monto'))['total'] or 0
        gastos = Decimal(str(gastos_raw))
        
        # Gastos fijos (por ahora 0)
        gastos_fijos = Decimal('0.00')
        
        # Calcular rentabilidad
        rentabilidad_bruta = ingresos - gastos
        rentabilidad_neta = rentabilidad_bruta - gastos_fijos
        
        # Margen de rentabilidad
        margen_rentabilidad = (rentabilidad_neta / ingresos * 100) if ingresos > 0 else Decimal('0.00')
        
        # Análisis por proyecto
        proyectos_rentabilidad = []
        proyectos = Proyecto.objects.filter(activo=True)
        
        for proyecto in proyectos:
            # Usar la función auxiliar para calcular rentabilidad
            rentabilidad_data = calcular_rentabilidad_proyecto(proyecto, fecha_inicio_dt, fecha_fin_dt)
            ingresos_proyecto = rentabilidad_data['ingresos']
            gastos_proyecto = rentabilidad_data['gastos']
            rentabilidad_proyecto = rentabilidad_data['rentabilidad']
            margen_proyecto = rentabilidad_data['margen']
            
            proyectos_rentabilidad.append({
                'proyecto': proyecto,
                'ingresos': ingresos_proyecto,
                'gastos': gastos_proyecto,
                'rentabilidad': rentabilidad_proyecto,
                'margen': margen_proyecto
            })
        
        # Ordenar por rentabilidad
        proyectos_rentabilidad.sort(key=lambda x: x['rentabilidad'], reverse=True)
        
        # Análisis por categoría de gasto
        gastos_categoria_filter = {'aprobado': True}
        if fecha_inicio_dt and fecha_fin_dt:
            gastos_categoria_filter['fecha_gasto__range'] = [fecha_inicio_dt, fecha_fin_dt]
        
        gastos_por_categoria = Gasto.objects.filter(**gastos_categoria_filter).values('categoria__nombre').annotate(
            total=Sum('monto'),
            cantidad=Count('id')
        ).order_by('-total')[:8]
        
        # Tendencias mensuales (últimos 12 meses)
        tendencias_mensuales = []
        for i in range(12):
            fecha = hoy - timedelta(days=30*i)
            mes = fecha.month
            año = fecha.year
            
            ingresos_mes = Factura.objects.filter(
                fecha_emision__month=mes,
                fecha_emision__year=año,
                monto_pagado__gt=0
            ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
            
            gastos_mes_raw = Gasto.objects.filter(
                fecha_gasto__month=mes,
                fecha_gasto__year=año,
                aprobado=True
            ).aggregate(total=Sum('monto'))['total'] or 0
            gastos_mes = Decimal(str(gastos_mes_raw))
            
            rentabilidad_mes = ingresos_mes - gastos_mes
            
            tendencias_mensuales.append({
                'mes': fecha.strftime('%b %Y'),
                'ingresos': ingresos_mes,
                'gastos': gastos_mes,
                'rentabilidad': rentabilidad_mes
            })
        
        # Ordenar tendencias cronológicamente
        tendencias_mensuales.reverse()
        
        # Crear datos JSON para JavaScript
        import json
        tendencias_json = json.dumps(tendencias_mensuales, default=str)
        
        # Calcular rentabilidad del mes actual para el dashboard
        mes_actual = hoy.month
        año_actual = hoy.year
        
        ingresos_mes_actual = Factura.objects.filter(
            fecha_emision__month=mes_actual,
            fecha_emision__year=año_actual,
            monto_pagado__gt=0
        ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
        
        gastos_mes_actual_raw = Gasto.objects.filter(
            fecha_gasto__month=mes_actual,
            fecha_gasto__year=año_actual,
            aprobado=True
        ).aggregate(total=Sum('monto'))['total'] or 0
        gastos_mes_actual = Decimal(str(gastos_mes_actual_raw))
        
        rentabilidad_mes_actual = ingresos_mes_actual - gastos_mes_actual
        margen_mes_actual = (rentabilidad_mes_actual / ingresos_mes_actual * 100) if ingresos_mes_actual > 0 else Decimal('0.00')
        
        # ===== CALCULAR TENDENCIAS VS MES ANTERIOR =====
        # Obtener mes anterior
        mes_anterior = mes_actual - 1 if mes_actual > 1 else 12
        año_anterior = año_actual if mes_actual > 1 else año_actual - 1
        
        # Ingresos mes anterior
        ingresos_mes_anterior = Factura.objects.filter(
            fecha_emision__month=mes_anterior,
            fecha_emision__year=año_anterior,
            monto_pagado__gt=0
        ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
        
        # Gastos mes anterior
        gastos_mes_anterior_raw = Gasto.objects.filter(
            fecha_gasto__month=mes_anterior,
            fecha_gasto__year=año_anterior,
            aprobado=True
        ).aggregate(total=Sum('monto'))['total'] or 0
        gastos_mes_anterior = Decimal(str(gastos_mes_anterior_raw))
        
        # Calcular tendencias (% de cambio)
        tendencia_ingresos = Decimal('0')
        tendencia_gastos = Decimal('0')
        tendencia_rentabilidad = Decimal('0')
        
        if ingresos_mes_anterior > 0:
            tendencia_ingresos = ((ingresos_mes_actual - ingresos_mes_anterior) / ingresos_mes_anterior * 100)
        
        if gastos_mes_anterior > 0:
            tendencia_gastos = ((gastos_mes_actual - gastos_mes_anterior) / gastos_mes_anterior * 100)
        
        rentabilidad_mes_anterior = ingresos_mes_anterior - gastos_mes_anterior
        if rentabilidad_mes_anterior != 0:
            tendencia_rentabilidad = ((rentabilidad_mes_actual - rentabilidad_mes_anterior) / abs(rentabilidad_mes_anterior) * 100)
        
        # Preparar datos para gráficos (formato JSON)
        import json
        meses_tendencia = [t['mes'] for t in tendencias_mensuales]
        ingresos_tendencia = [float(t['ingresos']) for t in tendencias_mensuales]
        gastos_tendencia = [float(t['gastos']) for t in tendencias_mensuales]
        
        categorias_gastos_labels = [cat['categoria__nombre'] if cat['categoria__nombre'] else 'Sin Categoría' for cat in gastos_por_categoria]
        categorias_gastos_data = [float(cat['total']) for cat in gastos_por_categoria]
        
        context = {
            'periodo': periodo,
            'periodo_seleccionado': periodo,  # Para el template
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'fecha_inicio_str': fecha_inicio,  # Para los inputs
            'fecha_fin_str': fecha_fin,  # Para los inputs
            'ingresos': ingresos,
            'gastos': gastos,
            'gastos_fijos': gastos_fijos,
            'rentabilidad_bruta': rentabilidad_bruta,
            'rentabilidad_neta': rentabilidad_neta,
            'margen_rentabilidad': margen_rentabilidad,
            'proyectos_rentabilidad': proyectos_rentabilidad,
            'gastos_por_categoria': gastos_por_categoria,
            'tendencias_mensuales': tendencias_mensuales,
            'tendencias_json': tendencias_json,
            # Datos para gráficos Chart.js
            'meses_tendencia': json.dumps(meses_tendencia),
            'ingresos_tendencia': json.dumps(ingresos_tendencia),
            'gastos_tendencia': json.dumps(gastos_tendencia),
            'categorias_gastos_chart_labels': json.dumps(categorias_gastos_labels),
            'categorias_gastos_chart_data': json.dumps(categorias_gastos_data),
            # Datos para el dashboard
            'ingresos_mes': ingresos_mes_actual,
            'gastos_mes': gastos_mes_actual,
            'rentabilidad_mes': margen_mes_actual,
            # Tendencias vs mes anterior
            'tendencia_ingresos': tendencia_ingresos,
            'tendencia_gastos': tendencia_gastos,
            'tendencia_rentabilidad': tendencia_rentabilidad,
        }
        
        return render(request, 'core/rentabilidad/index.html', context)
        
    except Exception as e:
        # En caso de error, devolver valores por defecto
        context = {
            'periodo': 'mes',
            'fecha_inicio': (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'),
            'fecha_fin': timezone.now().strftime('%Y-%m-%d'),
            'ingresos': Decimal('0.00'),
            'gastos': Decimal('0.00'),
            'gastos_fijos': Decimal('0.00'),
            'rentabilidad_bruta': Decimal('0.00'),
            'rentabilidad_neta': Decimal('0.00'),
            'margen_rentabilidad': Decimal('0.00'),
            'proyectos_rentabilidad': [],
            'gastos_por_categoria': [],
            'tendencias_mensuales': [],
            'tendencias_json': '[]',
            'ingresos_mes': Decimal('0.00'),
            'gastos_mes': Decimal('0.00'),
            'rentabilidad_mes': Decimal('0.00'),
            'error': str(e)
        }
        
        return render(request, 'core/rentabilidad/index.html', context)
 
# ===== VISTAS DE PRESUPUESTOS ELIMINADAS - YA NO SE USAN =====
# VISTA presupuestos_list ELIMINADA - YA NO SE USA

@login_required
def presupuesto_create(request):
    """Crear nuevo presupuesto - DESHABILITADO: Modelo Presupuesto no existe"""
    messages.error(request, 'Funcionalidad de presupuestos deshabilitada - modelo no disponible')
    return redirect('dashboard')
    # Código comentado - modelo Presupuesto no existe
    if False and request.method == 'POST':
        form = PresupuestoForm(request.POST)
        proyecto_id = request.POST.get('proyecto')
        
        # Validar que se seleccionó un proyecto
        if not proyecto_id:
            messages.error(request, 'Debes seleccionar un proyecto para crear el presupuesto')
            context = {
                'form': form,
                'proyectos': Proyecto.objects.filter(activo=True),
            }
            return render(request, 'core/presupuestos/create.html', context)
        
        # Verificar que el proyecto existe
        try:
            proyecto = Proyecto.objects.get(id=proyecto_id, activo=True)
        except Proyecto.DoesNotExist:
            messages.error(request, 'El proyecto seleccionado no existe o no está activo')
            context = {
                'form': form,
                'proyectos': Proyecto.objects.filter(activo=True),
            }
            return render(request, 'core/presupuestos/create.html', context)
        
        if form.is_valid():
            presupuesto = form.save(commit=False)
            presupuesto.proyecto = proyecto
            presupuesto.creado_por = request.user
            presupuesto.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear',
                modulo='Presupuestos',
                descripcion=f'Presupuesto creado: {presupuesto.nombre} para proyecto {presupuesto.proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Presupuesto creado exitosamente')
            return redirect('presupuesto_detail', presupuesto_id=presupuesto.id)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario')
    else:
        form = PresupuestoForm()
    
    context = {
        'form': form,
        'proyectos': Proyecto.objects.filter(activo=True),
    }
    
    return render(request, 'core/presupuestos/create.html', context)

@login_required
def presupuesto_detail(request, presupuesto_id):
    """Detalle de un presupuesto - DESHABILITADO: Modelo Presupuesto no existe"""
    messages.error(request, 'Funcionalidad de presupuestos deshabilitada - modelo no disponible')
    return redirect('dashboard')
    # Código comentado - modelo Presupuesto no existe
    if False:
        presupuesto = None  # get_object_or_404(Presupuesto, id=presupuesto_id, activo=True)
    partidas = presupuesto.partidas.all()
    variaciones = presupuesto.variaciones.all()
    
    # Obtener datos de variación
    variacion_data = presupuesto.obtener_variacion()
    
    # Agregar valores absolutos para el template
    if variacion_data:
        variacion_data['variacion_abs'] = abs(variacion_data['variacion'])
        variacion_data['porcentaje_variacion_abs'] = abs(variacion_data['porcentaje_variacion'])
    
    # Obtener gastos reales por categoría para comparar
    gastos_reales = Gasto.objects.filter(
        proyecto=presupuesto.proyecto,
        aprobado=True
    ).values('categoria__nombre').annotate(
        total=Sum('monto')
    ).order_by('-total')
    
    context = {
        'presupuesto': presupuesto,
        'partidas': partidas,
        'variaciones': variaciones,
        'variacion_data': variacion_data,
        'gastos_reales': gastos_reales,
    }
    
    return render(request, 'core/presupuestos/detail.html', context)

@login_required
def presupuesto_edit(request, presupuesto_id):
    """Editar presupuesto - DESHABILITADO: Modelo Presupuesto no existe"""
    messages.error(request, 'Funcionalidad de presupuestos deshabilitada - modelo no disponible')
    return redirect('dashboard')
    # Código comentado - modelo Presupuesto no existe
    if False:
        presupuesto = None  # get_object_or_404(Presupuesto, id=presupuesto_id, activo=True)
    
    if request.method == 'POST':
        form = PresupuestoForm(request.POST, instance=presupuesto)
        if form.is_valid():
            presupuesto = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar',
                modulo='Presupuestos',
                descripcion=f'Presupuesto editado: {presupuesto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Presupuesto actualizado exitosamente')
            return redirect('presupuesto_detail', presupuesto_id=presupuesto.id)
    else:
        form = PresupuestoForm(instance=presupuesto)
    
    context = {
        'form': form,
        'presupuesto': presupuesto,
    }
    
    return render(request, 'core/presupuestos/edit.html', context)

@login_required
def partida_create(request, presupuesto_id):
    """Crear nueva partida en un presupuesto - DESHABILITADO: Modelo Presupuesto no existe"""
    messages.error(request, 'Funcionalidad de presupuestos deshabilitada - modelo no disponible')
    return redirect('dashboard')
    # Código comentado - modelo Presupuesto no existe
    if False:
        presupuesto = None  # get_object_or_404(Presupuesto, id=presupuesto_id, activo=True)
    
    # Verificar que el presupuesto puede recibir partidas
    if presupuesto.estado in ['aprobado', 'rechazado', 'obsoleto']:
        messages.error(request, f'No se pueden agregar partidas a un presupuesto en estado "{presupuesto.get_estado_display()}"')
        return redirect('presupuesto_detail', presupuesto_id=presupuesto.id)
    
    if request.method == 'POST':
        form = PartidaPresupuestoForm(request.POST)
        if form.is_valid():
            partida = form.save(commit=False)
            partida.presupuesto = presupuesto
            partida.save()
            
            # Si es la primera partida y el presupuesto está en borrador, cambiar a en_revision
            if presupuesto.estado == 'borrador' and presupuesto.partidas.count() == 1:
                presupuesto.estado = 'en_revision'
                presupuesto.save()
                messages.info(request, 'El presupuesto ha sido marcado como "En Revisión" automáticamente')
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear Partida',
                modulo='Presupuestos',
                descripcion=f'Partida creada: {partida.descripcion} - ${partida.monto_estimado} en presupuesto {presupuesto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Partida "{partida.descripcion}" creada exitosamente por ${partida.monto_estimado}')
            return redirect('presupuesto_detail', presupuesto_id=presupuesto.id)
    else:
        form = PartidaPresupuestoForm()
    
    context = {
        'form': form,
        'presupuesto': presupuesto,
        'partidas_count': presupuesto.partidas.count(),
    }
    
    return render(request, 'core/presupuestos/partida_create.html', context)

@login_required
def presupuesto_aprobar(request, presupuesto_id):
    """Aprobar un presupuesto - DESHABILITADO: Modelo Presupuesto no existe"""
    messages.error(request, 'Funcionalidad de presupuestos deshabilitada - modelo no disponible')
    return redirect('dashboard')
    # Código comentado - modelo Presupuesto no existe
    if False:
        presupuesto = None  # get_object_or_404(Presupuesto, id=presupuesto_id, activo=True)
    
    # Verificar que el presupuesto puede ser aprobado
    if presupuesto.estado not in ['borrador', 'en_revision']:
        messages.error(request, f'No se puede aprobar un presupuesto en estado "{presupuesto.get_estado_display()}"')
        return redirect('presupuesto_detail', presupuesto_id=presupuesto.id)
    
    # Verificar que el presupuesto tiene partidas
    if presupuesto.partidas.count() == 0:
        messages.warning(request, 'El presupuesto no tiene partidas. Se recomienda agregar partidas antes de aprobar.')
    
    if request.method == 'POST':
        try:
            presupuesto.estado = 'aprobado'
            presupuesto.fecha_aprobacion = timezone.now()
            presupuesto.aprobado_por = request.user
            presupuesto.monto_aprobado = presupuesto.monto_total
            presupuesto.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Aprobar',
                modulo='Presupuestos',
                descripcion=f'Presupuesto aprobado: {presupuesto.nombre} - ${presupuesto.monto_total}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Presupuesto "{presupuesto.nombre}" aprobado exitosamente por ${presupuesto.monto_total}')
            return redirect('presupuesto_detail', presupuesto_id=presupuesto.id)
            
        except Exception as e:
            messages.error(request, f'Error al aprobar el presupuesto: {str(e)}')
            return redirect('presupuesto_detail', presupuesto_id=presupuesto.id)
    
    context = {
        'presupuesto': presupuesto,
        'partidas_count': presupuesto.partidas.count(),
        'monto_total': presupuesto.monto_total,
    }
    
    return render(request, 'core/presupuestos/aprobar.html', context)
 
# ==================== VISTAS DE NOTIFICACIONES ====================

@login_required
def notificaciones_list(request):
    """Lista de notificaciones del usuario"""
    notificaciones = NotificacionSistema.objects.filter(usuario=request.user)
    
    # Filtros
    tipo = request.GET.get('tipo')
    prioridad = request.GET.get('prioridad')
    leida = request.GET.get('leida')
    
    if tipo:
        notificaciones = notificaciones.filter(tipo=tipo)
    if prioridad:
        notificaciones = notificaciones.filter(prioridad=prioridad)
    if leida is not None:
        notificaciones = notificaciones.filter(leida=leida == 'true')
    
    # Estadísticas
    total_notificaciones = notificaciones.count()
    no_leidas = notificaciones.filter(leida=False).count()
    urgentes = notificaciones.filter(prioridad='urgente', leida=False).count()
    
    context = {
        'notificaciones': notificaciones,
        'total_notificaciones': total_notificaciones,
        'no_leidas': no_leidas,
        'urgentes': urgentes,
        'tipos': NotificacionSistema.TIPO_CHOICES,
        'prioridades': NotificacionSistema.PRIORIDAD_CHOICES,
    }
    
    return render(request, 'core/notificaciones/list.html', context)


@login_required
def notificacion_marcar_leida(request, notificacion_id):
    """Marca una notificación como leída"""
    if request.method == 'POST':
        success = NotificacionService.marcar_como_leida(notificacion_id, request.user)
        if success:
            messages.success(request, 'Notificación marcada como leída')
        else:
            messages.error(request, 'Error al marcar la notificación')
    
    return redirect('notificaciones_list')


@login_required
def notificacion_marcar_todas_leidas(request):
    """Marca todas las notificaciones como leídas"""
    if request.method == 'POST':
        notificaciones = NotificacionSistema.objects.filter(
            usuario=request.user,
            leida=False
        )
        notificaciones.update(leida=True, fecha_lectura=timezone.now())
        messages.success(request, 'Todas las notificaciones han sido marcadas como leídas')
    
    return redirect('notificaciones_list')


@login_required
def notificaciones_configurar(request):
    """Configuración de notificaciones del usuario"""
    config, created = ConfiguracionNotificaciones.objects.get_or_create(usuario=request.user)
    
    if request.method == 'POST':
        # Actualizar configuración
        config.facturas_vencidas = request.POST.get('facturas_vencidas') == 'on'
        config.pagos_pendientes = request.POST.get('pagos_pendientes') == 'on'
        config.gastos_aprobacion = request.POST.get('gastos_aprobacion') == 'on'
        config.cambios_proyecto = request.POST.get('cambios_proyecto') == 'on'
        config.anticipos_disponibles = request.POST.get('anticipos_disponibles') == 'on'
        config.presupuestos_revision = request.POST.get('presupuestos_revision') == 'on'
        config.archivos_subidos = request.POST.get('archivos_subidos') == 'on'
        config.notificaciones_sistema = request.POST.get('notificaciones_sistema') == 'on'
        
        config.frecuencia_email = request.POST.get('frecuencia_email', 'inmediato')
        config.horario_inicio = request.POST.get('horario_inicio', '08:00:00')
        config.horario_fin = request.POST.get('horario_fin', '18:00:00')
        
        config.save()
        messages.success(request, 'Configuración de notificaciones actualizada')
        return redirect('notificaciones_configurar')
    
    context = {
        'config': config,
    }
    
    return render(request, 'core/notificaciones/configurar.html', context)


@login_required
def notificaciones_historial(request):
    """Historial de notificaciones enviadas"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    historial = HistorialNotificaciones.objects.all()
    
    # Filtros
    usuario_id = request.GET.get('usuario')
    tipo = request.GET.get('tipo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if usuario_id:
        historial = historial.filter(usuario_id=usuario_id)
    if tipo:
        historial = historial.filter(tipo=tipo)
    if fecha_inicio:
        historial = historial.filter(fecha_envio__gte=fecha_inicio)
    if fecha_fin:
        historial = historial.filter(fecha_envio__lte=fecha_fin)
    
    # Paginación
    paginator = Paginator(historial, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'usuarios': User.objects.filter(is_active=True),
        'tipos': NotificacionSistema.TIPO_CHOICES,
    }
    
    return render(request, 'core/notificaciones/historial.html', context)


# ==================== API PARA NOTIFICACIONES EN TIEMPO REAL ====================

@login_required
def api_notificaciones_no_leidas(request):
    """API para obtener notificaciones no leídas (para AJAX)"""
    notificaciones = NotificacionService.obtener_notificaciones_no_leidas(request.user)
    
    data = []
    for notif in notificaciones:
        data.append({
            'id': notif.id,
            'tipo': notif.tipo,
            'titulo': notif.titulo,
            'mensaje': notif.mensaje,
            'prioridad': notif.prioridad,
            'fecha_creacion': notif.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'icono': notif.get_icono(),
            'color_clase': notif.get_color_clase(),
        })
    
    return JsonResponse({
        'notificaciones': data,
        'total': len(data)
    })


@login_required
def api_marcar_leida(request, notificacion_id):
    """API para marcar notificación como leída (para AJAX)"""
    success = NotificacionService.marcar_como_leida(notificacion_id, request.user)
    
    return JsonResponse({
        'success': success
    })


# ==================== VISTAS DE ADMINISTRACIÓN DE NOTIFICACIONES ====================

@login_required
def admin_notificaciones_sistema(request):
    """Administración de notificaciones del sistema (solo superusuarios)"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        titulo = request.POST.get('titulo')
        mensaje = request.POST.get('mensaje')
        prioridad = request.POST.get('prioridad')
        usuarios_ids = request.POST.getlist('usuarios')
        
        if tipo and titulo and mensaje and usuarios_ids:
            usuarios = User.objects.filter(id__in=usuarios_ids)
            
            for usuario in usuarios:
                NotificacionService.crear_notificacion(
                    usuario=usuario,
                    tipo=tipo,
                    titulo=titulo,
                    mensaje=mensaje,
                    prioridad=prioridad
                )
            
            messages.success(request, f'Notificación enviada a {len(usuarios)} usuarios')
            return redirect('admin_notificaciones_sistema')
        else:
            messages.error(request, 'Todos los campos son requeridos')
    
    context = {
        'tipos': NotificacionSistema.TIPO_CHOICES,
        'prioridades': NotificacionSistema.PRIORIDAD_CHOICES,
        'usuarios': User.objects.filter(is_active=True),
    }
    
    return render(request, 'core/notificaciones/admin_sistema.html', context)


@login_required
def admin_ejecutar_verificaciones(request):
    """Ejecuta las verificaciones automáticas del sistema"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            # TODO: Implementar SistemaNotificacionesAutomaticas
            # SistemaNotificacionesAutomaticas.ejecutar_verificaciones_diarias()
            messages.success(request, 'Verificaciones automáticas ejecutadas correctamente')
        except Exception as e:
            messages.error(request, f'Error ejecutando verificaciones: {str(e)}')
    
    return redirect('admin_notificaciones_sistema')

@login_required
def test_notification_email(request):
    """
    Vista de prueba para enviar una notificación por email
    """
    if request.method == 'POST':
        try:
            # Crear una notificación de prueba
            notificacion = NotificacionSistema.objects.create(
                usuario=request.user,
                tipo='sistema',
                titulo='🧪 Notificación de Prueba',
                mensaje='Esta es una notificación de prueba para verificar el sistema de emails. Si recibes este email, significa que el sistema está funcionando correctamente.',
                prioridad='normal'
            )
            
            # Enviar email
            resultado = NotificacionService.enviar_email_notificacion(notificacion)
            
            if resultado:
                # Enviar notificación push también
                NotificacionService.enviar_notificacion_push(notificacion)
                
                context = {
                    'mensaje': '¡Notificación enviada exitosamente! Revisa la terminal del servidor para ver el email.',
                    'tipo_mensaje': 'success',
                    'titulo_mensaje': 'Éxito',
                    'icono': 'check-circle',
                    'debug': settings.DEBUG
                }
            else:
                context = {
                    'mensaje': 'Error al enviar la notificación por email.',
                    'tipo_mensaje': 'danger',
                    'titulo_mensaje': 'Error',
                    'icono': 'exclamation-triangle',
                    'debug': settings.DEBUG
                }
        except Exception as e:
            context = {
                'mensaje': f'Error: {str(e)}',
                'tipo_mensaje': 'danger',
                'titulo_mensaje': 'Error',
                'icono': 'exclamation-triangle',
                'debug': settings.DEBUG
            }
    else:
        context = {
            'debug': settings.DEBUG
        }
    
    return render(request, 'core/test_notification.html', context)


@login_required
def push_notifications_setup(request):
    """
    Vista para configurar notificaciones push
    """
    if request.method == 'POST':
        try:
            # Aquí se procesaría la suscripción push
            # Por ahora solo simulamos el éxito
            context = {
                'mensaje': 'Notificaciones push configuradas exitosamente.',
                'tipo_mensaje': 'success',
                'titulo_mensaje': 'Configurado',
                'icono': 'check-circle'
            }
        except Exception as e:
            context = {
                'mensaje': f'Error al configurar: {str(e)}',
                'tipo_mensaje': 'danger',
                'titulo_mensaje': 'Error',
                'icono': 'exclamation-triangle'
            }
    else:
        context = {}
    
    return render(request, 'core/push_notifications_setup.html', context)


@login_required
def api_push_subscription(request):
    """
    API para manejar suscripciones push
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            subscription_info = data.get('subscription')
            
            # Guardar la suscripción en la base de datos
            # Por ahora solo simulamos el éxito
            return JsonResponse({'status': 'success', 'message': 'Suscripción guardada'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@login_required
def perfil(request):
    """
    Vista del perfil del usuario
    """
    context = {
        'usuario': request.user,
    }
    return render(request, 'core/perfil.html', context)


@login_required
def rentabilidad_exportar_pdf(request):
    """Exportar reporte de rentabilidad a PDF"""
    try:
        # Obtener los mismos datos que la vista principal
        periodo = request.GET.get('periodo', 'mes')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Fechas por defecto
        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = fecha_inicio or (hoy - timedelta(days=30)).strftime('%Y-%m-%d')
            fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
        elif periodo == 'trimestre':
            fecha_inicio = fecha_inicio or (hoy - timedelta(days=90)).strftime('%Y-%m-%d')
            fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
        elif periodo == 'año':
            fecha_inicio = fecha_inicio or (hoy - timedelta(days=365)).strftime('%Y-%m-%d')
            fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
        
        # Convertir fechas a datetime
        fecha_inicio_dt = timezone.make_aware(datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        fecha_fin_dt = timezone.make_aware(datetime.strptime(fecha_fin, '%Y-%m-%d'))
        
        # Calcular datos de rentabilidad (mismo código que la vista principal)
        ingresos = Factura.objects.filter(
            fecha_emision__range=[fecha_inicio_dt, fecha_fin_dt],
            estado='pagada'
        ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        gastos_raw = Gasto.objects.filter(
            fecha_gasto__range=[fecha_inicio_dt, fecha_fin_dt],
            aprobado=True
        ).aggregate(total=Sum('monto'))['total'] or 0
        gastos = Decimal(str(gastos_raw))
        
        rentabilidad_bruta = ingresos - gastos
        margen_rentabilidad = (rentabilidad_bruta / ingresos * 100) if ingresos > 0 else Decimal('0.00')
        
        # Crear respuesta HTTP con contenido PDF
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        story.append(Paragraph("Reporte de Rentabilidad", title_style))
        story.append(Spacer(1, 12))
        
        # Información del período
        periodo_text = f"Período: {fecha_inicio} a {fecha_fin}"
        story.append(Paragraph(periodo_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Resumen financiero
        story.append(Paragraph("Resumen Financiero", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        data = [
            ['Concepto', 'Monto (Q)'],
            ['Ingresos', f"{ingresos:,.2f}"],
            ['Gastos', f"{gastos:,.2f}"],
            ['Rentabilidad Bruta', f"{rentabilidad_bruta:,.2f}"],
            ['Margen de Rentabilidad', f"{margen_rentabilidad:.2f}%"]
        ]
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        
        # Crear respuesta
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_rentabilidad_{fecha_inicio}_a_{fecha_fin}.pdf"'
        
        return response
        
    except Exception as e:
        logger.error(f'Error al exportar PDF de rentabilidad: {e}')
        messages.error(request, 'Error al generar el reporte PDF')
        return redirect('rentabilidad')


@login_required
def rentabilidad_exportar_excel(request):
    """Exportar reporte de rentabilidad a Excel"""
    try:
        # Obtener los mismos datos que la vista principal
        periodo = request.GET.get('periodo', 'mes')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Fechas por defecto
        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = fecha_inicio or (hoy - timedelta(days=30)).strftime('%Y-%m-%d')
            fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
        elif periodo == 'trimestre':
            fecha_inicio = fecha_inicio or (hoy - timedelta(days=90)).strftime('%Y-%m-%d')
            fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
        elif periodo == 'año':
            fecha_inicio = fecha_inicio or (hoy - timedelta(days=365)).strftime('%Y-%m-%d')
            fecha_fin = fecha_fin or hoy.strftime('%Y-%m-%d')
        
        # Convertir fechas a datetime
        fecha_inicio_dt = timezone.make_aware(datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        fecha_fin_dt = timezone.make_aware(datetime.strptime(fecha_fin, '%Y-%m-%d'))
        
        # Calcular datos de rentabilidad
        ingresos = Factura.objects.filter(
            fecha_emision__range=[fecha_inicio_dt, fecha_fin_dt],
            estado='pagada'
        ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        gastos_raw = Gasto.objects.filter(
            fecha_gasto__range=[fecha_inicio_dt, fecha_fin_dt],
            aprobado=True
        ).aggregate(total=Sum('monto'))['total'] or 0
        gastos = Decimal(str(gastos_raw))
        
        rentabilidad_bruta = ingresos - gastos
        margen_rentabilidad = (rentabilidad_bruta / ingresos * 100) if ingresos > 0 else Decimal('0.00')
        
        # Crear archivo Excel
        from django.http import HttpResponse
        from io import BytesIO
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messages.error(request, 'El módulo openpyxl no está instalado. Instálalo con: pip install openpyxl')
            return redirect('rentabilidad')
        
        buffer = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reporte de Rentabilidad"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Título
        worksheet.merge_cells('A1:E1')
        worksheet['A1'] = "Reporte de Rentabilidad"
        worksheet['A1'].font = Font(bold=True, size=16)
        worksheet['A1'].alignment = center_alignment
        
        # Período
        worksheet['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        worksheet.merge_cells('A2:E2')
        worksheet['A2'].alignment = center_alignment
        
        # Resumen financiero
        worksheet['A4'] = "Resumen Financiero"
        worksheet['A4'].font = Font(bold=True, size=14)
        
        # Encabezados
        headers = ['Concepto', 'Monto (Q)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Datos
        data = [
            ['Ingresos', float(ingresos)],
            ['Gastos', float(gastos)],
            ['Rentabilidad Bruta', float(rentabilidad_bruta)],
            ['Margen de Rentabilidad (%)', float(margen_rentabilidad)]
        ]
        
        for row, (concepto, monto) in enumerate(data, 6):
            worksheet.cell(row=row, column=1, value=concepto)
            worksheet.cell(row=row, column=2, value=monto)
        
        # Ajustar ancho de columnas
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 15
        
        # Guardar archivo
        workbook.save(buffer)
        buffer.seek(0)
        
        # Crear respuesta
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_rentabilidad_{fecha_inicio}_a_{fecha_fin}.xlsx"'
        
        return response
        
    except Exception as e:
        logger.error(f'Error al exportar Excel de rentabilidad: {e}')
        messages.error(request, 'Error al generar el reporte Excel')
        return redirect('rentabilidad')


# ===== VISTAS DEL MÓDULO DE INVENTARIO =====

@login_required
def inventario_dashboard(request):
    """Dashboard del módulo de inventario"""
    try:
        # Estadísticas generales
        total_items = ItemInventario.objects.count()
        total_categorias = CategoriaInventario.objects.count()
        total_asignaciones = AsignacionInventario.objects.count()
        items_bajo_stock = ItemInventario.objects.filter(stock_actual__lte=F('stock_minimo')).count()
        
        # Items recientes (últimos 5)
        items_recientes = ItemInventario.objects.select_related('categoria').order_by('-id')[:5]
        
        # Categorías con más items
        categorias_con_items = CategoriaInventario.objects.annotate(
            items_count=Count('iteminventario')
        ).filter(items_count__gt=0).order_by('-items_count')[:5]
        
        context = {
            'total_items': total_items,
            'total_categorias': total_categorias,
            'total_asignaciones': total_asignaciones,
            'items_bajo_stock': items_bajo_stock,
            'items_recientes': items_recientes,
            'categorias_con_items': categorias_con_items,
        }
        
        return render(request, 'core/inventario/dashboard.html', context)
        
    except Exception as e:
        logger.error(f'Error en inventario_dashboard: {e}')
        messages.error(request, 'Error al cargar el dashboard de inventario')
        return redirect('item_list')

# Vistas para Categorías
@login_required
def categoria_list(request):
    """Lista de categorías de inventario"""
    categorias = CategoriaInventario.objects.all().order_by('nombre')
    return render(request, 'core/inventario/categoria/list.html', {'categorias': categorias})

@login_required
def categoria_create(request):
    """Crear nueva categoría"""
    if request.method == 'POST':
        form = CategoriaInventarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente')
            return redirect('categoria_list')
    else:
        form = CategoriaInventarioForm()
    
    return render(request, 'core/inventario/categoria/create.html', {'form': form})

@login_required
def categoria_detail(request, pk):
    """Detalle de categoría"""
    categoria = get_object_or_404(CategoriaInventario, pk=pk)
    items = ItemInventario.objects.filter(categoria=categoria)
    return render(request, 'core/inventario/categoria/detail.html', {
        'categoria': categoria, 'items': items
    })

@login_required
def categoria_edit(request, pk):
    """Editar categoría"""
    categoria = get_object_or_404(CategoriaInventario, pk=pk)
    if request.method == 'POST':
        form = CategoriaInventarioForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente')
            return redirect('categoria_list')
    else:
        form = CategoriaInventarioForm(instance=categoria)
    
    return render(request, 'core/inventario/categoria/edit.html', {
        'form': form, 'categoria': categoria
    })

@login_required
def categoria_delete(request, pk):
    """Eliminar categoría"""
    categoria = get_object_or_404(CategoriaInventario, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente')
        return redirect('categoria_list')
    
    return render(request, 'core/inventario/categoria/delete.html', {'categoria': categoria})

# Vistas para Items
@login_required
def item_list(request):
    """Lista de items del inventario"""
    from django.core.paginator import Paginator
    
    items = ItemInventario.objects.select_related('categoria').all()
    
    # Filtros
    search = request.GET.get('search', '')
    categoria_filter = request.GET.get('categoria', '')
    stock_filter = request.GET.get('stock', '')
    
    # Aplicar filtros
    if search:
        items = items.filter(
            Q(nombre__icontains=search) | 
            Q(codigo__icontains=search) | 
            Q(descripcion__icontains=search)
        )
    
    if categoria_filter:
        items = items.filter(categoria_id=categoria_filter)
    
    if stock_filter == 'bajo':
        items = items.filter(stock_actual__lte=F('stock_minimo'))
    elif stock_filter == 'ok':
        items = items.filter(stock_actual__gt=F('stock_minimo'))
    
    # Ordenamiento
    items = items.order_by('nombre')
    
    # Paginación
    paginator = Paginator(items, 10)  # 10 items por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener categorías para el filtro
    categorias = CategoriaInventario.objects.all().order_by('nombre')
    
    context = {
        'items': page_obj,
        'page_obj': page_obj,
        'categorias': categorias,
        'search': search,
        'categoria_filter': categoria_filter,
        'stock_filter': stock_filter,
    }
    
    return render(request, 'core/inventario/items/list.html', context)

@login_required
def item_create(request):
    """Crear nuevo item"""
    if request.method == 'POST':
        form = ItemInventarioForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.stock_disponible = item.stock_actual
            item.save()
            messages.success(request, '✅ Item creado exitosamente')
            return redirect('item_list')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario')
    else:
        form = ItemInventarioForm()
    
    context = {
        'form': form,
        'form_title': 'Crear Nuevo Item',
        'form_description': 'Completa la información para agregar un nuevo item al inventario',
        'form_icon': 'plus',
        'submit_text': 'Crear Item'
    }
    
    return render(request, 'core/inventario/items/form.html', context)

@login_required
def item_detail(request, pk):
    """Detalle de item"""
    item = get_object_or_404(ItemInventario, pk=pk)
    asignaciones = AsignacionInventario.objects.filter(item=item).select_related(
        'proyecto', 'asignado_por'
    ).order_by('-fecha_asignacion')
    
    context = {
        'item': item,
        'asignaciones': asignaciones
    }
    
    return render(request, 'core/inventario/item/detail.html', context)

@login_required
def item_edit(request, pk):
    """Editar item"""
    item = get_object_or_404(ItemInventario, pk=pk)
    if request.method == 'POST':
        form = ItemInventarioForm(request.POST, instance=item)
        if form.is_valid():
            # Actualizar stock disponible si cambió el stock actual
            old_stock = item.stock_actual
            item = form.save(commit=False)
            if item.stock_actual != old_stock:
                diferencia = item.stock_actual - old_stock
                item.stock_disponible += diferencia
            item.save()
            messages.success(request, '✅ Item actualizado exitosamente')
            return redirect('item_list')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario')
    else:
        form = ItemInventarioForm(instance=item)
    
    context = {
        'form': form,
        'item': item,
        'form_title': f'Editar Item: {item.nombre}',
        'form_description': 'Modifica la información del item seleccionado',
        'form_icon': 'edit',
        'submit_text': 'Actualizar Item'
    }
    
    return render(request, 'core/inventario/items/form.html', context)

@login_required
def item_delete(request, pk):
    """Eliminar item"""
    item = get_object_or_404(ItemInventario, pk=pk)
    
    if request.method == 'POST':
        confirm = request.POST.get('confirm')
        if confirm == 'true':
            # Verificar si hay asignaciones activas
            asignaciones_activas = AsignacionInventario.objects.filter(
                item=item, 
                fecha_devolucion__isnull=True
            ).count()
            
            if asignaciones_activas > 0:
                messages.error(request, f'❌ No se puede eliminar el item porque tiene {asignaciones_activas} asignación(es) activa(s). Debe devolver todas las asignaciones primero.')
                return redirect('item_detail', pk=pk)
            
            # Eliminar el item
            item.delete()
            messages.success(request, '✅ Item eliminado exitosamente')
            return redirect('item_list')
        else:
            messages.error(request, '❌ Confirmación requerida para eliminar el item')
    
    # Contar asignaciones para mostrar en el template
    asignaciones_count = AsignacionInventario.objects.filter(item=item).count()
    
    return render(request, 'core/inventario/items/delete.html', {
        'item': item,
        'asignaciones_count': asignaciones_count
    })

# Vistas para Asignaciones
@login_required
def asignacion_list(request):
    """Lista de asignaciones de inventario"""
    asignaciones = AsignacionInventario.objects.select_related(
        'item', 'proyecto', 'asignado_por'
    ).all().order_by('-fecha_asignacion')
    
    # Filtros
    estado = request.GET.get('estado')
    if estado:
        asignaciones = asignaciones.filter(estado=estado)
    
    proyecto_id = request.GET.get('proyecto')
    if proyecto_id:
        asignaciones = asignaciones.filter(proyecto_id=proyecto_id)
    
    # Búsqueda
    query = request.GET.get('q')
    if query:
        asignaciones = asignaciones.filter(
            Q(item__nombre__icontains=query) | 
            Q(proyecto__nombre__icontains=query) |
            Q(notas__icontains=query)
        )
    
    proyectos = Proyecto.objects.all()
    
    context = {
        'asignaciones': asignaciones,
        'proyectos': proyectos,
        'estado_seleccionado': estado,
        'proyecto_seleccionado': proyecto_id,
        'query': query
    }
    
    return render(request, 'core/inventario/asignacion/list.html', context)

@login_required
def asignacion_create(request):
    """Crear nueva asignación"""
    if request.method == 'POST':
        form = AsignacionInventarioForm(request.POST)
        if form.is_valid():
            asignacion = form.save(commit=False)
            asignacion.asignado_por = request.user
            
            # Verificar stock disponible
            if asignacion.item.stock_disponible < asignacion.cantidad_asignada:
                messages.error(request, f'Stock insuficiente. Disponible: {asignacion.item.stock_disponible}')
                return render(request, 'core/inventario/asignacion/create.html', {'form': form})
            
            asignacion.save()
            messages.success(request, 'Asignación creada exitosamente')
            return redirect('asignacion_list')
    else:
        form = AsignacionInventarioForm()
    
    return render(request, 'core/inventario/asignacion/create.html', {'form': form})

@login_required
def asignacion_detail(request, pk):
    """Detalle de asignación"""
    asignacion = get_object_or_404(AsignacionInventario, pk=pk)
    return render(request, 'core/inventario/asignacion/detail.html', {'asignacion': asignacion})


# ============================================
# GESTIÓN DE USUARIOS Y ROLES
# ============================================

@login_required
def usuarios_lista(request):
    """Lista de usuarios del sistema - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    usuarios = User.objects.all().order_by('-date_joined')
    
    context = {
        'usuarios': usuarios,
        'total_usuarios': usuarios.count(),
    }
    
    return render(request, 'core/usuarios_lista.html', context)


@login_required
def usuario_crear(request):
    """Crear nuevo usuario - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            # Datos básicos del usuario
            username = request.POST.get('username')
            email = request.POST.get('email')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            password = request.POST.get('password')
            rol_id = request.POST.get('rol')
            
            # Validaciones básicas
            if not username or not password or not rol_id:
                messages.error(request, 'Todos los campos marcados son obligatorios')
                return redirect('usuario_crear')
            
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Ya existe un usuario con ese nombre')
                return redirect('usuario_crear')
            
            # Crear usuario
            usuario = User.objects.create_user(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                password=password
            )
            
            # Crear perfil de usuario
            rol = Rol.objects.get(id=rol_id)
            PerfilUsuario.objects.create(
                usuario=usuario,
                rol=rol,
                telefono=request.POST.get('telefono', ''),
                direccion=request.POST.get('direccion', '')
            )
            
            messages.success(request, f'Usuario {username} creado exitosamente con rol {rol.nombre}')
            return redirect('usuarios_lista')
            
        except Exception as e:
            messages.error(request, f'Error al crear usuario: {str(e)}')
            return redirect('usuario_crear')
    
    # GET - Mostrar formulario
    roles = Rol.objects.all()
    
    context = {
        'roles': roles,
    }
    
    return render(request, 'core/usuario_crear.html', context)


@login_required
def usuario_editar(request, usuario_id):
    """Editar usuario existente - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        usuario = User.objects.get(id=usuario_id)
        perfil, created = PerfilUsuario.objects.get_or_create(usuario=usuario)
        
        if request.method == 'POST':
            # Actualizar datos básicos
            usuario.first_name = request.POST.get('first_name', '')
            usuario.last_name = request.POST.get('last_name', '')
            usuario.email = request.POST.get('email', '')
            usuario.is_active = request.POST.get('is_active') == 'on'
            usuario.save()
            
            # Actualizar perfil
            rol_id = request.POST.get('rol')
            if rol_id:
                perfil.rol = Rol.objects.get(id=rol_id)
            perfil.telefono = request.POST.get('telefono', '')
            perfil.direccion = request.POST.get('direccion', '')
            perfil.activo = request.POST.get('activo') == 'on'
            perfil.save()
            
            # Cambiar contraseña si se proporciona
            nueva_password = request.POST.get('password')
            if nueva_password:
                usuario.set_password(nueva_password)
                usuario.save()
            
            messages.success(request, f'Usuario {usuario.username} actualizado exitosamente')
            return redirect('usuarios_lista')
        
        # GET - Mostrar formulario
        roles = Rol.objects.all()
        
        context = {
            'usuario_obj': usuario,
            'perfil': perfil,
            'roles': roles,
        }
        
        return render(request, 'core/usuario_editar.html', context)
        
    except User.DoesNotExist:
        messages.error(request, 'Usuario no encontrado')
        return redirect('usuarios_lista')
    except Exception as e:
        messages.error(request, f'Error al editar usuario: {str(e)}')
        return redirect('usuarios_lista')


@login_required
def roles_lista(request):
    """Lista de roles del sistema - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    roles = Rol.objects.all().order_by('nombre')
    
    # Obtener permisos por rol
    roles_con_permisos = []
    for rol in roles:
        permisos = RolPermiso.objects.filter(rol=rol, activo=True).count()
        usuarios_count = PerfilUsuario.objects.filter(rol=rol).count()
        roles_con_permisos.append({
            'rol': rol,
            'permisos_count': permisos,
            'usuarios_count': usuarios_count
        })
    
    context = {
        'roles_con_permisos': roles_con_permisos,
        'total_roles': roles.count(),
    }
    
    return render(request, 'core/roles_lista.html', context)


@login_required
def rol_crear(request):
    """Crear nuevo rol - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        
        if nombre:
            # Verificar que el nombre no exista
            if Rol.objects.filter(nombre=nombre).exists():
                messages.error(request, 'Ya existe un rol con ese nombre')
            else:
                rol = Rol.objects.create(
                    nombre=nombre,
                    descripcion=descripcion
                )
                messages.success(request, f'Rol "{rol.nombre}" creado exitosamente')
                return redirect('roles_lista')
        else:
            messages.error(request, 'El nombre del rol es obligatorio')
    
    return render(request, 'core/rol_crear.html')


@login_required
def rol_editar(request, rol_id):
    """Editar rol - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        rol = Rol.objects.get(id=rol_id)
        
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            
            if nombre:
                # Verificar que el nombre no exista en otro rol
                if Rol.objects.filter(nombre=nombre).exclude(id=rol_id).exists():
                    messages.error(request, 'Ya existe un rol con ese nombre')
                else:
                    rol.nombre = nombre
                    rol.descripcion = descripcion
                    rol.save()
                    messages.success(request, f'Rol "{rol.nombre}" actualizado exitosamente')
                    return redirect('roles_lista')
            else:
                messages.error(request, 'El nombre del rol es obligatorio')
        
        context = {'rol': rol}
        return render(request, 'core/rol_editar.html', context)
        
    except Rol.DoesNotExist:
        messages.error(request, 'Rol no encontrado')
        return redirect('roles_lista')


@login_required
def rol_eliminar(request, rol_id):
    """Eliminar rol - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        rol = Rol.objects.get(id=rol_id)
        
        # Verificar si hay usuarios usando este rol
        usuarios_con_rol = PerfilUsuario.objects.filter(rol=rol).count()
        
        if request.method == 'POST':
            if usuarios_con_rol > 0:
                messages.error(request, f'No se puede eliminar el rol "{rol.nombre}" porque tiene {usuarios_con_rol} usuario(s) asignado(s)')
            else:
                # Eliminar permisos del rol
                RolPermiso.objects.filter(rol=rol).delete()
                # Eliminar el rol
                rol.delete()
                messages.success(request, f'Rol "{rol.nombre}" eliminado exitosamente')
            return redirect('roles_lista')
        
        context = {
            'rol': rol,
            'usuarios_con_rol': usuarios_con_rol
        }
        return render(request, 'core/rol_eliminar.html', context)
        
    except Rol.DoesNotExist:
        messages.error(request, 'Rol no encontrado')
        return redirect('roles_lista')


@login_required
def pwa_test(request):
    """Página de prueba para verificar el estado de la PWA"""
    return render(request, 'pwa-test.html')


@login_required
def rol_permisos(request, rol_id):
    """Gestionar permisos de un rol - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        rol = Rol.objects.get(id=rol_id)
        
        if request.method == 'POST':
            # Limpiar permisos actuales
            RolPermiso.objects.filter(rol=rol).delete()
            
            # Agregar nuevos permisos
            permisos_seleccionados = request.POST.getlist('permisos')
            for permiso_id in permisos_seleccionados:
                permiso = Permiso.objects.get(id=permiso_id)
                RolPermiso.objects.create(rol=rol, permiso=permiso)
            
            messages.success(request, f'Permisos del rol {rol.nombre} actualizados exitosamente')
            return redirect('roles_lista')
        
        # GET - Mostrar formulario
        modulos = Modulo.objects.all().order_by('nombre')
        permisos_actuales = RolPermiso.objects.filter(rol=rol, activo=True).values_list('permiso_id', flat=True)
        
        # Organizar permisos por módulo
        modulos_con_permisos = []
        for modulo in modulos:
            permisos = Permiso.objects.filter(modulo=modulo)
            modulos_con_permisos.append({
                'modulo': modulo,
                'permisos': permisos
            })
        
        context = {
            'rol': rol,
            'modulos_con_permisos': modulos_con_permisos,
            'permisos_actuales': list(permisos_actuales),
        }
        
        return render(request, 'core/rol_permisos.html', context)
        
    except Rol.DoesNotExist:
        messages.error(request, 'Rol no encontrado')
        return redirect('roles_lista')
    except Exception as e:
        messages.error(request, f'Error al gestionar permisos: {str(e)}')
        return redirect('roles_lista')

@login_required
def asignacion_edit(request, pk):
    """Editar asignación"""
    asignacion = get_object_or_404(AsignacionInventario, pk=pk)
    if request.method == 'POST':
        form = AsignacionInventarioForm(request.POST, instance=asignacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignación actualizada exitosamente')
            return redirect('asignacion_list')
    else:
        form = AsignacionInventarioForm(instance=asignacion)
    
    return render(request, 'core/inventario/asignacion/edit.html', {
        'form': form, 'asignacion': asignacion
    })

@login_required
def asignacion_delete(request, pk):
    """Eliminar asignación"""
    asignacion = get_object_or_404(AsignacionInventario, pk=pk)
    if request.method == 'POST':
        asignacion.delete()
        messages.success(request, 'Asignación eliminada exitosamente')
        return redirect('asignacion_list')
    
    return render(request, 'core/inventario/asignacion/delete.html', {'asignacion': asignacion})

@login_required
def asignacion_devolver(request, pk):
    """Marcar asignación como devuelta"""
    asignacion = get_object_or_404(AsignacionInventario, pk=pk)
    if request.method == 'POST':
        asignacion.estado = 'devuelto'
        asignacion.fecha_devolucion = timezone.now().date()
        asignacion.save()
        messages.success(request, 'Item marcado como devuelto exitosamente')
        return redirect('asignacion_list')
    
    return render(request, 'core/inventario/asignacion/devolver.html', {'asignacion': asignacion})


# ============================================
# DASHBOARD INTELIGENTE - FUNCIONES AVANZADAS
# ============================================

@login_required
def dashboard_intelligent_data(request):
    """API para obtener datos del dashboard inteligente"""
    try:
        # Datos de rentabilidad por proyecto
        proyectos_rentabilidad = []
        proyectos = Proyecto.objects.filter(activo=True)[:10]
        
        for proyecto in proyectos:
            # Calcular ingresos del proyecto
            ingresos_proyecto_raw = Factura.objects.filter(
                proyecto=proyecto,
                estado__in=['pagada', 'enviada']
            ).aggregate(total=Sum('monto_total'))['total'] or 0
            ingresos_proyecto = Decimal(str(ingresos_proyecto_raw))
            
            # Calcular gastos del proyecto
            gastos_proyecto_raw = Gasto.objects.filter(
                proyecto=proyecto,
                aprobado=True
            ).aggregate(total=Sum('monto'))['total'] or 0
            gastos_proyecto = Decimal(str(gastos_proyecto_raw))
            
            # Calcular rentabilidad
            if ingresos_proyecto > 0:
                rentabilidad = ((ingresos_proyecto - gastos_proyecto) / ingresos_proyecto) * 100
            else:
                rentabilidad = 0
            
            proyectos_rentabilidad.append({
                'nombre': proyecto.nombre,
                'rentabilidad': round(rentabilidad, 2),
                'ingresos': float(ingresos_proyecto),
                'gastos': float(gastos_proyecto)
            })
        
        # Datos de flujo de caja mensual
        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio']
        flujo_caja = []
        
        for i, mes in enumerate(meses):
            # Simular datos de flujo de caja (en producción se calcularían reales)
            ingresos_mes = float(Factura.objects.filter(
                fecha_emision__month=i+1,
                estado__in=['pagada', 'enviada']
            ).aggregate(total=Sum('monto_total'))['total'] or 0)
            
            gastos_mes = float(Gasto.objects.filter(
                fecha_gasto__month=i+1,
                aprobado=True
            ).aggregate(total=Sum('monto'))['total'] or 0)
            
            flujo_caja.append({
                'mes': mes,
                'ingresos': ingresos_mes,
                'gastos': gastos_mes,
                'neto': ingresos_mes - gastos_mes
            })
        
        # Métricas de productividad
        total_proyectos = Proyecto.objects.filter(activo=True).count()
        proyectos_completados = Proyecto.objects.filter(estado='completado').count()
        eficiencia_proyectos = (proyectos_completados / total_proyectos * 100) if total_proyectos > 0 else 0
        
        # CÁLCULO DE TIEMPO PROMEDIO ELIMINADO - YA NO SE USA fecha_fin
        tiempo_promedio = 0
        
        # KPIs inteligentes
        satisfaccion_cliente = 85  # Simulado - en producción se calcularía de encuestas
        calidad_obra = 90  # Simulado - en producción se calcularía de inspecciones
        rentabilidad_general = 25  # Simulado - en producción se calcularía real
        eficiencia_general = round(eficiencia_proyectos, 1)
        
        # Datos de tendencias
        tendencias = {
            'ventas': '+12%',
            'clientes': '+8%',
            'costos': '-5%',
            'eficiencia': '+15%'
        }
        
        # Predicciones del sistema
        predicciones = {
            'ventas_proximo_mes': float(Factura.objects.filter(
                fecha_emision__month=timezone.now().month + 1
            ).aggregate(total=Sum('monto_total'))['total'] or 25000),
            'rentabilidad_esperada': 28.5
        }
        
        # Comparación de períodos
        mes_actual = timezone.now().month
        mes_anterior = mes_actual - 1 if mes_actual > 1 else 12
        
        periodo_actual = float(Factura.objects.filter(
            fecha_emision__month=mes_actual
        ).aggregate(total=Sum('monto_total'))['total'] or 0)
        
        periodo_anterior = float(Factura.objects.filter(
            fecha_emision__month=mes_anterior
        ).aggregate(total=Sum('monto_total'))['total'] or 0)
        
        crecimiento = ((periodo_actual - periodo_anterior) / periodo_anterior * 100) if periodo_anterior > 0 else 0
        
        data = {
            'rentabilidad': {
                'proyectos': proyectos_rentabilidad,
                'total': round(rentabilidad_general, 1),
                'proyectos_activos': total_proyectos
            },
            'flujoCaja': {
                'mensual': flujo_caja,
                'ingresos_mes': periodo_actual,
                'gastos_mes': float(Gasto.objects.filter(
                    fecha_gasto__month=mes_actual,
                    aprobado=True
                ).aggregate(total=Sum('monto'))['total'] or 0)
            },
            'productividad': {
                'eficiencia': round(eficiencia_proyectos, 1),
                'tiempo_promedio': round(tiempo_promedio, 0),
                'proyectos_completados': proyectos_completados,
                'total_proyectos': total_proyectos
            },
            'kpis': {
                'satisfaccion': satisfaccion_cliente,
                'calidad': calidad_obra,
                'rentabilidad': rentabilidad_general,
                'eficiencia': eficiencia_general
            },
            'tendencias': tendencias,
            'predicciones': predicciones,
            'comparacion': {
                'periodo_anterior': periodo_anterior,
                'periodo_actual': periodo_actual,
                'crecimiento': round(crecimiento, 1)
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        logger.error(f'Error obteniendo datos del dashboard inteligente: {str(e)}')
        return JsonResponse({
            'error': 'Error obteniendo datos',
            'message': str(e)
        }, status=500)


@login_required
def dashboard_intelligent_analytics(request):
    """Vista para análisis avanzado del dashboard inteligente"""
    try:
        # Obtener métricas avanzadas
        context = {
            'titulo': 'Análisis Inteligente del Sistema',
            'fecha_analisis': timezone.now().strftime('%d/%m/%Y %H:%M'),
        }
        
        return render(request, 'core/dashboard_intelligent_analytics.html', context)
        
    except Exception as e:
        logger.error(f'Error en análisis inteligente: {str(e)}')
        messages.error(request, f'Error al cargar análisis: {str(e)}')
        return redirect('dashboard')

@login_required
def roles_resumen(request):
    """Mostrar resumen completo de todos los roles y sus permisos - SOLO SUPERUSUARIOS"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    try:
        # Obtener todos los roles con sus permisos
        roles = Rol.objects.all().order_by('nombre')
        modulos = Modulo.objects.all().order_by('nombre')
        
        # Crear resumen detallado
        resumen_roles = []
        for rol in roles:
            permisos_rol = RolPermiso.objects.filter(rol=rol, activo=True).select_related('permiso__modulo')
            
            # Organizar permisos por módulo
            permisos_por_modulo = {}
            for rp in permisos_rol:
                modulo_nombre = rp.permiso.modulo.nombre
                if modulo_nombre not in permisos_por_modulo:
                    permisos_por_modulo[modulo_nombre] = []
                permisos_por_modulo[modulo_nombre].append(rp.permiso.nombre)
            
            resumen_roles.append({
                'rol': rol,
                'permisos_por_modulo': permisos_por_modulo,
                'total_permisos': permisos_rol.count()
            })
        
        # Estadísticas generales
        total_roles = roles.count()
        total_modulos = modulos.count()
        total_permisos = Permiso.objects.count()
        
        context = {
            'resumen_roles': resumen_roles,
            'modulos': modulos,
            'total_roles': total_roles,
            'total_modulos': total_modulos,
            'total_permisos': total_permisos,
        }
        
        return render(request, 'core/roles_resumen.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al generar resumen: {str(e)}')
        return redirect('roles_lista')


def offline_page(request):
    """Página offline para PWA"""
    return render(request, 'offline.html')

def test_view(request):
    """Vista de prueba para diagnosticar problemas"""
    from django.utils import timezone
    context = {
        'now': timezone.now(),
    }
    return render(request, 'core/test.html', context)

@login_required
def planilla_proyecto(request, proyecto_id):
    """Planilla de personal del proyecto con control de anticipos"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener colaboradores asignados al proyecto
    colaboradores_asignados = proyecto.colaboradores.all().order_by('nombre')
    
    # Obtener configuración de planilla (retenciones y bonos)
    try:
        configuracion_planilla = proyecto.configuracion_planilla
    except ConfiguracionPlanilla.DoesNotExist:
        # Crear configuración por defecto si no existe
        configuracion_planilla = ConfiguracionPlanilla.objects.create(
            proyecto=proyecto,
            retencion_seguro_social=0,
            retencion_seguro_educativo=0,
            bono_general=0,
            bono_produccion=0,
            aplicar_retenciones=True,
            aplicar_bonos=True
        )
    
    # Obtener anticipos del proyecto (pendientes y liquidados)
    anticipos = AnticipoProyecto.objects.filter(
        proyecto=proyecto
    ).select_related('colaborador')
    
    # Calcular totales de anticipos
    total_anticipos_pendientes = anticipos.filter(estado='pendiente').aggregate(total=Sum('monto'))['total'] or 0
    total_anticipos_liquidados = anticipos.filter(estado='liquidado').aggregate(total=Sum('monto'))['total'] or 0
    total_anticipos = total_anticipos_pendientes + total_anticipos_liquidados
    
    # Calcular salarios netos y totales de la planilla
    total_salarios = 0
    total_anticipos_aplicados = 0
    total_salarios_netos = 0
    total_retenciones = 0
    total_bonos = 0
    
    for colaborador in colaboradores_asignados:
        salario_colaborador = colaborador.salario or 0
        
        # Calcular anticipos del colaborador
        anticipos_colaborador = anticipos.filter(colaborador=colaborador)
        anticipos_pendientes = anticipos_colaborador.filter(estado='pendiente').aggregate(total=Sum('monto'))['total'] or 0
        anticipos_liquidados = anticipos_colaborador.filter(estado='liquidado').aggregate(total=Sum('monto'))['total'] or 0
        total_anticipos_colaborador = anticipos_pendientes + anticipos_liquidados
        
        # Salario neto = Salario base - total de anticipos
        salario_neto = salario_colaborador - total_anticipos_colaborador
        
        # Determinar el estado actual del colaborador
        if anticipos_pendientes > 0:
            colaborador.estado_actual = 'pendiente'
            colaborador.estado_display = 'Pendiente'
        elif anticipos_liquidados > 0:
            colaborador.estado_actual = 'liquidado'
            colaborador.estado_display = 'Liquidado'
        else:
            colaborador.estado_actual = 'sin_anticipos'
            colaborador.estado_display = 'Sin Anticipos'
        
        # Calcular retenciones y bonos solo si el colaborador los tiene habilitados
        retenciones_monto = Decimal('0')
        if colaborador.aplica_retenciones:
            retenciones_monto = Decimal(str(configuracion_planilla.calcular_retenciones(salario_colaborador)))
        
        bonos_monto = Decimal('0')
        if configuracion_planilla.aplicar_bonos:
            # Bono general
            if colaborador.aplica_bono_general:
                bonos_monto += Decimal(str(configuracion_planilla.bono_general))
            
            # Bono de producción
            if colaborador.aplica_bono_produccion:
                bono_prod = (Decimal(str(salario_colaborador)) * Decimal(str(configuracion_planilla.bono_produccion))) / Decimal('100')
                bonos_monto += bono_prod
        
        # Salario con bonos y retenciones (todo en Decimal)
        salario_colaborador_decimal = Decimal(str(salario_colaborador))
        salario_bruto = salario_colaborador_decimal + bonos_monto
        salario_neto_con_retenciones = salario_bruto - retenciones_monto
        
        # Calcular valores quincenales
        salario_quincenal = salario_colaborador_decimal / Decimal('2')
        bonos_quincenal = bonos_monto / Decimal('2')  # Bonos se dividen entre 2 quincenas
        retenciones_quincenal = retenciones_monto / Decimal('2')  # Retenciones se dividen entre 2 quincenas
        
        # Salario quincenal neto = Salario quincenal + Bonos quincenal - Retenciones quincenal - Anticipos
        total_anticipos_decimal = Decimal(str(total_anticipos_colaborador))
        salario_quincenal_neto = salario_quincenal + bonos_quincenal - retenciones_quincenal - total_anticipos_decimal
        
        # Agregar campos calculados al colaborador
        colaborador.salario_mensual = salario_colaborador
        colaborador.bonos_monto = bonos_monto
        colaborador.bonos_monto_quincenal = bonos_quincenal
        colaborador.retenciones_monto = retenciones_monto
        colaborador.retenciones_monto_quincenal = retenciones_quincenal
        colaborador.salario_bruto = salario_bruto
        colaborador.salario_quincenal = salario_quincenal
        colaborador.salario_quincenal_bruto = salario_quincenal + bonos_quincenal
        colaborador.salario_neto = salario_neto
        colaborador.salario_neto_quincenal = salario_quincenal_neto
        colaborador.deuda_anticipos = anticipos_pendientes
        colaborador.anticipos_liquidados = anticipos_liquidados
        
        total_salarios += salario_colaborador
        total_anticipos_aplicados += total_anticipos_colaborador
        total_salarios_netos += salario_neto
        total_retenciones += retenciones_monto
        total_bonos += bonos_monto
    
    # Calcular histórico de pagos usando el nuevo modelo PlanillaLiquidada
    planillas_liquidadas = PlanillaLiquidada.objects.filter(proyecto=proyecto)
    total_historico_pagado = planillas_liquidadas.aggregate(total=Sum('total_planilla'))['total'] or Decimal('0.00')
    planillas_generadas = planillas_liquidadas.count()
    
    # Promedio por planilla
    promedio_por_planilla = total_historico_pagado / planillas_generadas if planillas_generadas > 0 else Decimal('0.00')
    
    # Obtener las últimas 5 planillas liquidadas
    ultimas_planillas = PlanillaLiquidada.objects.filter(proyecto=proyecto).order_by('-año', '-mes', '-fecha_liquidacion')[:5]
    
    # Calcular totales quincenales con retenciones
    total_salarios_quincenal = total_salarios / 2
    total_retenciones_quincenal = total_retenciones / 2
    total_bonos_quincenal = total_bonos / 2
    
    # Salario bruto quincenal (base + bonos)
    total_salarios_bruto_quincenal = total_salarios_quincenal + total_bonos_quincenal
    
    # Salario neto quincenal (bruto - retenciones - anticipos)
    total_neto_pagar_quincenal = total_salarios_bruto_quincenal - total_retenciones_quincenal - total_anticipos_aplicados
    
    # NOTA: Los anticipos se eliminan al liquidar la planilla, por lo que cada planilla empieza limpia
    
    context = {
        'proyecto': proyecto,
        'colaboradores_asignados': colaboradores_asignados,
        'anticipos': anticipos,
        'total_anticipos': total_anticipos,
        'total_liquidado': total_anticipos_liquidados,
        'saldo_pendiente': total_anticipos_pendientes,
        'total_salarios': total_salarios,
        'total_salarios_quincenal': total_salarios_quincenal,
        'total_retenciones': total_retenciones,
        'total_retenciones_quincenal': total_retenciones_quincenal,
        'total_bonos': total_bonos,
        'total_bonos_quincenal': total_bonos_quincenal,
        'total_anticipos_aplicados': total_anticipos_aplicados,
        'total_salarios_netos': total_salarios_netos,
        'total_neto_pagar_quincenal': total_neto_pagar_quincenal,
        'total_historico_pagado': total_historico_pagado,
        'planillas_generadas': planillas_generadas,
        'promedio_por_planilla': promedio_por_planilla,
        'ultimas_planillas': ultimas_planillas,
        'configuracion_planilla': configuracion_planilla,
    }
    
    return render(request, 'core/proyectos/planilla.html', context)


@login_required
def resetear_historico_nomina(request, proyecto_id):
    """Resetear el histórico de nómina del proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        try:
            # Contar registros antes de eliminar
            planillas_personal = PlanillaLiquidada.objects.filter(proyecto=proyecto).count()
            trabajadores_inactivos = TrabajadorDiario.objects.filter(proyecto=proyecto, activo=False).count()
            
            # Eliminar planillas liquidadas del personal
            PlanillaLiquidada.objects.filter(proyecto=proyecto).delete()
            
            # Eliminar trabajadores diarios inactivos
            TrabajadorDiario.objects.filter(proyecto=proyecto, activo=False).delete()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Reset',
                modulo='Histórico Nómina',
                descripcion=f'Histórico reseteado - {planillas_personal} planillas + {trabajadores_inactivos} trabajadores eliminados',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(
                request,
                f'✅ <strong>Histórico de nómina reseteado exitosamente</strong><br>'
                f'🗑️ {planillas_personal} planilla(s) de personal eliminadas<br>'
                f'🗑️ {trabajadores_inactivos} trabajador(es) diario(s) eliminados',
                extra_tags='html'
            )
            
            return redirect('proyecto_dashboard', proyecto_id=proyecto_id)
            
        except Exception as e:
            logger.error(f"Error reseteando histórico de nómina del proyecto {proyecto_id}: {e}")
            messages.error(request, f'Error al resetear el histórico: {str(e)}')
            return redirect('proyecto_dashboard', proyecto_id=proyecto_id)
    
    return redirect('proyecto_dashboard', proyecto_id=proyecto_id)


@login_required
def liquidar_y_generar_planilla(request, proyecto_id):
    """Liquidar anticipos pendientes y generar PDF de la planilla"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        try:
            # Obtener mes, año y quincena del formulario
            mes = int(request.POST.get('mes', timezone.now().month))
            año = int(request.POST.get('año', timezone.now().year))
            quincena = int(request.POST.get('quincena', 1))  # Siempre requerida
            
            # Verificar si ya existe una planilla para este mes/año/quincena
            if PlanillaLiquidada.objects.filter(proyecto=proyecto, mes=mes, año=año, quincena=quincena).exists():
                mes_nombre = dict(PlanillaLiquidada.MESES_CHOICES).get(mes, '')
                quincena_nombre = dict(PlanillaLiquidada.QUINCENA_CHOICES).get(quincena, '')
                messages.error(request, f'Ya existe una planilla liquidada para {mes_nombre} {año} - {quincena_nombre} en este proyecto.')
                return redirect('planilla_proyecto', proyecto_id=proyecto_id)
            
            # Obtener colaboradores asignados al proyecto
            colaboradores_asignados = proyecto.colaboradores.all()
            
            # Calcular total de salarios MENSUAL
            total_salarios_mensual = sum(c.salario or 0 for c in colaboradores_asignados)
            
            # Calcular salario QUINCENAL (dividir entre 2)
            total_salarios_quincenal = total_salarios_mensual / 2
            
            # Liquidar TODOS los anticipos (pendientes y liquidados)
            anticipos_a_liquidar = AnticipoProyecto.objects.filter(
                proyecto=proyecto
            ).exclude(estado='procesado')
            
            total_anticipos_liquidados = anticipos_a_liquidar.aggregate(total=Sum('monto'))['total'] or 0
            
            # ELIMINAR todos los anticipos después de liquidar la planilla
            # Los anticipos ya fueron descontados del salario, no deben persistir
            anticipos_a_liquidar.delete()
            
            # Total de la planilla = Salario QUINCENAL - Anticipos liquidados
            total_planilla = Decimal(str(total_salarios_quincenal)) - Decimal(str(total_anticipos_liquidados))
            
            # Crear registro de planilla liquidada
            planilla = PlanillaLiquidada.objects.create(
                proyecto=proyecto,
                mes=mes,
                año=año,
                quincena=quincena,
                total_salarios=Decimal(str(total_salarios_quincenal)),
                total_anticipos=Decimal(str(total_anticipos_liquidados)),
                total_planilla=total_planilla,
                cantidad_personal=colaboradores_asignados.count(),
                liquidada_por=request.user
            )
            
            # Guardar registro de la planilla generada en el log de actividad
            mes_nombre = dict(PlanillaLiquidada.MESES_CHOICES).get(mes, '')
            quincena_nombre = dict(PlanillaLiquidada.QUINCENA_CHOICES).get(quincena, '')
            
            LogActividad.objects.create(
                usuario=request.user,
                accion='Liquidar Planilla',
                modulo='Planilla Personal',
                descripcion=f'Planilla {mes_nombre} {año} - {quincena_nombre} liquidada - Salarios Quincenales: ${total_salarios_quincenal:,.2f} - Anticipos: ${total_anticipos_liquidados:,.2f} - Total: ${total_planilla:,.2f}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(
                request,
                f'✅ <strong>Planilla Quincenal liquidada exitosamente</strong><br>'
                f'📅 <strong>Período:</strong> {mes_nombre} {año} - {quincena_nombre}<br>'
                f'💼 Total Salarios Quincenales: <strong>${total_salarios_quincenal:,.2f}</strong> (Mensual: ${total_salarios_mensual:,.2f})<br>'
                f'💰 Anticipos Descontados: <strong>${total_anticipos_liquidados:,.2f}</strong><br>'
                f'📋 <strong>TOTAL A PAGAR: ${total_planilla:,.2f}</strong><br>'
                f'👥 Personal: <strong>{colaboradores_asignados.count()}</strong>',
                extra_tags='html'
            )
            
            # Redirigir a generar el PDF
            return redirect('planilla_proyecto_pdf', proyecto_id=proyecto_id)
            
        except Exception as e:
            logger.error(f"Error liquidando planilla del proyecto {proyecto_id}: {e}")
            messages.error(request, f'Error al liquidar la planilla: {str(e)}')
            return redirect('planilla_proyecto', proyecto_id=proyecto_id)
    
    return redirect('planilla_proyecto', proyecto_id=proyecto_id)


@login_required
def configurar_planilla_proyecto(request, proyecto_id):
    """Configurar retenciones y bonos para la planilla del proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener colaboradores del proyecto
    colaboradores = proyecto.colaboradores.all().order_by('nombre')
    
    # Obtener o crear configuración
    configuracion, created = ConfiguracionPlanilla.objects.get_or_create(
        proyecto=proyecto,
        defaults={
            'retencion_seguro_social': 0,
            'retencion_seguro_educativo': 0,
            'bono_general': 0,
            'bono_produccion': 0,
            'aplicar_retenciones': True,
            'aplicar_bonos': True
        }
    )
    
    if request.method == 'POST':
        form = ConfiguracionPlanillaForm(request.POST, instance=configuracion)
        
        # Procesar asignación de bonos a colaboradores
        for colaborador in colaboradores:
            # Bonos (pueden ser individuales)
            aplica_bono_general = f'bono_general_{colaborador.id}' in request.POST
            aplica_bono_produccion = f'bono_produccion_{colaborador.id}' in request.POST
            
            # Actualizar colaborador
            colaborador.aplica_bono_general = aplica_bono_general
            colaborador.aplica_bono_produccion = aplica_bono_produccion
            colaborador.aplica_retenciones = True  # Retenciones SIEMPRE para todos
            colaborador.save()
        
        if form.is_valid():
            config = form.save(commit=False)
            config.modificado_por = request.user
            config.save()
            messages.success(request, '✅ Configuración de planilla y asignación de bonos actualizada exitosamente.')
            return redirect('planilla_proyecto', proyecto_id=proyecto_id)
    else:
        form = ConfiguracionPlanillaForm(instance=configuracion)
    
    context = {
        'proyecto': proyecto,
        'form': form,
        'configuracion': configuracion,
        'colaboradores': colaboradores,
    }
    
    return render(request, 'core/proyectos/configurar_planilla.html', context)


@login_required
def planilla_proyecto_pdf(request, proyecto_id):
    """Generar PDF de la planilla del proyecto con desglose completo"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener configuración de planilla (retenciones y bonos)
    try:
        configuracion_planilla = proyecto.configuracion_planilla
    except ConfiguracionPlanilla.DoesNotExist:
        configuracion_planilla = ConfiguracionPlanilla.objects.create(
            proyecto=proyecto,
            retencion_seguro_social=Decimal('0'),
            retencion_seguro_educativo=Decimal('0'),
            bono_general=Decimal('0'),
            bono_produccion=Decimal('0'),
            aplicar_retenciones=True,
            aplicar_bonos=True
        )
    
    # Obtener colaboradores asignados al proyecto
    colaboradores_asignados = proyecto.colaboradores.all().order_by('nombre')
    
    # Obtener anticipos del proyecto
    anticipos = AnticipoProyecto.objects.filter(proyecto=proyecto).select_related('colaborador')
    
    # Calcular totales iniciales
    total_anticipos = anticipos.filter(estado='pendiente').aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0')
    
    total_liquidado = anticipos.filter(estado='liquidado').aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0')
    
    # Variables para acumular totales
    total_salarios = Decimal('0')
    total_bonos = Decimal('0')
    total_retenciones = Decimal('0')
    total_anticipos_aplicados = Decimal('0')
    total_salarios_netos = Decimal('0')
    
    # Calcular salarios con bonos y retenciones
    for colaborador in colaboradores_asignados:
        salario_colaborador = colaborador.salario or Decimal('0')
        
        # Calcular anticipos del colaborador
        anticipos_colaborador = anticipos.filter(colaborador=colaborador)
        total_anticipos_colaborador = anticipos_colaborador.filter(estado='pendiente').aggregate(total=Sum('monto'))['total'] or Decimal('0')
        
        # Calcular retenciones y bonos solo si el colaborador los tiene habilitados
        retenciones_monto = Decimal('0')
        if colaborador.aplica_retenciones and configuracion_planilla.aplicar_retenciones:
            retenciones_monto = configuracion_planilla.calcular_retenciones(salario_colaborador)
        
        bonos_monto = Decimal('0')
        if configuracion_planilla.aplicar_bonos:
            if colaborador.aplica_bono_general:
                bonos_monto += configuracion_planilla.bono_general
            if colaborador.aplica_bono_produccion:
                bonos_monto += (salario_colaborador * configuracion_planilla.bono_produccion) / Decimal('100')
        
        # Salario quincenal (todo dividido entre 2)
        salario_quincenal = salario_colaborador / Decimal('2')
        bonos_quincenal = bonos_monto / Decimal('2')
        retenciones_quincenal = retenciones_monto / Decimal('2')
        
        # Salario neto quincenal
        salario_neto = salario_quincenal + bonos_quincenal - retenciones_quincenal - total_anticipos_colaborador
        
        # Acumular totales
        total_salarios += salario_colaborador
        total_bonos += bonos_monto
        total_retenciones += retenciones_monto
        total_anticipos_aplicados += total_anticipos_colaborador
        total_salarios_netos += salario_neto
    
    # Totales quincenales
    total_salarios_quincenal = total_salarios / Decimal('2')
    total_bonos_quincenal = total_bonos / Decimal('2')
    total_retenciones_quincenal = total_retenciones / Decimal('2')
    
    # Crear el PDF en orientación horizontal
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    elements = []
    
    # Estilos personalizados
    styles = getSampleStyleSheet()
    
    # Estilo para el encabezado principal
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e3a8a'),
        fontName='Helvetica-Bold',
        leading=28
    )
    
    # Estilo para subtítulo
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=14,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#3b82f6'),
        fontName='Helvetica-Bold'
    )
    
    # Estilo para secciones
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=10,
        textColor=colors.HexColor('#1e3a8a'),
        fontName='Helvetica-Bold'
    )
    
    # Logo y encabezado
    try:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'LOGO-TELECOM-small.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=120, height=60)
        else:
            logo = Paragraph("<b>TELECOM</b><br/>Technology Panama INC.", styles['Normal'])
    except:
        logo = Paragraph("<b>TELECOM</b><br/>Technology Panama INC.", styles['Normal'])
    
    # Fecha de generación
    from datetime import datetime
    import pytz
    guatemala_tz = pytz.timezone('America/Guatemala')
    fecha_generacion = datetime.now(guatemala_tz).strftime('%d/%m/%Y %H:%M')
    
    # Header con logo y fecha
    header_data = [[
        logo,
        Paragraph(
            f'<b>PLANILLA DE PERSONAL QUINCENAL</b><br/>'
            f'<font size="10">Generado: {fecha_generacion}</font>',
            ParagraphStyle('HeaderRight', parent=styles['Normal'], fontSize=12, alignment=TA_RIGHT)
        )
    ]]
    header_table = Table(header_data, colWidths=[3*inch, 7*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))
    
    # Título del proyecto
    elements.append(Paragraph(f"<b>PROYECTO: {proyecto.nombre.upper()}</b>", header_style))
    
    # Cliente
    if proyecto.cliente:
        elements.append(Paragraph(
            f'<b>Cliente:</b> {proyecto.cliente.razon_social}', 
            subtitle_style
        ))
    
    elements.append(Spacer(1, 15))
    
    # Resumen de configuración de planilla
    config_data = [[
        Paragraph('<b>CONFIGURACIÓN DE PLANILLA</b>', section_style)
    ]]
    config_table = Table(config_data, colWidths=[10*inch])
    config_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e0e7ff')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(config_table)
    elements.append(Spacer(1, 5))
    
    # Detalles de configuración
    config_details = [
        ['Retenciones', ''],
        ['• Seguro Social:', f'${configuracion_planilla.retencion_seguro_social:,.2f} mensuales'],
        ['• Seguro Educativo:', f'${configuracion_planilla.retencion_seguro_educativo:,.2f} mensuales'],
        ['Bonos', ''],
        ['• Bono General:', f'${configuracion_planilla.bono_general:,.2f} mensuales'],
        ['• Bono de Producción:', f'{configuracion_planilla.bono_produccion}% sobre salario base'],
    ]
    
    config_details_table = Table(config_details, colWidths=[2.5*inch, 7.5*inch])
    config_details_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 3), (0, 3), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#374151')),
    ]))
    elements.append(config_details_table)
    elements.append(Spacer(1, 20))
    
    # Sección de desglose por colaborador
    if colaboradores_asignados:
        # Encabezado de tabla
        elements.append(Paragraph('<b>DESGLOSE QUINCENAL POR COLABORADOR</b>', section_style))
        elements.append(Spacer(1, 10))
        
        # Encabezados de la tabla
        headers = [
            'No.', 'Colaborador', 
            'Salario\nQuincenal', 
            'Bonos', 
            'Retenciones', 
            'Anticipos',
            'Pago Neto\nQuincenal'
        ]
        data = [headers]
        
        # Variables para totales
        total_salarios_q = Decimal('0')
        total_bonos_q = Decimal('0')
        total_retenciones_q = Decimal('0')
        total_anticipos_q = Decimal('0')
        total_neto_q = Decimal('0')
        
        for i, colaborador in enumerate(colaboradores_asignados, 1):
            salario_colaborador = colaborador.salario or Decimal('0')
            
            # Calcular anticipos del colaborador
            anticipos_colaborador = anticipos.filter(colaborador=colaborador)
            total_anticipos_colaborador = anticipos_colaborador.filter(estado='pendiente').aggregate(total=Sum('monto'))['total'] or Decimal('0')
            
            # Calcular retenciones y bonos
            retenciones_monto = Decimal('0')
            if colaborador.aplica_retenciones and configuracion_planilla.aplicar_retenciones:
                retenciones_monto = configuracion_planilla.calcular_retenciones(salario_colaborador)
            
            bonos_monto = Decimal('0')
            if configuracion_planilla.aplicar_bonos:
                if colaborador.aplica_bono_general:
                    bonos_monto += configuracion_planilla.bono_general
                if colaborador.aplica_bono_produccion:
                    bonos_monto += (salario_colaborador * configuracion_planilla.bono_produccion) / Decimal('100')
            
            # Valores quincenales
            salario_quincenal = salario_colaborador / Decimal('2')
            bonos_quincenal = bonos_monto / Decimal('2')
            retenciones_quincenal = retenciones_monto / Decimal('2')
            
            # Salario neto quincenal
            salario_neto = salario_quincenal + bonos_quincenal - retenciones_quincenal - total_anticipos_colaborador
            
            data.append([
                str(i),
                colaborador.nombre,
                f"${float(salario_quincenal):,.2f}",
                f"+${float(bonos_quincenal):,.2f}",
                f"-${float(retenciones_quincenal):,.2f}",
                f"-${float(total_anticipos_colaborador):,.2f}",
                f"${float(salario_neto):,.2f}"
            ])
            
            # Acumular totales
            total_salarios_q += salario_quincenal
            total_bonos_q += bonos_quincenal
            total_retenciones_q += retenciones_quincenal
            total_anticipos_q += total_anticipos_colaborador
            total_neto_q += salario_neto
        
        # Agregar fila de totales
        data.append([
            '', 
            'TOTALES:', 
            f"${float(total_salarios_q):,.2f}", 
            f"+${float(total_bonos_q):,.2f}", 
            f"-${float(total_retenciones_q):,.2f}",
            f"-${float(total_anticipos_q):,.2f}",
            f"${float(total_neto_q):,.2f}"
        ])
        
        # Crear tabla con columnas ajustadas
        table = Table(data, colWidths=[0.4*inch, 2.5*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.3*inch])
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            
            # Fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('ALIGN', (2, -1), (-1, -1), 'RIGHT'),
            ('ALIGN', (1, -1), (1, -1), 'LEFT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
            ('TOPPADDING', (0, -1), (-1, -1), 10),
            
            # Datos
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (2, 1), (-1, -2), 'RIGHT'),
            ('ALIGN', (0, 1), (1, -2), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 1), (-1, -2), 6),
            ('RIGHTPADDING', (0, 1), (-1, -2), 6),
            ('TOPPADDING', (0, 1), (-1, -2), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 6),
            
            # Colores alternos en filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1e3a8a')),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    # Resumen financiero ejecutivo
    elements.append(Spacer(1, 15))
    elements.append(Paragraph('<b>RESUMEN FINANCIERO QUINCENAL</b>', section_style))
    elements.append(Spacer(1, 10))
    
    # Cálculo del total neto a pagar
    total_bruto_quincenal = total_salarios_quincenal + total_bonos_quincenal
    total_deducciones = total_retenciones_quincenal + total_anticipos_aplicados
    total_neto_pagar = total_bruto_quincenal - total_deducciones
    
    resumen_data = [
        [Paragraph('<b>CONCEPTO</b>', styles['Normal']), Paragraph('<b>MONTO QUINCENAL</b>', styles['Normal'])],
        ['Salarios Base (Quincenal)', f"${float(total_salarios_quincenal):,.2f}"],
        ['(+) Bonos Totales', f"${float(total_bonos_quincenal):,.2f}"],
        ['', ''],
        [Paragraph('<b>Salario Bruto Quincenal</b>', styles['Normal']), Paragraph(f'<b>${float(total_bruto_quincenal):,.2f}</b>', styles['Normal'])],
        ['', ''],
        ['(-) Retenciones Totales', f"${float(total_retenciones_quincenal):,.2f}"],
        ['(-) Anticipos Aplicados', f"${float(total_anticipos_aplicados):,.2f}"],
        ['', ''],
        [Paragraph('<b>TOTAL NETO A PAGAR</b>', ParagraphStyle('TotalStyle', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold')), 
         Paragraph(f'<b>${float(total_neto_pagar):,.2f}</b>', ParagraphStyle('TotalValueStyle', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', textColor=colors.HexColor('#059669')))],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[5*inch, 2.5*inch])
    resumen_table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        
        # Fila de salario bruto (índice 4)
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#dbeafe')),
        
        # Fila de total neto (última fila)
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1e3a8a')),
    ]))
    
    elements.append(resumen_table)
    
    # Pie de página con información corporativa
    elements.append(Spacer(1, 30))
    
    # Configuración del sistema
    config = ConfiguracionSistema.get_config()
    
    footer_text = f"""
    <para align="center">
        <b>{config.nombre_empresa}</b><br/>
        Technology Panama INC.<br/>
        <font size="8">Correo: info@telecompanama.com | Tel: +507 206-3456</font><br/>
        <font size="7" color="#9ca3af">Documento generado electrónicamente el {fecha_generacion}</font>
    </para>
    """
    
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta HTTP
    filename = f"planilla_proyecto_{proyecto.nombre.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.pdf"
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def crear_anticipo_masivo(request, proyecto_id):
    """Crear anticipo masivo para todos los colaboradores del proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        monto = request.POST.get('monto')
        concepto = request.POST.get('concepto', 'Anticipo masivo por proyecto')
        
        if monto and monto.isdigit():
            monto = Decimal(monto)
            colaboradores = proyecto.colaboradores.all()
            
            # Crear anticipos para todos los colaboradores
            anticipos_creados = []
            for colaborador in colaboradores:
                anticipo = AnticipoProyecto.objects.create(
                    proyecto=proyecto,
                    colaborador=colaborador,
                    monto=monto,
                    tipo='masivo',
                    concepto=concepto,
                    observaciones=f"Anticipo masivo - {concepto}"
                )
                anticipos_creados.append(anticipo)
            
            messages.success(
                request, 
                f'✅ Se crearon exitosamente {len(anticipos_creados)} anticipos de ${monto:,.2f} cada uno para el proyecto "{proyecto.nombre}". Total desembolsado: ${(monto * len(anticipos_creados)):,.2f}'
            )
            
            return redirect('planilla_proyecto', proyecto_id=proyecto.id)
    
    context = {
        'proyecto': proyecto,
        'total_colaboradores': proyecto.colaboradores.count()
    }
    
    return render(request, 'core/proyectos/crear_anticipo_masivo.html', context)


@login_required
def crear_anticipo_individual(request, proyecto_id):
    """Crear anticipo individual para un colaborador específico"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        colaborador_id = request.POST.get('colaborador')
        monto = request.POST.get('monto')
        concepto = request.POST.get('concepto', 'Anticipo individual')
        observaciones = request.POST.get('observaciones', '')
        
        try:
            # Validar que todos los campos requeridos estén presentes
            if not colaborador_id:
                messages.error(request, '❌ Error: Debes seleccionar un colaborador.')
            elif not monto:
                messages.error(request, '❌ Error: Debes ingresar un monto.')
            else:
                # Intentar convertir el monto a Decimal
                monto_decimal = Decimal(str(monto))
                
                if monto_decimal <= 0:
                    messages.error(request, '❌ Error: El monto debe ser mayor a cero.')
                else:
                    colaborador = get_object_or_404(Colaborador, id=colaborador_id)
                    
                    # Verificar que el colaborador esté asignado al proyecto
                    if colaborador in proyecto.colaboradores.all():
                        anticipo = AnticipoProyecto.objects.create(
                            proyecto=proyecto,
                            colaborador=colaborador,
                            monto=monto_decimal,
                            tipo='individual',
                            concepto=concepto,
                            observaciones=observaciones
                        )
                        
                        messages.success(
                            request, 
                            f'✅ Anticipo creado exitosamente: ${monto_decimal:,.2f} para {colaborador.nombre} en el proyecto "{proyecto.nombre}". Concepto: {concepto}'
                        )
                        
                        return redirect('planilla_proyecto', proyecto_id=proyecto.id)
                    else:
                        messages.error(request, '❌ Error: El colaborador seleccionado no está asignado a este proyecto. Por favor verifica la selección.')
        except (ValueError, InvalidOperation) as e:
            logger.error(f"Error al crear anticipo individual: {e}")
            messages.error(request, '❌ Error: El monto ingresado no es válido. Por favor ingresa un número válido.')
    
    context = {
        'proyecto': proyecto,
        'colaboradores_disponibles': proyecto.colaboradores.all().order_by('nombre')
    }
    
    return render(request, 'core/proyectos/crear_anticipo_individual.html', context)


@login_required
def liquidar_anticipo(request, anticipo_id):
    """Liquidar un anticipo específico"""
    anticipo = get_object_or_404(AnticipoProyecto, id=anticipo_id)

    if request.method == 'POST':
        if anticipo.estado == 'pendiente':
            anticipo.liquidar_anticipo(request.user)
            messages.success(
                request, 
                f'Anticipo de ${anticipo.monto} para {anticipo.colaborador.nombre} ha sido liquidado'
            )
        else:
            messages.warning(request, 'Este anticipo ya no está pendiente de liquidación')

    return redirect('planilla_proyecto', proyecto_id=anticipo.proyecto.id)


@login_required
def calendario_pagos_proyecto(request, proyecto_id):
    """Calendario de pagos y anticipos del proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener mes y año de la URL o usar el actual
    mes = int(request.GET.get('mes', timezone.now().month))
    año = int(request.GET.get('año', timezone.now().year))
    
    # Obtener colaboradores asignados
    colaboradores = proyecto.colaboradores.all().order_by('nombre')
    
    # Obtener anticipos del mes
    anticipos_mes = AnticipoProyecto.objects.filter(
        proyecto=proyecto,
        fecha_anticipo__month=mes,
        fecha_anticipo__year=año
    ).order_by('fecha_anticipo')
    
    # Obtener anticipos liquidados del mes
    anticipos_liquidados_mes = AnticipoProyecto.objects.filter(
        proyecto=proyecto,
        fecha_liquidacion__month=mes,
        fecha_liquidacion__year=año
    ).order_by('fecha_liquidacion')
    
    # Calcular días del mes
    import calendar
    cal = calendar.monthcalendar(año, mes)
    nombre_mes = calendar.month_name[mes]
    
    # Crear eventos del calendario
    eventos_calendario = {}
    
    # Agregar días de pago (quincenales: día 15 y último día del mes)
    ultimo_dia = calendar.monthrange(año, mes)[1]
    total_salarios = sum(c.salario or 0 for c in colaboradores)
    pago_quincenal = total_salarios / 2
    
    # Primera quincena (día 15)
    eventos_calendario[15] = {
        'tipo': 'pago_salario',
        'titulo': '1ra Quincena',
        'descripcion': f'Pago primera quincena - Total: ${pago_quincenal:.2f}',
        'color': 'success',
        'icono': 'fas fa-money-bill-wave'
    }
    
    # Segunda quincena (último día del mes)
    eventos_calendario[ultimo_dia] = {
        'tipo': 'pago_salario',
        'titulo': '2da Quincena',
        'descripcion': f'Pago segunda quincena - Total: ${pago_quincenal:.2f}',
        'color': 'success',
        'icono': 'fas fa-money-bill-wave'
    }
    
    # Agregar anticipos realizados
    for anticipo in anticipos_mes:
        dia = anticipo.fecha_anticipo.day
        if dia not in eventos_calendario:
            eventos_calendario[dia] = []
        elif not isinstance(eventos_calendario[dia], list):
            eventos_calendario[dia] = [eventos_calendario[dia]]
        
        eventos_calendario[dia].append({
            'tipo': 'anticipo_realizado',
            'titulo': f'Anticipo: {anticipo.colaborador.nombre}',
            'descripcion': f'${anticipo.monto} - {anticipo.concepto}',
            'color': 'warning',
            'icono': 'fas fa-hand-holding-usd',
            'anticipo_id': anticipo.id
        })
    
    # Agregar anticipos liquidados
    for anticipo in anticipos_liquidados_mes:
        dia = anticipo.fecha_liquidacion.day
        if dia not in eventos_calendario:
            eventos_calendario[dia] = []
        elif not isinstance(eventos_calendario[dia], list):
            eventos_calendario[dia] = [eventos_calendario[dia]]
        
        eventos_calendario[dia].append({
            'tipo': 'anticipo_liquidado',
            'titulo': f'Liquidado: {anticipo.colaborador.nombre}',
            'descripcion': f'${anticipo.monto} liquidado por {anticipo.liquidado_por.username if anticipo.liquidado_por else "Sistema"}',
            'color': 'success',
            'icono': 'fas fa-check-circle',
            'anticipo_id': anticipo.id
        })
    
    # Estadísticas del mes
    total_anticipos_mes = anticipos_mes.aggregate(total=Sum('monto'))['total'] or 0
    total_liquidado_mes = anticipos_liquidados_mes.aggregate(total=Sum('monto'))['total'] or 0
    saldo_pendiente_mes = total_anticipos_mes - total_liquidado_mes
    
    context = {
        'proyecto': proyecto,
        'colaboradores': colaboradores,
        'calendario': cal,
        'nombre_mes': nombre_mes,
        'mes': mes,
        'año': año,
        'eventos_calendario': eventos_calendario,
        'anticipos_mes': anticipos_mes,
        'anticipos_liquidados_mes': anticipos_liquidados_mes,
        'total_anticipos_mes': total_anticipos_mes,
        'total_liquidado_mes': total_liquidado_mes,
        'saldo_pendiente_mes': saldo_pendiente_mes,
        'meses': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'años': range(año-2, año+3)
    }
    
    return render(request, 'core/proyectos/calendario_pagos.html', context)


@login_required
def administrar_anticipos_proyecto(request, proyecto_id):
    """Administrar anticipos del proyecto (editar, eliminar, cambiar estado)"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener todos los anticipos del proyecto
    anticipos = AnticipoProyecto.objects.filter(proyecto=proyecto).select_related('colaborador').order_by('-fecha_anticipo')
    
    # Estadísticas de anticipos
    total_anticipos = anticipos.aggregate(total=Sum('monto'))['total'] or 0
    anticipos_pendientes = anticipos.filter(estado='pendiente').aggregate(total=Sum('monto'))['total'] or 0
    anticipos_liquidados = anticipos.filter(estado='liquidado').aggregate(total=Sum('monto'))['total'] or 0
    anticipos_cancelados = anticipos.filter(estado='cancelado').aggregate(total=Sum('monto'))['total'] or 0
    
    context = {
        'proyecto': proyecto,
        'anticipos': anticipos,
        'total_anticipos': total_anticipos,
        'anticipos_pendientes': anticipos_pendientes,
        'anticipos_liquidados': anticipos_liquidados,
        'anticipos_cancelados': anticipos_cancelados,
    }
    
    return render(request, 'core/proyectos/administrar_anticipos.html', context)


@login_required
def editar_anticipo(request, anticipo_id):
    """Editar un anticipo existente"""
    anticipo = get_object_or_404(AnticipoProyecto, id=anticipo_id)
    
    if request.method == 'POST':
        monto = request.POST.get('monto')
        concepto = request.POST.get('concepto')
        estado = request.POST.get('estado')
        
        if monto and concepto:
            try:
                anticipo.monto = Decimal(monto)
                anticipo.concepto = concepto
                anticipo.estado = estado
                anticipo.save()
                
                messages.success(request, f'Anticipo de {anticipo.colaborador.nombre} actualizado exitosamente')
                return redirect('administrar_anticipos_proyecto', proyecto_id=anticipo.proyecto.id)
            except ValueError:
                messages.error(request, 'El monto debe ser un número válido')
        else:
            messages.error(request, 'Todos los campos son obligatorios')
    
    context = {
        'anticipo': anticipo,
        'estados_choices': AnticipoProyecto.ESTADO_CHOICES,
    }
    
    return render(request, 'core/proyectos/editar_anticipo.html', context)


@login_required
def eliminar_anticipo(request, anticipo_id):
    """Eliminar un anticipo"""
    anticipo = get_object_or_404(Anticipo, id=anticipo_id)
    proyecto_id = anticipo.proyecto.id
    
    if request.method == 'POST':
        numero_anticipo = anticipo.numero_anticipo
        anticipo.delete()
        messages.success(request, f'Anticipo {numero_anticipo} eliminado exitosamente')
        return redirect('anticipos_list')
    
    context = {
        'anticipo': anticipo,
    }
    
    return render(request, 'core/anticipos/delete.html', context)


@login_required
def cambiar_estado_anticipo(request, anticipo_id):
    """Cambiar el estado de un anticipo"""
    anticipo = get_object_or_404(Anticipo, id=anticipo_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('nuevo_estado')
        if nuevo_estado in ['pendiente', 'aplicado', 'devuelto', 'cancelado']:
            anticipo.estado = nuevo_estado
            anticipo.save()
            
            estado_display = dict(Anticipo.ESTADO_CHOICES)[nuevo_estado]
            messages.success(request, f'Estado del anticipo {anticipo.numero_anticipo} cambiado a {estado_display}')
            return redirect('anticipos_list')
        else:
            messages.error(request, 'Estado no válido')
    
    context = {
        'anticipo': anticipo,
        'estados_choices': Anticipo.ESTADO_CHOICES,
    }
    
    return render(request, 'core/proyectos/cambiar_estado_anticipo.html', context)


@login_required
def carpeta_create(request, proyecto_id):
    """Crear nueva carpeta en un proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id, activo=True)
    
    if request.method == 'POST':
        form = CarpetaProyectoForm(request.POST, proyecto=proyecto)
        if form.is_valid():
            carpeta = form.save(commit=False)
            carpeta.proyecto = proyecto
            carpeta.creada_por = request.user
            carpeta.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear Carpeta',
                modulo='Archivos',
                descripcion=f'Carpeta creada: {carpeta.nombre} en el proyecto {proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Carpeta creada exitosamente')
            return redirect('archivos_proyecto_list', proyecto_id=proyecto.id)
    else:
        form = CarpetaProyectoForm(proyecto=proyecto)
    
    context = {
        'form': form,
        'proyecto': proyecto,
        'accion': 'Crear',
        'titulo': 'Nueva Carpeta'
    }
    
    return render(request, 'core/archivos/carpeta_form.html', context)


@login_required
def carpeta_edit(request, carpeta_id):
    """Editar una carpeta existente"""
    carpeta = get_object_or_404(CarpetaProyecto, id=carpeta_id, activa=True)
    
    if request.method == 'POST':
        form = CarpetaProyectoForm(request.POST, instance=carpeta, proyecto=carpeta.proyecto, carpeta_actual=carpeta)
        if form.is_valid():
            carpeta = form.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar Carpeta',
                modulo='Archivos',
                descripcion=f'Carpeta editada: {carpeta.nombre} en el proyecto {carpeta.proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Carpeta actualizada exitosamente')
            return redirect('archivos_proyecto_list', proyecto_id=carpeta.proyecto.id)
    else:
        form = CarpetaProyectoForm(instance=carpeta, proyecto=carpeta.proyecto, carpeta_actual=carpeta)
    
    context = {
        'form': form,
        'carpeta': carpeta,
        'proyecto': carpeta.proyecto,
        'accion': 'Editar',
        'titulo': 'Editar Carpeta'
    }
    
    return render(request, 'core/archivos/carpeta_form.html', context)


@login_required
def carpeta_delete(request, carpeta_id):
    """Eliminar una carpeta"""
    carpeta = get_object_or_404(CarpetaProyecto, id=carpeta_id, activa=True)
    
    if request.method == 'POST':
        # Verificar que la carpeta se pueda eliminar
        if not carpeta.puede_eliminarse():
            messages.error(request, 'No se puede eliminar la carpeta porque contiene archivos o subcarpetas')
            return redirect('archivos_proyecto_list', proyecto_id=carpeta.proyecto.id)
        
        # Registrar actividad antes de eliminar
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar Carpeta',
            modulo='Archivos',
            descripcion=f'Carpeta eliminada: {carpeta.nombre} del proyecto {carpeta.proyecto.nombre}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        carpeta.delete()
        messages.success(request, 'Carpeta eliminada exitosamente')
        return redirect('archivos_proyecto_list', proyecto_id=carpeta.proyecto.id)
    
    return render(request, 'core/archivos/carpeta_delete.html', {'carpeta': carpeta})


@login_required
def carpeta_detail(request, carpeta_id):
    """Ver detalles de una carpeta y su contenido"""
    carpeta = get_object_or_404(CarpetaProyecto, id=carpeta_id, activa=True)
    
    # Obtener archivos en esta carpeta
    archivos = ArchivoProyecto.objects.filter(carpeta=carpeta, activo=True)
    
    # Obtener subcarpetas activas
    subcarpetas = carpeta.get_subcarpetas_activas()
    
    # Filtros
    tipo = request.GET.get('tipo')
    if tipo:
        archivos = archivos.filter(tipo=tipo)
    
    context = {
        'carpeta': carpeta,
        'proyecto': carpeta.proyecto,
        'archivos': archivos,
        'subcarpetas': subcarpetas,
        'tipos': ArchivoProyecto.TIPO_CHOICES,
        'total_archivos': archivos.count(),
        'total_subcarpetas': subcarpetas.count(),
    }
    
    return render(request, 'core/archivos/carpeta_detail.html', context)

# ===== REPORTES DE FACTURAS =====

@login_required
def facturas_reporte_lista(request):
    """Vista para generar reportes de facturas"""
    # Obtener filtros de la URL
    estado = request.GET.get('estado', '')
    cliente_id = request.GET.get('cliente', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    tipo_reporte = request.GET.get('tipo', 'lista')
    
    # Filtrar facturas
    facturas = Factura.objects.select_related('cliente', 'proyecto').all()
    
    if estado:
        facturas = facturas.filter(estado=estado)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if fecha_inicio:
        facturas = facturas.filter(fecha_emision__gte=fecha_inicio)
    if fecha_fin:
        facturas = facturas.filter(fecha_emision__lte=fecha_fin)
    
    # Ordenar por fecha de emisión (más recientes primero)
    facturas = facturas.order_by('-fecha_emision')
    
    # Calcular estadísticas
    total_facturado = facturas.aggregate(total=Sum('monto_total'))['total'] or 0
    total_cobrado = facturas.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_pendiente = total_facturado - total_cobrado
    
    # Estadísticas por estado
    stats_por_estado = facturas.values('estado').annotate(
        cantidad=Count('id'),
        monto_total=Sum('monto_total')
    ).order_by('estado')
    
    # Obtener clientes para el filtro
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    
    context = {
        'facturas': facturas,
        'clientes': clientes,
        'total_facturado': total_facturado,
        'total_cobrado': total_cobrado,
        'total_pendiente': total_pendiente,
        'stats_por_estado': stats_por_estado,
        'filtros': {
            'estado': estado,
            'cliente_id': cliente_id,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'tipo_reporte': tipo_reporte
        }
    }
    
    return render(request, 'core/facturas/reportes/lista.html', context)


@login_required
def facturas_reporte_pdf(request):
    """Generar reporte PDF de facturas"""
    # Obtener filtros
    estado = request.GET.get('estado', '')
    cliente_id = request.GET.get('cliente', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Filtrar facturas
    facturas = Factura.objects.select_related('cliente', 'proyecto').all()
    
    if estado:
        facturas = facturas.filter(estado=estado)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if fecha_inicio:
        facturas = facturas.filter(fecha_emision__gte=fecha_inicio)
    if fecha_fin:
        facturas = facturas.filter(fecha_emision__lte=fecha_fin)
    
    facturas = facturas.order_by('-fecha_emision')
    
    # Calcular totales
    total_facturado = facturas.aggregate(total=Sum('monto_total'))['total'] or 0
    total_cobrado = facturas.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_pendiente = total_facturado - total_cobrado
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    # Título
    elements.append(Paragraph("REPORTE DE FACTURAS", title_style))
    
    # Información del reporte
    filtros_info = []
    if estado:
        filtros_info.append(f"Estado: {dict(Factura.ESTADO_CHOICES).get(estado, estado)}")
    if cliente_id:
        cliente = Cliente.objects.get(id=cliente_id)
        filtros_info.append(f"Cliente: {cliente.razon_social}")
    if fecha_inicio:
        filtros_info.append(f"Desde: {fecha_inicio}")
    if fecha_fin:
        filtros_info.append(f"Hasta: {fecha_fin}")
    
    if filtros_info:
        filtros_text = " | ".join(filtros_info)
        elements.append(Paragraph(filtros_text, subtitle_style))
    
    # Resumen financiero
    resumen_data = [
        ['CONCEPTO', 'MONTO'],
        ['Total Facturado', f"${total_facturado:,.2f}"],
        ['Total Cobrado', f"${total_cobrado:,.2f}"],
        ['Total Pendiente', f"${total_pendiente:,.2f}"],
        ['Cantidad Facturas', str(facturas.count())],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (0, -1), 10),
        ('BOTTOMPADDING', (0, 0), (0, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.white),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))
    
    # Tabla de facturas
    if facturas:
        headers = ['Número', 'Cliente', 'Proyecto', 'Monto', 'Estado', 'Emisión', 'Vencimiento']
        data = [headers]
        
        for factura in facturas:
            data.append([
                factura.numero_factura,
                factura.cliente.razon_social if factura.cliente else 'N/A',
                factura.proyecto.nombre if factura.proyecto else 'N/A',
                f"${factura.monto_total:,.2f}",
                factura.get_estado_display(),
                factura.fecha_emision.strftime('%d/%m/%Y') if factura.fecha_emision else 'N/A',
                factura.fecha_vencimiento.strftime('%d/%m/%Y') if factura.fecha_vencimiento else 'N/A',
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Alinear montos a la derecha
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
    
    # Pie de página
    elements.append(Spacer(1, 30))
    fecha_generacion = timezone.now().strftime('%d/%m/%Y %H:%M')
    elements.append(Paragraph(f"Reporte generado el: {fecha_generacion}", styles['Normal']))
    elements.append(Paragraph(f"Sistema ARCA Construcción", styles['Normal']))
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta HTTP
    filename = f"reporte_facturas_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def facturas_reporte_excel(request):
    """Generar reporte Excel de facturas"""
    import csv
    
    # Obtener filtros
    estado = request.GET.get('estado', '')
    cliente_id = request.GET.get('cliente', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Filtrar facturas
    facturas = Factura.objects.select_related('cliente', 'proyecto').all()
    
    if estado:
        facturas = facturas.filter(estado=estado)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if fecha_inicio:
        facturas = facturas.filter(fecha_emision__gte=fecha_inicio)
    if fecha_fin:
        facturas = facturas.filter(fecha_emision__lte=fecha_fin)
    
    facturas = facturas.order_by('-fecha_emision')
    
    # Crear respuesta CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="reporte_facturas_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    
    # Crear writer CSV
    writer = csv.writer(response)
    
    # Escribir encabezados
    writer.writerow([
        'Número Factura',
        'Cliente',
        'Proyecto',
        'Monto Total',
        'Monto Pagado',
        'Monto Pendiente',
        'Estado',
        'Fecha Emisión',
        'Fecha Vencimiento',
        'Fecha Creación'
    ])
    
    # Escribir datos
    for factura in facturas:
        writer.writerow([
            factura.numero_factura,
            factura.cliente.razon_social if factura.cliente else 'N/A',
            factura.proyecto.nombre if factura.proyecto else 'N/A',
            factura.monto_total,
            factura.monto_pagado,
            factura.monto_pendiente,
            factura.get_estado_display(),
            factura.fecha_emision.strftime('%d/%m/%Y') if factura.fecha_emision else 'N/A',
            factura.fecha_vencimiento.strftime('%d/%m/%Y') if factura.fecha_vencimiento else 'N/A',
            factura.fecha_creacion.strftime('%d/%m/%Y %H:%M') if factura.fecha_creacion else 'N/A',
        ])
    
    return response


@login_required
def facturas_reporte_detallado(request):
    """Reporte detallado de facturas con análisis"""
    # Obtener filtros
    estado = request.GET.get('estado', '')
    cliente_id = request.GET.get('cliente', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Filtrar facturas
    facturas = Factura.objects.select_related('cliente', 'proyecto').all()
    
    if estado:
        facturas = facturas.filter(estado=estado)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if fecha_inicio:
        facturas = facturas.filter(fecha_emision__gte=fecha_inicio)
    if fecha_fin:
        facturas = facturas.filter(fecha_emision__lte=fecha_fin)
    
    # Calcular estadísticas
    total_facturado = facturas.aggregate(total=Sum('monto_total'))['total'] or 0
    total_cobrado = facturas.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_pendiente = total_facturado - total_cobrado
    
    # Estadísticas por estado
    stats_por_estado = []
    for estado in Factura.ESTADO_CHOICES:
        estado_key = estado[0]
        facturas_estado = facturas.filter(estado=estado_key)
        cantidad = facturas_estado.count()
        monto_total = facturas_estado.aggregate(total=Sum('monto_total'))['total'] or 0
        monto_pagado = facturas_estado.aggregate(total=Sum('monto_pagado'))['total'] or 0
        
        if cantidad > 0:
            porcentaje = (monto_pagado / monto_total * 100) if monto_total > 0 else 0
            stats_por_estado.append({
                'estado': estado_key,
                'estado_display': estado[1],
                'cantidad': cantidad,
                'monto_total': monto_total,
                'monto_pagado': monto_pagado,
                'porcentaje': round(porcentaje, 1)
            })
    
    # Estadísticas por cliente
    stats_por_cliente = []
    clientes_con_facturas = facturas.values('cliente__razon_social', 'cliente__id').distinct()
    
    for cliente_data in clientes_con_facturas:
        cliente_id = cliente_data['cliente__id']
        cliente_nombre = cliente_data['cliente__razon_social']
        
        facturas_cliente = facturas.filter(cliente_id=cliente_id)
        cantidad = facturas_cliente.count()
        monto_total = facturas_cliente.aggregate(total=Sum('monto_total'))['total'] or 0
        monto_pagado = facturas_cliente.aggregate(total=Sum('monto_pagado'))['total'] or 0
        
        # Obtener el presupuesto del proyecto más reciente del cliente
        proyecto_mas_reciente = Proyecto.objects.filter(cliente_id=cliente_id).order_by('-creado_en').first()
        # presupuesto_proyecto ELIMINADO - YA NO SE USA
        presupuesto_proyecto = 0
        
        # Calcular porcentaje basado en el presupuesto del proyecto
        if presupuesto_proyecto and presupuesto_proyecto > 0:
            porcentaje = (monto_pagado / presupuesto_proyecto * 100)
        else:
            # Si no hay presupuesto, calcular basado en el monto total facturado
            porcentaje = (monto_pagado / monto_total * 100) if monto_total > 0 else 0
        
        stats_por_cliente.append({
            'cliente_nombre': cliente_nombre,
            'cliente_id': cliente_id,
            'cantidad': cantidad,
            'monto_total': monto_total,
            'monto_pagado': monto_pagado,
            'presupuesto_proyecto': presupuesto_proyecto,
            'porcentaje': round(porcentaje, 1)
        })
    
    # Ordenar por monto total descendente
    stats_por_cliente.sort(key=lambda x: x['monto_total'], reverse=True)
    
    # Facturas vencidas
    facturas_vencidas = facturas.filter(
        estado__in=['emitida', 'enviada'],
        fecha_vencimiento__lt=timezone.now().date()
    )
    
    # Total de facturas vencidas
    total_vencidas = facturas_vencidas.aggregate(total=Sum('monto_total'))['total'] or 0
    
    # Obtener clientes para el filtro
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    
    context = {
        'facturas': facturas,
        'clientes': clientes,
        'total_facturado': total_facturado,
        'total_cobrado': total_cobrado,
        'total_pendiente': total_pendiente,
        'total_vencidas': total_vencidas,
        'stats_por_estado': stats_por_estado,
        'stats_por_cliente': stats_por_cliente,
        'facturas_vencidas': facturas_vencidas,
        'filtros': {
            'estado': estado,
            'cliente_id': cliente_id,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
    }
    
    return render(request, 'core/facturas/reportes/detallado.html', context)


# ==================== TRABAJADORES DIARIOS ====================

@login_required
def trabajadores_diarios_list(request, proyecto_id):
    """Lista de trabajadores diarios de un proyecto - SIGUIENDO DOCUMENTACIÓN COMPLETA"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # IMPORTANTE: Obtener SOLO las planillas ACTIVAS para el selector (máximo 2)
    planillas_activas = PlanillaTrabajadoresDiarios.objects.filter(
        proyecto=proyecto,
        estado__in=['activa', 'pendiente']
    ).order_by('-fecha_creacion').distinct()[:2]
    
    # Obtener todas las planillas para histórico
    planillas_todas = PlanillaTrabajadoresDiarios.objects.filter(
        proyecto=proyecto
    ).order_by('-fecha_creacion')
    
    # Obtener planilla seleccionada desde la URL
    planilla_id = request.GET.get('planilla_id')
    planilla_seleccionada = None
    if planilla_id:
        try:
            planilla_seleccionada = PlanillaTrabajadoresDiarios.objects.get(
                id=planilla_id, 
                proyecto=proyecto
            )
        except PlanillaTrabajadoresDiarios.DoesNotExist:
            pass
    
    # Si no hay planilla seleccionada, usar la primera activa
    if not planilla_seleccionada:
        planilla_seleccionada = planillas_activas.first()
    
    # Filtrar trabajadores: Mostrar TODOS los trabajadores activos del proyecto
    # Si hay una planilla seleccionada, se puede usar para filtrar opcionalmente
    # pero por defecto mostrar todos los trabajadores activos del proyecto
    if planilla_seleccionada and planilla_seleccionada.estado == 'finalizada':
        # Si está finalizada, mostrar trabajadores de esa planilla (para historial)
        trabajadores = TrabajadorDiario.objects.filter(
            proyecto=proyecto, 
            planilla=planilla_seleccionada
        ).order_by('nombre')
    else:
        # Mostrar TODOS los trabajadores activos del proyecto (no solo los de la planilla seleccionada)
        # Si hay planilla seleccionada, mostrar trabajadores de esa planilla activos
        # Si no hay planilla seleccionada, mostrar todos los trabajadores activos del proyecto
        if planilla_seleccionada:
            # Mostrar TODOS los trabajadores de la planilla seleccionada (activos e inactivos)
            # para que se vean todos los trabajadores de esa planilla
            trabajadores = TrabajadorDiario.objects.filter(
                proyecto=proyecto,
                planilla=planilla_seleccionada
            ).order_by('nombre', 'id')
        else:
            # Mostrar TODOS los trabajadores activos del proyecto
            trabajadores = TrabajadorDiario.objects.filter(
                proyecto=proyecto, 
                activo=True
            ).order_by('nombre', 'id')
    
    # Calcular anticipos aplicados y total bruto para cada trabajador
    for trabajador in trabajadores:
        anticipos_qs = AnticipoTrabajadorDiario.objects.filter(
            trabajador=trabajador,
            estado='aplicado'
        )
        total_anticipos = sum(anticipo.monto for anticipo in anticipos_qs)
        trabajador.anticipos_monto = Decimal(str(total_anticipos))
        trabajador.anticipos_count = anticipos_qs.count()
        
        # Calcular total bruto (días trabajados * pago diario)
        dias_trabajados = Decimal(str(trabajador.total_dias_trabajados)) if trabajador.total_dias_trabajados else Decimal('0')
        pago_diario = Decimal(str(trabajador.pago_diario))
        trabajador.total_bruto = dias_trabajados * pago_diario
        
        # Calcular total a pagar (total bruto - anticipos)
        trabajador.total_a_pagar_calculado = trabajador.total_bruto - trabajador.anticipos_monto
    
    # Calcular totales generales según documentación
    total_bruto_general = sum(t.total_dias_trabajados * t.pago_diario for t in trabajadores)
    total_anticipos_general = sum(t.anticipos_monto for t in trabajadores)
    total_neto_general = total_bruto_general - total_anticipos_general
    
    # Contar trabajadores liquidados
    trabajadores_liquidados_count = trabajadores.filter(activo=False).count()
    
    # Calcular histórico de planillas finalizadas
    planillas_liquidadas = PlanillaLiquidada.objects.filter(
        proyecto=proyecto,
        observaciones__icontains='trabajadores diarios'
    )
    total_planillas_finalizadas = planillas_liquidadas.count()
    total_historico_gastos = planillas_liquidadas.aggregate(total=Sum('total_planilla'))['total'] or Decimal('0.00')
    promedio_por_planilla = total_historico_gastos / total_planillas_finalizadas if total_planillas_finalizadas > 0 else Decimal('0.00')
    
    context = {
        'proyecto': proyecto,
        'trabajadores': trabajadores,
        'planillas_activas': planillas_activas,  # Máximo 2 para selector
        'planillas_todas': planillas_todas,  # Todas para histórico
        'planilla_seleccionada': planilla_seleccionada,
        'total_bruto_general': total_bruto_general,
        'total_anticipos_general': total_anticipos_general,
        'total_neto_general': total_neto_general,
        'total_trabajadores': trabajadores.count(),
        'trabajadores_liquidados_count': trabajadores_liquidados_count,
        'total_planillas_finalizadas': total_planillas_finalizadas,
        'total_historico_gastos': total_historico_gastos,
        'promedio_por_planilla': promedio_por_planilla,
    }
    
    return render(request, 'core/trabajadores_diarios/list.html', context)


@login_required
def reactivar_trabajador_diario(request, proyecto_id, trabajador_id):
    """Reactivar un trabajador diario liquidado"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    
    if request.method == 'POST':
        try:
            # Reactivar el trabajador
            trabajador.activo = True
            trabajador.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Reactivar Trabajador Diario',
                modulo='Trabajadores Diarios',
                descripcion=f'Trabajador {trabajador.nombre} reactivado en el proyecto {proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'✅ Trabajador "{trabajador.nombre}" reactivado exitosamente. Ahora puedes editarlo y pagarle nuevamente.')
            
        except Exception as e:
            messages.error(request, f'❌ Error al reactivar trabajador: {str(e)}')
    
    # Redirigir de vuelta a la lista con la planilla seleccionada
    planilla_id = request.GET.get('planilla_id')
    if planilla_id:
        return redirect(f'{reverse("trabajadores_diarios_list", args=[proyecto_id])}?planilla_id={planilla_id}')
    else:
        return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)

@login_required
def reactivar_todos_trabajadores_diarios(request, proyecto_id):
    """Reactivar todos los trabajadores diarios liquidados de una planilla"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        try:
            # Obtener planilla seleccionada
            planilla_id = request.POST.get('planilla_id')
            if planilla_id:
                planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
                trabajadores = TrabajadorDiario.objects.filter(proyecto=proyecto, planilla=planilla, activo=False)
            else:
                trabajadores = TrabajadorDiario.objects.filter(proyecto=proyecto, activo=False)
            
            # Reactivar todos los trabajadores liquidados
            trabajadores_reactivados = trabajadores.update(activo=True)
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Reactivar Todos los Trabajadores Diarios',
                modulo='Trabajadores Diarios',
                descripcion=f'{trabajadores_reactivados} trabajadores reactivados en el proyecto {proyecto.nombre}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'✅ {trabajadores_reactivados} trabajadores reactivados exitosamente. Ahora puedes editarlos y pagarles nuevamente.')
            
        except Exception as e:
            messages.error(request, f'❌ Error al reactivar trabajadores: {str(e)}')
    
    # Redirigir de vuelta a la lista con la planilla seleccionada
    planilla_id = request.POST.get('planilla_id') or request.GET.get('planilla_id')
    if planilla_id:
        return redirect(f'{reverse("trabajadores_diarios_list", args=[proyecto_id])}?planilla_id={planilla_id}')
    else:
        return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)

@login_required
def finalizar_planilla_trabajadores(request, proyecto_id):
    """Finalizar planilla de trabajadores diarios: generar PDF, guardarlo y limpiar lista"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    try:
        # Verificar que hay trabajadores para finalizar
        trabajadores = TrabajadorDiario.objects.filter(proyecto=proyecto, activo=True)
        if not trabajadores.exists():
            messages.warning(request, 'No hay trabajadores activos para finalizar la planilla.')
            return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)
        
        # 1. Generar PDF de la planilla usando la misma lógica que trabajadores_diarios_pdf
        from django.core.files.base import ContentFile
        from django.utils import timezone
        from io import BytesIO
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
        import os
        
        # Crear el buffer para el PDF
        buffer = BytesIO()
        
        # Crear el documento PDF en orientación horizontal
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Obtener estilos
        styles = getSampleStyleSheet()
        
        # Crear estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkgreen
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=12
        )
        
        # Contenido del PDF
        story = []
        
        # Título principal
        story.append(Paragraph("PLANILLA DE TRABAJADORES DIARIOS", title_style))
        story.append(Spacer(1, 12))
        
        # Información del proyecto
        story.append(Paragraph(f"<b>Proyecto:</b> {proyecto.nombre}", normal_style))
        story.append(Paragraph(f"<b>Cliente:</b> {proyecto.cliente.razon_social}", normal_style))
        story.append(Paragraph(f"<b>Fecha de Generación:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
        story.append(Spacer(1, 20))
        
        # Calcular totales con anticipos
        total_trabajadores = trabajadores.count()
        total_bruto_general = 0
        total_anticipos_general = 0
        total_neto_general = 0
        
        # Crear tabla con columnas de anticipos
        data = [['No.', 'Nombre del Trabajador', 'Pago Diario', 'Días Trabajados', 'Total Bruto', 'Anticipos', 'Total Neto']]
        
        for i, trabajador in enumerate(trabajadores, 1):
            dias_trabajados = sum(registro.dias_trabajados for registro in trabajador.registros_trabajo.all())
            if dias_trabajados == 0:
                dias_trabajados = 1  # Valor por defecto si no hay registros
            
            total_bruto = float(trabajador.pago_diario) * dias_trabajados
            
            # Calcular anticipos del trabajador
            from core.models import AnticipoTrabajadorDiario
            anticipos_trabajador = AnticipoTrabajadorDiario.objects.filter(
                trabajador=trabajador,
                estado='aplicado'
            )
            total_anticipos_trabajador = sum(anticipo.monto_aplicado for anticipo in anticipos_trabajador)
            
            # Total neto = Total bruto - Anticipos
            total_neto = total_bruto - float(total_anticipos_trabajador)
            
            total_bruto_general += total_bruto
            total_anticipos_general += float(total_anticipos_trabajador)
            total_neto_general += total_neto
            
            data.append([
                str(i),
                trabajador.nombre,
                f"${trabajador.pago_diario:.2f}",
                str(dias_trabajados),
                f"${total_bruto:.2f}",
                f"${total_anticipos_trabajador:.2f}",
                f"${total_neto:.2f}"
            ])
        
        # Agregar fila de totales
        data.append(['', '', '', 'TOTAL GENERAL:', f"${total_bruto_general:.2f}", f"${total_anticipos_general:.2f}", f"${total_neto_general:.2f}"])
        
        # Crear la tabla con columnas adicionales (7 columnas total)
        table = Table(data, colWidths=[0.6*inch, 2.5*inch, 1.2*inch, 1.2*inch, 1.4*inch, 1.4*inch, 1.4*inch])
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 30))
        
        # Información adicional
        story.append(Paragraph(f"<b>Total de Trabajadores:</b> {total_trabajadores}", normal_style))
        story.append(Paragraph(f"<b>Total Bruto a Pagar:</b> ${total_bruto_general:.2f}", normal_style))
        story.append(Paragraph(f"<b>Total Anticipos Aplicados:</b> ${total_anticipos_general:.2f}", normal_style))
        story.append(Paragraph(f"<b>Total Neto a Pagar:</b> ${total_neto_general:.2f}", normal_style))
        
        # Construir el PDF
        doc.build(story)
        
        # Obtener el contenido del buffer
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # 2. Crear o obtener carpeta "Trabajadores Diarios" en archivos del proyecto
        from core.models import CarpetaProyecto
        
        carpeta, created = CarpetaProyecto.objects.get_or_create(
            proyecto=proyecto,
            nombre='Trabajadores Diarios',
            defaults={
                'creada_por': request.user,
                'descripcion': 'Carpeta para almacenar planillas de trabajadores diarios'
            }
        )
        
        if created:
            print(f"✅ Carpeta 'Trabajadores Diarios' creada para proyecto {proyecto.nombre}")
        else:
            print(f"✅ Carpeta 'Trabajadores Diarios' ya existe para proyecto {proyecto.nombre}")
        
        # 3. Guardar archivo PDF en los archivos del proyecto
        from core.models import ArchivoProyecto
        
        nombre_archivo = f"planilla_trabajadores_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Crear ContentFile con el contenido PDF
        archivo_pdf = ContentFile(pdf_content)
        archivo_pdf.name = nombre_archivo
        
        archivo = ArchivoProyecto.objects.create(
            proyecto=proyecto,
            carpeta=carpeta,
            nombre=nombre_archivo,
            archivo=archivo_pdf,
            descripcion=f'Planilla de trabajadores diarios finalizada el {timezone.now().strftime("%d/%m/%Y %H:%M")}',
            subido_por=request.user,
            activo=True
        )
        
        print(f"✅ PDF guardado: {nombre_archivo}")
        print(f"✅ Archivo ID: {archivo.id}")
        print(f"✅ Tiene archivo físico: {bool(archivo.archivo)}")
        if archivo.archivo:
            print(f"✅ Tamaño: {archivo.archivo.size} bytes")
            print(f"✅ Ruta: {archivo.archivo.path}")
            print(f"✅ Existe archivo: {os.path.exists(archivo.archivo.path)}")
        
        # 4. Crear registro de planilla liquidada
        from core.models import PlanillaLiquidada
        planilla_liquidada = PlanillaLiquidada.objects.create(
            proyecto=proyecto,
            total_salarios=Decimal(str(total_neto_general)),
            total_anticipos=Decimal(str(total_anticipos_general)),
            total_planilla=Decimal(str(total_neto_general)),
            cantidad_personal=trabajadores.count(),
            liquidada_por=request.user,
            observaciones=f'Planilla de trabajadores diarios finalizada - Total: ${total_neto_general:.2f}'
        )
        
        print(f"✅ Planilla liquidada creada: ID {planilla_liquidada.id}, Total: ${total_neto_general:.2f}")
        
        # 4.5. Actualizar total histórico del proyecto
        proyecto.total_diarios = (proyecto.total_diarios or Decimal('0.00')) + Decimal(str(total_neto_general))
        proyecto.save()
        print(f"✅ Total diarios actualizado: ${proyecto.total_diarios}")
        
        # 5. Limpiar lista de trabajadores (marcar como inactivos)
        trabajadores_eliminados = trabajadores.count()
        trabajadores.update(activo=False)
        
        print(f"✅ {trabajadores_eliminados} trabajadores marcados como inactivos")
        
        # 6. Registrar actividad
        from core.models import LogActividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Finalizar Planilla',
            modulo='Trabajadores Diarios',
            descripcion=f'Planilla de trabajadores finalizada para proyecto {proyecto.nombre}. Archivo guardado: {nombre_archivo}. Trabajadores procesados: {trabajadores_eliminados}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        # 7. Mostrar mensaje de éxito
        messages.success(request, f'Planilla finalizada exitosamente. Archivo guardado como "{nombre_archivo}". Se procesaron {trabajadores_eliminados} trabajadores.')
        
        return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)
        
    except Exception as e:
        print(f"❌ Error al finalizar planilla: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error al finalizar la planilla: {str(e)}')
        return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)


@login_required
def reabrir_planilla_trabajadores(request, proyecto_id, planilla_id):
    """Reabrir una planilla finalizada para continuar editando - SIGUIENDO DOCUMENTACIÓN"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
    
    if request.method != 'POST':
        messages.error(request, 'Método no permitido.')
        return redirect(f'{reverse("trabajadores_diarios_list", args=[proyecto_id])}?planilla_id={planilla.id}')
    
    if planilla.estado != 'finalizada':
        messages.info(request, f'La planilla "{planilla.nombre}" ya está activa.')
        return redirect(f'{reverse("trabajadores_diarios_list", args=[proyecto_id])}?planilla_id={planilla.id}')
    
    # Reabrir planilla
    planilla.estado = 'activa'
    planilla.fecha_finalizacion = None
    planilla.finalizada_por = None
    planilla.save()
    
    # Reactivar trabajadores
    trabajadores_reactivados = planilla.trabajadores.update(activo=True)
    
    LogActividad.objects.create(
        usuario=request.user,
        accion='Reabrir Planilla',
        modulo='Trabajadores Diarios',
        descripcion=f'Planilla "{planilla.nombre}" reabierta. Trabajadores reactivados: {trabajadores_reactivados}',
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    messages.success(request, 
        f'✅ Planilla "{planilla.nombre}" reabierta. Puedes continuar editando y registrando días.'
    )
    
    return redirect(f'{reverse("trabajadores_diarios_list", args=[proyecto_id])}?planilla_id={planilla.id}')


@login_required
def trabajador_diario_create(request, proyecto_id):
    """Crear trabajador diario - SIGUIENDO DOCUMENTACIÓN COMPLETA"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener planilla_id de la URL (GET) o del POST si existe
    planilla_id = request.GET.get('planilla_id') or request.POST.get('planilla_id')
    planilla_seleccionada = None
    if planilla_id:
        try:
            planilla_seleccionada = PlanillaTrabajadoresDiarios.objects.get(id=planilla_id, proyecto=proyecto)
        except PlanillaTrabajadoresDiarios.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = TrabajadorDiarioForm(
            request.POST, 
            planilla=planilla_seleccionada, 
            proyecto=proyecto
        )
        if form.is_valid():
            # Validar que no exista el mismo nombre en otras planillas ACTIVAS
            nombre_trabajador = form.cleaned_data.get('nombre')
            trabajadores_duplicados = TrabajadorDiario.objects.filter(
                proyecto=proyecto,
                nombre__iexact=nombre_trabajador,
                activo=True,
                planilla__estado__in=['activa', 'pendiente']
            ).exclude(planilla=planilla_seleccionada)
            
            if trabajadores_duplicados.exists():
                planilla_duplicada = trabajadores_duplicados.first().planilla
                messages.error(
                    request,
                    f'❌ El trabajador "{nombre_trabajador}" ya existe en la planilla activa "{planilla_duplicada.nombre}". '
                    f'No se permiten trabajadores duplicados entre planillas activas.'
                )
            else:
                trabajador = form.save(commit=False)
                trabajador.proyecto = proyecto
                if planilla_seleccionada:
                    trabajador.planilla = planilla_seleccionada
                trabajador.creado_por = request.user
                trabajador.save()
                
                # Refrescar la instancia de la planilla para asegurar que se actualice la relación
                if planilla_seleccionada:
                    planilla_seleccionada.refresh_from_db()
                
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Crear',
                    modulo='Trabajadores Diarios',
                    descripcion=f'Trabajador "{trabajador.nombre}" creado en planilla "{planilla_seleccionada.nombre if planilla_seleccionada else "sin planilla"}"',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, 
                    f'✅ Trabajador "{nombre_trabajador}" agregado a la planilla "{planilla_seleccionada.nombre if planilla_seleccionada else "sin planilla"}".'
                )
                # Si hay planilla seleccionada, redirigir a la vista de detalle de la planilla
                if planilla_seleccionada:
                    return redirect('planilla_trabajadores_diarios_detail', proyecto_id=proyecto_id, planilla_id=planilla_seleccionada.id)
                # Si no hay planilla, redirigir a la lista general
                return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = TrabajadorDiarioForm(planilla=planilla_seleccionada, proyecto=proyecto)
    
    return render(request, 'core/trabajadores_diarios/create.html', {
        'form': form,
        'proyecto': proyecto,
        'planilla_seleccionada': planilla_seleccionada
    })


@login_required
def trabajador_diario_detail(request, proyecto_id, trabajador_id):
    """Detalle de trabajador diario"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    registros = trabajador.registros_trabajo.all().order_by('-fecha_inicio')
    
    return render(request, 'core/trabajadores_diarios/detail.html', {
        'proyecto': proyecto,
        'trabajador': trabajador,
        'registros': registros
    })


@login_required
def trabajador_diario_edit(request, proyecto_id, trabajador_id):
    """Editar trabajador diario - SIGUIENDO DOCUMENTACIÓN: Preservar planilla original"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    
    planilla_id = request.GET.get('planilla_id')
    
    # IMPORTANTE: Preservar la planilla original
    planilla_original = trabajador.planilla
    
    if request.method == 'POST':
        form = TrabajadorDiarioForm(request.POST, instance=trabajador, proyecto=proyecto)
        if form.is_valid():
            trabajador_editado = form.save(commit=False)
            trabajador_editado.proyecto = proyecto
            
            # Preservar la planilla original (no debe cambiar)
            if planilla_original:
                trabajador_editado.planilla = planilla_original
            elif trabajador.planilla:
                trabajador_editado.planilla = trabajador.planilla
            
            trabajador_editado.save()
            
            messages.success(request, 
                f'✅ Trabajador "{trabajador_editado.nombre}" actualizado correctamente.'
            )
            
            planilla_redirect = planilla_id or (planilla_original.id if planilla_original else None)
            if planilla_redirect:
                return redirect(
                    f'{reverse("trabajadores_diarios_list", args=[proyecto_id])}?planilla_id={planilla_redirect}'
                )
            return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)
    else:
        form = TrabajadorDiarioForm(instance=trabajador, proyecto=proyecto)
    
    return render(request, 'core/trabajadores_diarios/edit.html', {
        'form': form,
        'proyecto': proyecto,
        'trabajador': trabajador,
        'planilla_id': planilla_id or (planilla_original.id if planilla_original else None)
    })


@login_required
def trabajador_diario_delete(request, proyecto_id, trabajador_id):
    """Eliminar trabajador diario"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    
    if request.method == 'POST':
        trabajador.delete()
        messages.success(request, 'Trabajador diario eliminado correctamente.')
        return redirect('trabajadores_diarios_list', proyecto_id=proyecto_id)
    
    return render(request, 'core/trabajadores_diarios/delete.html', {
        'proyecto': proyecto,
        'trabajador': trabajador
    })


@login_required
def registro_trabajo_create(request, proyecto_id, trabajador_id):
    """Crear registro de trabajo"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    
    if request.method == 'POST':
        form = RegistroTrabajoForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.trabajador = trabajador
            registro.registrado_por = request.user
            registro.save()
            
            messages.success(request, 'Registro de trabajo creado correctamente.')
            return redirect('trabajador_diario_detail', proyecto_id=proyecto_id, trabajador_id=trabajador_id)
    else:
        form = RegistroTrabajoForm()
    
    return render(request, 'core/trabajadores_diarios/registro_create.html', {
        'form': form,
        'proyecto': proyecto,
        'trabajador': trabajador
    })


@login_required
def registro_trabajo_edit(request, proyecto_id, trabajador_id, registro_id):
    """Editar registro de trabajo"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    registro = get_object_or_404(RegistroTrabajo, id=registro_id, trabajador=trabajador)
    
    if request.method == 'POST':
        form = RegistroTrabajoForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro de trabajo actualizado correctamente.')
            return redirect('trabajador_diario_detail', proyecto_id=proyecto_id, trabajador_id=trabajador_id)
    else:
        form = RegistroTrabajoForm(instance=registro)
    
    return render(request, 'core/trabajadores_diarios/registro_edit.html', {
        'form': form,
        'proyecto': proyecto,
        'trabajador': trabajador,
        'registro': registro
    })


@login_required
def registro_trabajo_delete(request, proyecto_id, trabajador_id, registro_id):
    """Eliminar registro de trabajo"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    registro = get_object_or_404(RegistroTrabajo, id=registro_id, trabajador=trabajador)
    
    if request.method == 'POST':
        registro.delete()
        messages.success(request, 'Registro de trabajo eliminado correctamente.')
        return redirect('trabajador_diario_detail', proyecto_id=proyecto_id, trabajador_id=trabajador_id)
    
    return render(request, 'core/trabajadores_diarios/registro_delete.html', {
        'proyecto': proyecto,
        'trabajador': trabajador,
        'registro': registro
    })


@login_required
@login_required
def actualizar_dias_trabajados(request, proyecto_id, trabajador_id):
    """Actualizar días trabajados de un trabajador (AJAX) - SIGUIENDO DOCUMENTACIÓN"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto_id=proyecto_id)
    
    try:
        dias_trabajados = int(request.POST.get('dias_trabajados', 0))
        
        if dias_trabajados < 0:
            return JsonResponse({'success': False, 'error': 'Los días no pueden ser negativos'})
        
        # Actualizar o crear registro de trabajo
        registro, created = RegistroTrabajo.objects.get_or_create(
            trabajador=trabajador,
            defaults={
                'fecha_inicio': timezone.now().date(),
                'fecha_fin': timezone.now().date(),
                'dias_trabajados': dias_trabajados,
                'registrado_por': request.user
            }
        )
        
        if not created:
            registro.dias_trabajados = dias_trabajados
            registro.fecha_fin = timezone.now().date()
            registro.save()
        
        # Calcular totales según documentación
        total_bruto = trabajador.total_dias_trabajados * trabajador.pago_diario
        total_anticipos = trabajador.total_anticipos_aplicados
        total_neto = total_bruto - total_anticipos
        
        return JsonResponse({
            'success': True,
            'total_bruto': float(total_bruto),
            'total_anticipos': float(total_anticipos),
            'total_neto': float(total_neto),
            'dias_trabajados': trabajador.total_dias_trabajados
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def trabajadores_diarios_pdf(request, proyecto_id):
    """Generar PDF de la planilla de trabajadores diarios"""
    print("🚀 INICIANDO GENERACIÓN DE PDF - VERSIÓN ACTUALIZADA")
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    # Obtener planilla seleccionada
    planilla_id = request.POST.get('planilla_id')
    planilla_seleccionada = None
    if planilla_id:
        try:
            planilla_seleccionada = PlanillaTrabajadoresDiarios.objects.get(id=planilla_id, proyecto=proyecto)
        except PlanillaTrabajadoresDiarios.DoesNotExist:
            pass
    
    # Filtrar trabajadores por planilla seleccionada
    if planilla_seleccionada:
        # Incluir tanto trabajadores activos como inactivos de la planilla seleccionada
        trabajadores = TrabajadorDiario.objects.filter(proyecto=proyecto, planilla=planilla_seleccionada).order_by('nombre')
    else:
        # Mostrar trabajadores sin planilla asignada (solo activos)
        trabajadores = TrabajadorDiario.objects.filter(proyecto=proyecto, planilla__isnull=True, activo=True).order_by('nombre')
    
    # Obtener días trabajados temporales del POST o usar datos de la base de datos
    dias_trabajados_data = {}
    if request.method == 'POST':
        for trabajador in trabajadores:
            dias_key = f'dias_trabajador_{trabajador.id}'
            dias_trabajados_data[trabajador.id] = int(request.POST.get(dias_key, 0))
    else:
        # Si no hay datos POST, usar datos de la base de datos
        for trabajador in trabajadores:
            dias_trabajados_data[trabajador.id] = sum(registro.dias_trabajados for registro in trabajador.registros_trabajo.all())
    
    # Crear el buffer para el PDF
    buffer = BytesIO()
    
    # Crear el documento PDF en orientación horizontal
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Obtener estilos
    styles = getSampleStyleSheet()
    
    # Crear estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.darkgreen
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=12
    )
    
    # Contenido del PDF
    story = []
    
    # Título principal
    if planilla_seleccionada:
        story.append(Paragraph(f"PLANILLA DE TRABAJADORES DIARIOS - {planilla_seleccionada.nombre.upper()}", title_style))
    else:
        story.append(Paragraph("PLANILLA DE TRABAJADORES DIARIOS", title_style))
    story.append(Spacer(1, 12))
    
    # Información del proyecto
    story.append(Paragraph(f"<b>Proyecto:</b> {proyecto.nombre}", normal_style))
    story.append(Paragraph(f"<b>Cliente:</b> {proyecto.cliente.razon_social}", normal_style))
    if planilla_seleccionada:
        story.append(Paragraph(f"<b>Planilla:</b> {planilla_seleccionada.nombre}", normal_style))
        if planilla_seleccionada.fecha_inicio and planilla_seleccionada.fecha_fin:
            story.append(Paragraph(f"<b>Período:</b> {planilla_seleccionada.fecha_inicio.strftime('%d/%m/%Y')} - {planilla_seleccionada.fecha_fin.strftime('%d/%m/%Y')}", normal_style))
    # Obtener fecha local de Guatemala
    import pytz
    guatemala_tz = pytz.timezone('America/Guatemala')
    fecha_generacion = timezone.now().astimezone(guatemala_tz)
    story.append(Paragraph(f"<b>Fecha de Generación:</b> {fecha_generacion.strftime('%d/%m/%Y %H:%M')} (Guatemala)", normal_style))
    story.append(Spacer(1, 20))
    
    if trabajadores.exists():
        # Calcular totales usando datos temporales
        total_trabajadores = trabajadores.count()
        total_a_pagar = 0
        for trabajador in trabajadores:
            dias_trabajados = dias_trabajados_data.get(trabajador.id, 0)
            total_trabajador = float(trabajador.pago_diario) * dias_trabajados
            total_a_pagar += total_trabajador
        
        # Anticipos específicos de trabajadores diarios
        anticipos_trabajadores = AnticipoTrabajadorDiario.objects.filter(
            trabajador__proyecto=proyecto
        )
        total_anticipos = anticipos_trabajadores.aggregate(total=Sum('monto'))['total'] or Decimal('0')
        total_aplicado = sum(anticipo.monto_aplicado for anticipo in anticipos_trabajadores)
        # Convertir a Decimal para evitar errores de tipo
        saldo_pendiente = Decimal(str(total_a_pagar)) - Decimal(str(total_aplicado))
        
        # Crear tabla de trabajadores con columnas de anticipos
        data = [['No.', 'Nombre del Trabajador', 'Pago Diario', 'Días Trabajados', 'Total Bruto', 'Anticipos', 'Total Neto']]
        
        total_bruto_general = 0
        total_anticipos_general = 0
        total_neto_general = 0
        
        for i, trabajador in enumerate(trabajadores, 1):
            # Usar días trabajados temporales
            dias_trabajados = dias_trabajados_data.get(trabajador.id, 0)
            total_bruto = float(trabajador.pago_diario) * dias_trabajados
            
            # Calcular anticipos del trabajador
            anticipos_trabajador = AnticipoTrabajadorDiario.objects.filter(
                trabajador=trabajador,
                estado='aplicado'
            )
            total_anticipos_trabajador = sum(anticipo.monto_aplicado for anticipo in anticipos_trabajador)
            
            # Debug: Imprimir información del trabajador y sus anticipos
            print(f"🔍 DEBUG PDF - Trabajador: {trabajador.nombre} (ID: {trabajador.id})")
            print(f"🔍 DEBUG PDF - Anticipos encontrados: {anticipos_trabajador.count()}")
            for anticipo in anticipos_trabajador:
                print(f"🔍 DEBUG PDF - Anticipo ID: {anticipo.id}, Monto: {anticipo.monto}, Aplicado: {anticipo.monto_aplicado}, Estado: {anticipo.estado}")
            print(f"🔍 DEBUG PDF - Total anticipos trabajador: {total_anticipos_trabajador}")
            
            # Total neto = Total bruto - Anticipos
            total_neto = total_bruto - float(total_anticipos_trabajador)
            
            data.append([
                str(i),
                trabajador.nombre,
                f"${trabajador.pago_diario:.2f}",
                str(dias_trabajados),
                f"${total_bruto:.2f}",
                f"${total_anticipos_trabajador:.2f}",
                f"${total_neto:.2f}"
            ])
            
            total_bruto_general += total_bruto
            total_anticipos_general += float(total_anticipos_trabajador)
            total_neto_general += total_neto
        
        # Agregar fila de totales
        data.append(['', '', '', 'TOTAL GENERAL:', f"${total_bruto_general:.2f}", f"${total_anticipos_general:.2f}", f"${total_neto_general:.2f}"])
        
        # Crear la tabla con columnas adicionales (7 columnas total)
        # Ancho total: 11.7" - 2" (márgenes) = 9.7" disponibles
        table = Table(data, colWidths=[0.6*inch, 2.5*inch, 1.2*inch, 1.2*inch, 1.4*inch, 1.4*inch, 1.4*inch])
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            
            # Fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 8),
            
            # Datos
            ('FONTSIZE', (0, 1), (-1, -2), 7),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 30))
        
        # Resumen de anticipos
        story.append(Paragraph("RESUMEN DE ANTICIPOS", subtitle_style))
        
        anticipos_data = [
            ['Concepto', 'Monto (Q)'],
            ['Total Bruto a Pagar', f"${total_bruto_general:.2f}"],
            ['Total Anticipos Aplicados', f"${total_anticipos_general:.2f}"],
            ['Total Neto a Pagar', f"${total_neto_general:.2f}"]
        ]
        
        anticipos_table = Table(anticipos_data, colWidths=[3*inch, 1.5*inch])
        anticipos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(anticipos_table)
        story.append(Spacer(1, 20))
        
        # RESUMEN DE ANTICIPOS
        story.append(Paragraph("RESUMEN DE ANTICIPOS", title_style))
        story.append(Spacer(1, 12))
        
        # Crear tabla de resumen de anticipos
        resumen_data = [
            ['Concepto', 'Monto (Q)'],
            ['Total Bruto a Pagar', f"${total_bruto_general:.2f}"],
            ['Total Anticipos Aplicados', f"${total_anticipos_general:.2f}"],
            ['Total Neto a Pagar', f"${total_neto_general:.2f}"]
        ]
        
        resumen_table = Table(resumen_data, colWidths=[4*inch, 2*inch])
        resumen_table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(resumen_table)
        story.append(Spacer(1, 20))
        
        # Información adicional
        story.append(Paragraph(f"<b>Total de Trabajadores:</b> {total_trabajadores}", normal_style))
        story.append(Paragraph(f"<b>Total Bruto a Pagar:</b> ${total_bruto_general:.2f}", normal_style))
        story.append(Paragraph(f"<b>Total Anticipos Aplicados:</b> ${total_anticipos_general:.2f}", normal_style))
        story.append(Paragraph(f"<b>Total Neto a Pagar:</b> ${total_neto_general:.2f}", normal_style))
        
    else:
        story.append(Paragraph("No hay trabajadores diarios registrados en este proyecto.", normal_style))
    
    # Construir el PDF
    doc.build(story)
    
    # Obtener el contenido del buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Guardar el PDF en la carpeta de archivos del proyecto
    try:
        from core.models import CarpetaProyecto, ArchivoProyecto
        from django.core.files.base import ContentFile
        
        # Crear o obtener carpeta "Trabajadores Diarios"
        carpeta, created = CarpetaProyecto.objects.get_or_create(
            proyecto=proyecto,
            nombre='Trabajadores Diarios',
            defaults={
                'creada_por': request.user,
                'descripcion': 'Carpeta para almacenar planillas de trabajadores diarios'
            }
        )
        
        # Crear archivo PDF
        if planilla_seleccionada:
            nombre_archivo = f"planilla_{planilla_seleccionada.nombre}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        else:
            nombre_archivo = f"planilla_trabajadores_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Crear ContentFile con el contenido PDF
        archivo_pdf = ContentFile(pdf_content)
        archivo_pdf.name = nombre_archivo
        
        archivo = ArchivoProyecto.objects.create(
            proyecto=proyecto,
            carpeta=carpeta,
            nombre=nombre_archivo,
            archivo=archivo_pdf,
            descripcion=f'Planilla de trabajadores diarios generada el {timezone.now().strftime("%d/%m/%Y %H:%M")}',
            subido_por=request.user,
            activo=True
        )
        
        print(f"✅ PDF guardado automáticamente: {nombre_archivo}")
        print(f"✅ Archivo ID: {archivo.id}")
        print(f"✅ Tiene archivo físico: {bool(archivo.archivo)}")
        if archivo.archivo:
            print(f"✅ Tamaño: {archivo.archivo.size} bytes")
            print(f"✅ Ruta: {archivo.archivo.path}")
            print(f"✅ Existe archivo: {os.path.exists(archivo.archivo.path)}")
        
    except Exception as e:
        print(f"⚠️ Error guardando PDF automáticamente: {e}")
        # Continuar con la descarga aunque falle el guardado
    
    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    if planilla_seleccionada:
        filename = f"planilla_{planilla_seleccionada.nombre}_{proyecto.nombre}_{timezone.now().strftime('%Y%m%d')}.pdf"
    else:
        filename = f"planilla_trabajadores_diarios_{proyecto.nombre}_{timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf_content)
    
    # Registrar actividad
    LogActividad.objects.create(
        usuario=request.user,
        accion='Exportar',
        modulo='Trabajadores Diarios',
        descripcion=f'PDF de planilla generado y guardado para proyecto {proyecto.nombre}',
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return response


# ==================== ANTICIPOS TRABAJADORES DIARIOS ====================

@login_required
def anticipo_trabajador_diario_list(request, proyecto_id):
    """Lista de anticipos de trabajadores diarios"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    anticipos = AnticipoTrabajadorDiario.objects.filter(
        trabajador__proyecto=proyecto
    ).select_related('trabajador', 'creado_por').order_by('-fecha_creacion')
    
    # Calcular totales
    total_anticipos = anticipos.aggregate(total=Sum('monto'))['total'] or 0
    total_aplicado = sum(anticipo.monto_aplicado for anticipo in anticipos)
    
    # Calcular total a pagar de trabajadores diarios
    trabajadores = TrabajadorDiario.objects.filter(proyecto=proyecto, activo=True)
    total_a_pagar = sum(t.total_a_pagar for t in trabajadores)
    saldo_pendiente = total_a_pagar - total_aplicado
    
    return render(request, 'core/anticipos_trabajadores_diarios/list.html', {
        'proyecto': proyecto,
        'anticipos': anticipos,
        'total_anticipos': total_anticipos,
        'total_aplicado': total_aplicado,
        'total_a_pagar': total_a_pagar,
        'saldo_pendiente': saldo_pendiente
    })


@login_required
def anticipo_trabajador_diario_create(request, proyecto_id):
    """Crear anticipo de trabajador diario - SIGUIENDO DOCUMENTACIÓN"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    planilla_id = request.GET.get('planilla_id') or request.POST.get('planilla_id')
    trabajador_id = request.GET.get('trabajador_id') or request.POST.get('trabajador_id')
    
    planilla_seleccionada = None
    if planilla_id:
        try:
            planilla_seleccionada = PlanillaTrabajadoresDiarios.objects.get(
                id=planilla_id, 
                proyecto=proyecto
            )
        except PlanillaTrabajadoresDiarios.DoesNotExist:
            planilla_seleccionada = None
    
    if request.method == 'POST':
        form = AnticipoTrabajadorDiarioForm(
            request.POST, 
            proyecto_id=proyecto_id,
            trabajador_id=trabajador_id
        )
        
        # Filtrar trabajadores por planilla seleccionada
        if planilla_seleccionada:
            form.fields['trabajador'].queryset = TrabajadorDiario.objects.filter(
                proyecto=proyecto,
                planilla=planilla_seleccionada,
                activo=True
            ).order_by('nombre')
        
        if form.is_valid():
            trabajador_seleccionado = form.cleaned_data.get('trabajador')
            monto_ingresado = form.cleaned_data.get('monto')
            
            # Crear el anticipo
            anticipo = form.save(commit=False)
            anticipo.creado_por = request.user
            # Los anticipos se aplican directamente (estado='aplicado')
            anticipo.estado = 'aplicado'
            anticipo.save()
            
            messages.success(request, 
                f'✅ Anticipo creado y aplicado exitosamente<br>'
                f'👤 Trabajador: <strong>{anticipo.trabajador.nombre}</strong><br>'
                f'💰 Monto: <strong>${anticipo.monto:,.2f}</strong>',
                extra_tags='html'
            )
            
            if planilla_id:
                return redirect(
                    f'{reverse("anticipo_trabajador_diario_list", args=[proyecto_id])}?planilla_id={planilla_id}'
                )
            return redirect('anticipo_trabajador_diario_list', proyecto_id=proyecto_id)
    else:
        form = AnticipoTrabajadorDiarioForm(
            proyecto_id=proyecto_id,
            trabajador_id=trabajador_id
        )
        
        # Filtrar trabajadores por planilla seleccionada
        if planilla_seleccionada:
            form.fields['trabajador'].queryset = TrabajadorDiario.objects.filter(
                proyecto=proyecto,
                planilla=planilla_seleccionada,
                activo=True
            ).order_by('nombre')
    
    return render(request, 'core/anticipos_trabajadores_diarios/create.html', {
        'form': form,
        'proyecto': proyecto,
        'planilla_seleccionada': planilla_seleccionada,
        'trabajador_id': trabajador_id
    })


@login_required
def anticipos_trabajador_diario_list(request, proyecto_id, trabajador_id):
    """Lista de anticipos de un trabajador diario específico"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, proyecto=proyecto)
    
    anticipos = AnticipoTrabajadorDiario.objects.filter(
        trabajador=trabajador
    ).select_related('creado_por').order_by('-fecha_creacion')
    
    # Calcular totales
    total_anticipos = sum(anticipo.monto for anticipo in anticipos)
    total_aplicado = sum(anticipo.monto for anticipo in anticipos.filter(estado='aplicado'))
    
    # Obtener planilla_id de la URL si existe
    planilla_id = request.GET.get('planilla_id')
    
    return render(request, 'core/anticipos_trabajadores_diarios/trabajador_list.html', {
        'proyecto': proyecto,
        'trabajador': trabajador,
        'anticipos': anticipos,
        'total_anticipos': total_anticipos,
        'total_aplicado': total_aplicado,
        'planilla_id': planilla_id
    })


@login_required
def anticipo_trabajador_diario_detail(request, proyecto_id, anticipo_id):
    """Detalle de anticipo de trabajador diario"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    anticipo = get_object_or_404(AnticipoTrabajadorDiario, id=anticipo_id, trabajador__proyecto=proyecto)
    
    return render(request, 'core/anticipos_trabajadores_diarios/detail.html', {
        'proyecto': proyecto,
        'anticipo': anticipo
    })


@login_required
def anticipo_trabajador_diario_edit(request, proyecto_id, anticipo_id):
    """Editar anticipo de trabajador diario"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    anticipo = get_object_or_404(AnticipoTrabajadorDiario, id=anticipo_id, trabajador__proyecto=proyecto)
    trabajador = anticipo.trabajador
    
    planilla_id = request.GET.get('planilla_id')
    
    if request.method == 'POST':
        form = AnticipoTrabajadorDiarioForm(request.POST, instance=anticipo, proyecto_id=proyecto_id)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Anticipo actualizado correctamente.')
            
            # Redirigir a la lista del trabajador si es posible
            redirect_url = reverse('anticipos_trabajador_diario_list', args=[proyecto_id, trabajador.id])
            if planilla_id:
                redirect_url += f'?planilla_id={planilla_id}'
            return redirect(redirect_url)
    else:
        form = AnticipoTrabajadorDiarioForm(instance=anticipo, proyecto_id=proyecto_id)
    
    return render(request, 'core/anticipos_trabajadores_diarios/edit.html', {
        'form': form,
        'proyecto': proyecto,
        'anticipo': anticipo,
        'trabajador': trabajador,
        'planilla_id': planilla_id
    })


@login_required
def anticipo_trabajador_diario_delete(request, proyecto_id, anticipo_id):
    """Eliminar anticipo de trabajador diario"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    anticipo = get_object_or_404(AnticipoTrabajadorDiario, id=anticipo_id, trabajador__proyecto=proyecto)
    trabajador = anticipo.trabajador
    
    planilla_id = request.GET.get('planilla_id')
    
    if request.method == 'POST':
        trabajador_nombre = anticipo.trabajador.nombre
        monto = anticipo.monto
        anticipo.delete()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar',
            modulo='Anticipos Trabajadores Diarios',
            descripcion=f'Anticipo eliminado para {trabajador_nombre}: ${monto}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, '✅ Anticipo eliminado correctamente')
        
        # Redirigir a la lista del trabajador si es posible
        redirect_url = reverse('anticipos_trabajador_diario_list', args=[proyecto_id, trabajador.id])
        if planilla_id:
            redirect_url += f'?planilla_id={planilla_id}'
        return redirect(redirect_url)
    
    return render(request, 'core/anticipos_trabajadores_diarios/delete.html', {
        'proyecto': proyecto,
        'anticipo': anticipo,
        'trabajador': trabajador,
        'planilla_id': planilla_id
    })


@login_required
def anticipo_trabajador_diario_aplicar(request, proyecto_id, anticipo_id):
    """Aplicar anticipo de trabajador diario"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    anticipo = get_object_or_404(AnticipoTrabajadorDiario, id=anticipo_id, trabajador__proyecto=proyecto)
    
    if request.method == 'POST':
        try:
            # Cambiar estado a aplicado
            anticipo.estado = 'aplicado'
            anticipo.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Aplicar',
                modulo='Anticipos Trabajadores Diarios',
                descripcion=f'Anticipo aplicado para {anticipo.trabajador.nombre}: ${anticipo.monto}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Anticipo de ${anticipo.monto} aplicado correctamente para {anticipo.trabajador.nombre}.')
            return redirect('anticipo_trabajador_diario_list', proyecto_id=proyecto_id)
            
        except Exception as e:
            messages.error(request, f'Error al aplicar el anticipo: {str(e)}')
            return redirect('anticipo_trabajador_diario_list', proyecto_id=proyecto_id)
    
    return render(request, 'core/anticipos_trabajadores_diarios/aplicar.html', {
        'proyecto': proyecto,
        'anticipo': anticipo
    })


# ==================== NUEVAS FUNCIONES OPTIMIZADAS ====================

@api_view()
def api_login(request):
    """API de login para AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')
            
            if username and password:
                user = authenticate(request, username=username, password=password)
                if user is not None and user.is_active:
                    login(request, user)
                    return JsonResponse({
                        'success': True,
                        'message': 'Login exitoso',
                        'redirect_url': reverse('dashboard')
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Credenciales inválidas'
                    })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Datos incompletos'
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Método no permitido'
    })


@api_view()
def dashboard_data_api(request):
    """API para obtener datos del dashboard en tiempo real"""
    try:
        # Estadísticas generales
        estadisticas = DashboardService.obtener_estadisticas_generales()
        
        # Datos de gráficos
        gastos_por_categoria = Gasto.objects.filter(aprobado=True).values(
            'categoria__nombre'
        ).annotate(
            total=Sum('monto')
        ).order_by('-total')[:5]
        
        proyectos_por_estado = Proyecto.objects.filter(activo=True).values(
            'estado'
        ).annotate(
            total=Count('id')
        )
        
        # Facturas por mes (últimos 6 meses)
        fecha_inicio = timezone.now().date() - timedelta(days=180)
        facturas_por_mes = Factura.objects.filter(
            fecha_emision__gte=fecha_inicio
        ).extra(
            select={'mes': "strftime('%%Y-%%m', fecha_emision)"}
        ).values('mes').annotate(
            total=Sum('monto_total')
        ).order_by('mes')
        
        return JsonResponse({
            'success': True,
            'data': {
                'estadisticas': estadisticas,
                'gastos_por_categoria': list(gastos_por_categoria),
                'proyectos_por_estado': list(proyectos_por_estado),
                'facturas_por_mes': list(facturas_por_mes),
            }
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@api_view()
def dashboard_intelligent_data(request):
    """API para datos de análisis inteligente"""
    try:
        # Datos para gráficos avanzados
        data = {
            'proyectos_rentabilidad': [],
            'gastos_tendencia': [],
            'facturas_estado': [],
            'colaboradores_activos': 0,
        }
        
        # Proyectos con rentabilidad
        for proyecto in Proyecto.objects.filter(activo=True)[:5]:
            rentabilidad = ProyectoService.calcular_rentabilidad(proyecto)
            data['proyectos_rentabilidad'].append({
                'nombre': proyecto.nombre,
                'rentabilidad': rentabilidad.get('rentabilidad_porcentaje', 0)
            })
        
        # Tendencia de gastos últimos 12 meses
        fecha_inicio = timezone.now().date() - timedelta(days=365)
        gastos_tendencia = Gasto.objects.filter(
            fecha_gasto__gte=fecha_inicio,
            aprobado=True
        ).extra(
            select={'mes': "strftime('%%Y-%%m', fecha_gasto)"}
        ).values('mes').annotate(
            total=Sum('monto')
        ).order_by('mes')
        
        data['gastos_tendencia'] = list(gastos_tendencia)
        
        # Estado de facturas
        facturas_estado = Factura.objects.values('estado').annotate(
            total=Count('id')
        )
        data['facturas_estado'] = list(facturas_estado)
        
        # Colaboradores activos
        data['colaboradores_activos'] = Colaborador.objects.filter(activo=True).count()
        
        return JsonResponse({
            'success': True,
            'data': data
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def dashboard_intelligent_analytics(request):
    """Dashboard con análisis inteligente"""
    try:
        # Análisis de rentabilidad por proyecto
        proyectos_rentabilidad = []
        for proyecto in Proyecto.objects.filter(activo=True)[:10]:
            rentabilidad = ProyectoService.calcular_rentabilidad(proyecto)
            proyectos_rentabilidad.append({
                'proyecto': proyecto.nombre,
                'rentabilidad': rentabilidad
            })
        
        # Análisis de tendencias de gastos
        ultimos_6_meses = timezone.now().date() - timedelta(days=180)
        gastos_tendencia = Gasto.objects.filter(
            fecha_gasto__gte=ultimos_6_meses,
            aprobado=True
        ).extra(
            select={'mes': "strftime('%%Y-%%m', fecha_gasto)"}
        ).values('mes').annotate(
            total=Sum('monto')
        ).order_by('mes')
        
        # Proyectos con mayor riesgo (gastos altos vs presupuesto)
        proyectos_riesgo = []
        # FILTRO DE PRESUPUESTO ELIMINADO - YA NO SE USA
        for proyecto in Proyecto.objects.filter(activo=True):
            estadisticas = ProyectoService.obtener_estadisticas_proyecto(proyecto)
            if estadisticas.get('gastos', {}).get('total', 0) > 0:
                # CÁLCULO DE PORCENTAJE ELIMINADO - YA NO SE USA
                porcentaje_gastado = 0
                if porcentaje_gastado > 80:  # Más del 80% gastado
                    proyectos_riesgo.append({
                        'proyecto': proyecto.nombre,
                        'porcentaje_gastado': round(porcentaje_gastado, 2),
                        'presupuesto': 0,  # ELIMINADO - YA NO SE USA
                        'gastado': estadisticas['gastos']['total']
                    })
        
        context = {
            'proyectos_rentabilidad': proyectos_rentabilidad,
            'gastos_tendencia': list(gastos_tendencia),
            'proyectos_riesgo': proyectos_riesgo,
        }
        
        return render(request, 'core/dashboard/intelligent_analytics.html', context)
    
    except Exception as e:
        messages.error(request, f'Error en análisis inteligente: {str(e)}')
        return render(request, 'core/dashboard/intelligent_analytics.html', {
            'proyectos_rentabilidad': [],
            'gastos_tendencia': [],
            'proyectos_riesgo': [],
        })


@api_view()
def cliente_toggle_estado(request, cliente_id):
    """Activar/desactivar cliente"""
    try:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        cliente.activo = not cliente.activo
        cliente.save()
        
        estado = "activado" if cliente.activo else "desactivado"
        messages.success(request, f'Cliente "{cliente.razon_social}" {estado} correctamente')
        
        return JsonResponse({
            'success': True,
            'activo': cliente.activo,
            'message': f'Cliente {estado} correctamente.'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def cliente_estadisticas(request, cliente_id):
    """Estadísticas detalladas del cliente"""
    try:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        
        # Estadísticas de proyectos
        proyectos_stats = Proyecto.objects.filter(cliente=cliente).aggregate(
            total=Count('id'),
            # presupuesto_total ELIMINADO - YA NO SE USA
        )
        
        # Estadísticas de facturas
        facturas_stats = Factura.objects.filter(cliente=cliente).aggregate(
            total=Count('id'),
            monto_total=Sum('monto_total'),
            monto_pagado=Sum('monto_pagado')
        )
        
        # Proyectos por estado
        proyectos_por_estado = Proyecto.objects.filter(cliente=cliente).values(
            'estado'
        ).annotate(
            total=Count('id')
        )
        
        # Facturas por estado
        facturas_por_estado = Factura.objects.filter(cliente=cliente).values(
            'estado'
        ).annotate(
            total=Count('id')
        )
        
        # Gastos por proyecto (últimos 6 meses)
        fecha_inicio = timezone.now().date() - timedelta(days=180)
        
        gastos_por_proyecto = Gasto.objects.filter(
            proyecto__cliente=cliente,
            fecha_gasto__gte=fecha_inicio,
            aprobado=True
        ).values(
            'proyecto__nombre'
        ).annotate(
            total=Sum('monto')
        ).order_by('-total')[:10]
        
        context = {
            'cliente': cliente,
            'proyectos_stats': proyectos_stats,
            'facturas_stats': facturas_stats,
            'proyectos_por_estado': list(proyectos_por_estado),
            'facturas_por_estado': list(facturas_por_estado),
            'gastos_por_proyecto': list(gastos_por_proyecto),
        }
        
        return render(request, 'core/clientes/estadisticas.html', context)
    
    except Exception as e:
        messages.error(request, f'Error al cargar estadísticas: {str(e)}')
        return redirect('cliente_detail', cliente_id=cliente_id)


# ==================== VISTAS DEL SISTEMA DE PLANILLAS MÚLTIPLES ====================

@login_required
def planillas_trabajadores_diarios_list(request, proyecto_id):
    """Lista de planillas de trabajadores diarios de un proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planillas = PlanillaTrabajadoresDiarios.objects.filter(proyecto=proyecto).order_by('-fecha_creacion')
    
    context = {
        'proyecto': proyecto,
        'planillas': planillas,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/list.html', context)


@login_required
def planilla_trabajadores_diarios_create(request, proyecto_id):
    """Crear nueva planilla de trabajadores diarios"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        form = PlanillaTrabajadoresDiariosForm(request.POST, proyecto=proyecto)
        if form.is_valid():
            planilla = form.save(commit=False)
            planilla.proyecto = proyecto
            planilla.creada_por = request.user
            planilla.save()
            
            messages.success(request, f'Planilla "{planilla.nombre}" creada exitosamente.')
            return redirect('planilla_trabajadores_diarios_detail', proyecto_id=proyecto_id, planilla_id=planilla.id)
    else:
        form = PlanillaTrabajadoresDiariosForm(proyecto=proyecto)
    
    context = {
        'proyecto': proyecto,
        'form': form,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/create.html', context)


@login_required
def planilla_trabajadores_diarios_detail(request, proyecto_id, planilla_id):
    """Detalle de una planilla de trabajadores diarios"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
    
    # Obtener trabajadores de la planilla usando la relación inversa
    # Si no hay trabajadores por la relación, intentar obtenerlos directamente
    trabajadores = TrabajadorDiario.objects.filter(planilla=planilla).order_by('nombre')
    
    # Calcular totales
    total_trabajadores = trabajadores.count()
    total_a_pagar = sum(t.total_a_pagar for t in trabajadores)
    total_anticipos = sum(t.total_anticipos_aplicados for t in trabajadores)
    saldo_pendiente = total_a_pagar - total_anticipos
    
    context = {
        'proyecto': proyecto,
        'planilla': planilla,
        'trabajadores': trabajadores,
        'total_trabajadores': total_trabajadores,
        'total_a_pagar': total_a_pagar,
        'total_anticipos': total_anticipos,
        'saldo_pendiente': saldo_pendiente,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/detail.html', context)


@login_required
def planilla_trabajadores_diarios_edit(request, proyecto_id, planilla_id):
    """Editar planilla de trabajadores diarios"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
    
    if request.method == 'POST':
        form = PlanillaTrabajadoresDiariosForm(request.POST, instance=planilla, proyecto=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request, f'Planilla "{planilla.nombre}" actualizada exitosamente.')
            return redirect('planilla_trabajadores_diarios_detail', proyecto_id=proyecto_id, planilla_id=planilla.id)
    else:
        form = PlanillaTrabajadoresDiariosForm(instance=planilla, proyecto=proyecto)
    
    context = {
        'proyecto': proyecto,
        'planilla': planilla,
        'form': form,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/edit.html', context)


@login_required
def planilla_trabajadores_diarios_delete(request, proyecto_id, planilla_id):
    """Eliminar planilla de trabajadores diarios"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
    
    if request.method == 'POST':
        nombre_planilla = planilla.nombre
        planilla.delete()
        messages.success(request, f'Planilla "{nombre_planilla}" eliminada exitosamente.')
        return redirect('planillas_trabajadores_diarios_list', proyecto_id=proyecto_id)
    
    context = {
        'proyecto': proyecto,
        'planilla': planilla,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/delete.html', context)


@login_required
def planilla_trabajadores_diarios_finalizar(request, proyecto_id, planilla_id):
    """Finalizar planilla de trabajadores diarios"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
    
    if request.method == 'POST':
        try:
            # Marcar planilla como finalizada
            planilla.estado = 'finalizada'
            planilla.fecha_finalizacion = timezone.now()
            planilla.finalizada_por = request.user
            planilla.save()
            
            # Marcar trabajadores como inactivos
            planilla.trabajadores.update(activo=False)
            
            # Marcar anticipos como procesados
            AnticipoTrabajadorDiario.objects.filter(
                trabajador__planilla=planilla,
                estado='aplicado'
            ).update(
                estado='procesado',
                fecha_liquidacion=timezone.now().date(),
                liquidado_por=request.user
            )
            
            # Crear registro de planilla liquidada
            PlanillaLiquidada.objects.create(
                proyecto=proyecto,
                total_salarios=Decimal(str(planilla.total_a_pagar)),
                total_anticipos=Decimal(str(planilla.total_anticipos)),
                total_planilla=Decimal(str(planilla.total_a_pagar + planilla.total_anticipos)),
                cantidad_personal=planilla.total_trabajadores,
                liquidada_por=request.user,
                observaciones=f'Planilla de trabajadores diarios: {planilla.nombre}'
            )
            
            messages.success(request, f'Planilla "{planilla.nombre}" finalizada exitosamente.')
            return redirect('planillas_trabajadores_diarios_list', proyecto_id=proyecto_id)
            
        except Exception as e:
            print(f"Error al finalizar planilla: {e}")
            messages.error(request, f'Error al finalizar planilla: {str(e)}')
    
    context = {
        'proyecto': proyecto,
        'planilla': planilla,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/finalizar.html', context)


@login_required
def trabajador_diario_add_to_planilla(request, proyecto_id, planilla_id):
    """Agregar trabajador a una planilla específica"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
    
    if request.method == 'POST':
        form = TrabajadorDiarioForm(request.POST, planilla=planilla, proyecto=proyecto)
        if form.is_valid():
            nombre_trabajador = form.cleaned_data.get('nombre')
            
            # Validar que no exista el mismo nombre en otras planillas ACTIVAS
            trabajadores_duplicados = TrabajadorDiario.objects.filter(
                proyecto=proyecto,
                nombre__iexact=nombre_trabajador,
                activo=True,
                planilla__estado__in=['activa', 'pendiente']
            ).exclude(planilla=planilla)
            
            if trabajadores_duplicados.exists():
                planilla_duplicada = trabajadores_duplicados.first().planilla
                messages.error(
                    request,
                    f'❌ El trabajador "{nombre_trabajador}" ya existe en la planilla activa "{planilla_duplicada.nombre}". '
                    f'No se permiten trabajadores duplicados entre planillas activas.'
                )
            else:
                trabajador = form.save(commit=False)
                trabajador.proyecto = proyecto
                trabajador.planilla = planilla
                trabajador.creado_por = request.user
                trabajador.save()
                
                # Registrar actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='Crear',
                    modulo='Trabajadores Diarios',
                    descripcion=f'Trabajador "{trabajador.nombre}" agregado a la planilla "{planilla.nombre}"',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, 
                    f'✅ Trabajador "{trabajador.nombre}" agregado a la planilla "{planilla.nombre}".'
                )
                return redirect('planilla_trabajadores_diarios_detail', proyecto_id=proyecto_id, planilla_id=planilla_id)
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = TrabajadorDiarioForm(planilla=planilla, proyecto=proyecto)
    
    context = {
        'proyecto': proyecto,
        'planilla': planilla,
        'form': form,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/add_trabajador.html', context)


@login_required
def trabajador_diario_remove_from_planilla(request, proyecto_id, planilla_id, trabajador_id):
    """Remover trabajador de una planilla"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    planilla = get_object_or_404(PlanillaTrabajadoresDiarios, id=planilla_id, proyecto=proyecto)
    trabajador = get_object_or_404(TrabajadorDiario, id=trabajador_id, planilla=planilla)
    
    if request.method == 'POST':
        nombre_trabajador = trabajador.nombre
        trabajador.delete()
        messages.success(request, f'Trabajador "{nombre_trabajador}" removido de la planilla.')
        return redirect('planilla_trabajadores_diarios_detail', proyecto_id=proyecto_id, planilla_id=planilla_id)
    
    context = {
        'proyecto': proyecto,
        'planilla': planilla,
        'trabajador': trabajador,
    }
    
    return render(request, 'core/planillas_trabajadores_diarios/remove_trabajador.html', context)


# ========================================
# VISTAS PARA MÓDULO DE INGRESOS POR PROYECTO
# ========================================

@login_required
def ingresos_list(request):
    """Lista de ingresos por proyecto"""
    filtro_proyecto = request.GET.get('proyecto', '')
    filtro_estado = request.GET.get('estado', '')
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Query base
    ingresos = IngresoProyecto.objects.all().select_related('proyecto', 'factura', 'creado_por')
    
    # Aplicar filtros
    if filtro_proyecto:
        ingresos = ingresos.filter(proyecto_id=filtro_proyecto)
    
    if filtro_estado:
        if filtro_estado == 'pagado':
            ingresos = ingresos.filter(pagado=True)
        elif filtro_estado == 'pendiente':
            ingresos = ingresos.filter(pagado=False)
    
    if filtro_fecha_desde:
        ingresos = ingresos.filter(fecha_emision__gte=filtro_fecha_desde)
    
    if filtro_fecha_hasta:
        ingresos = ingresos.filter(fecha_emision__lte=filtro_fecha_hasta)
    
    # Ordenar por fecha de emisión descendente
    ingresos = ingresos.order_by('-fecha_emision', '-fecha_registro')
    
    # Obtener proyectos para el filtro
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    # Estadísticas
    total_ingresos = ingresos.aggregate(total=Sum('monto_total'))['total'] or 0
    ingresos_pagados = ingresos.filter(pagado=True).aggregate(total=Sum('monto_total'))['total'] or 0
    ingresos_pendientes = ingresos.filter(pagado=False).aggregate(total=Sum('monto_total'))['total'] or 0
    
    context = {
        'ingresos': ingresos,
        'proyectos': proyectos,
        'filtro_proyecto': filtro_proyecto,
        'filtro_estado': filtro_estado,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
        'total_ingresos': total_ingresos,
        'ingresos_pagados': ingresos_pagados,
        'ingresos_pendientes': ingresos_pendientes,
    }
    
    return render(request, 'core/ingresos/list.html', context)


@login_required
def ingreso_create(request):
    """Crear nuevo ingreso por proyecto"""
    if request.method == 'POST':
        form = IngresoProyectoForm(request.POST)
        if form.is_valid():
            ingreso = form.save(commit=False)
            ingreso.creado_por = request.user
            ingreso.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Crear Ingreso',
                modulo='Ingresos',
                descripcion=f'Ingreso creado: {ingreso.numero_documento} - {ingreso.proyecto.nombre} - ${ingreso.monto_total}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Ingreso "{ingreso.numero_documento}" creado exitosamente')
            return redirect('ingresos_list')
    else:
        form = IngresoProyectoForm()
    
    context = {
        'form': form,
        'titulo': 'Crear Nuevo Ingreso',
    }
    
    return render(request, 'core/ingresos/create.html', context)


@login_required
def ingreso_edit(request, ingreso_id):
    """Editar ingreso por proyecto"""
    ingreso = get_object_or_404(IngresoProyecto, id=ingreso_id)
    
    if request.method == 'POST':
        form = IngresoProyectoForm(request.POST, instance=ingreso)
        if form.is_valid():
            ingreso = form.save(commit=False)
            ingreso.modificado_por = request.user
            ingreso.save()
            
            # Registrar actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='Editar Ingreso',
                modulo='Ingresos',
                descripcion=f'Ingreso editado: {ingreso.numero_documento} - {ingreso.proyecto.nombre} - ${ingreso.monto_total}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Ingreso "{ingreso.numero_documento}" actualizado exitosamente')
            return redirect('ingresos_list')
    else:
        form = IngresoProyectoForm(instance=ingreso)
    
    context = {
        'form': form,
        'ingreso': ingreso,
        'titulo': 'Editar Ingreso',
    }
    
    return render(request, 'core/ingresos/edit.html', context)


@login_required
def ingreso_detail(request, ingreso_id):
    """Detalle de un ingreso por proyecto"""
    ingreso = get_object_or_404(IngresoProyecto, id=ingreso_id)
    
    context = {
        'ingreso': ingreso,
    }
    
    return render(request, 'core/ingresos/detail.html', context)


@login_required
def ingreso_delete(request, ingreso_id):
    """Eliminar ingreso por proyecto"""
    ingreso = get_object_or_404(IngresoProyecto, id=ingreso_id)
    
    if request.method == 'POST':
        numero_documento = ingreso.numero_documento
        proyecto_nombre = ingreso.proyecto.nombre
        
        # Registrar actividad antes de eliminar
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar Ingreso',
            modulo='Ingresos',
            descripcion=f'Ingreso eliminado: {numero_documento} - {proyecto_nombre}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        ingreso.delete()
        messages.success(request, f'Ingreso "{numero_documento}" eliminado exitosamente')
        return redirect('ingresos_list')
    
    context = {
        'ingreso': ingreso,
    }
    
    return render(request, 'core/ingresos/delete.html', context)


@login_required
def ingresos_proyecto(request, proyecto_id):
    """Lista de ingresos específicos de un proyecto"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    ingresos = IngresoProyecto.objects.filter(proyecto=proyecto).select_related('factura', 'creado_por').order_by('-fecha_emision')
    
    # Estadísticas del proyecto
    total_ingresos = ingresos.aggregate(total=Sum('monto_total'))['total'] or 0
    ingresos_pagados = ingresos.filter(pagado=True).aggregate(total=Sum('monto_total'))['total'] or 0
    ingresos_pendientes = ingresos.filter(pagado=False).aggregate(total=Sum('monto_total'))['total'] or 0
    
    context = {
        'proyecto': proyecto,
        'ingresos': ingresos,
        'total_ingresos': total_ingresos,
        'ingresos_pagados': ingresos_pagados,
        'ingresos_pendientes': ingresos_pendientes,
    }
    
    return render(request, 'core/ingresos/proyecto.html', context)


# ==================== VISTAS DE COTIZACIONES ====================

@login_required
def cotizaciones_list(request):
    """Lista todas las cotizaciones"""
    cotizaciones = Cotizacion.objects.select_related('proyecto', 'cliente', 'creado_por').order_by('-fecha_emision')
    
    # Filtros
    estado = request.GET.get('estado')
    proyecto_id = request.GET.get('proyecto')
    
    if estado:
        cotizaciones = cotizaciones.filter(estado=estado)
    
    if proyecto_id:
        cotizaciones = cotizaciones.filter(proyecto_id=proyecto_id)
    
    # Estadísticas
    total_cotizaciones = cotizaciones.count()
    cotizaciones_aceptadas = cotizaciones.filter(estado='aceptada').count()
    cotizaciones_pendientes = cotizaciones.filter(estado__in=['borrador', 'enviada']).count()
    monto_total_cotizado = cotizaciones.aggregate(total=Sum('monto_total'))['total'] or 0
    
    # Proyectos para filtro
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'cotizaciones': cotizaciones,
        'total_cotizaciones': total_cotizaciones,
        'cotizaciones_aceptadas': cotizaciones_aceptadas,
        'cotizaciones_pendientes': cotizaciones_pendientes,
        'monto_total_cotizado': monto_total_cotizado,
        'proyectos': proyectos,
        'estado_filtro': estado,
        'proyecto_filtro': proyecto_id,
    }
    
    return render(request, 'core/cotizaciones/list.html', context)


@login_required
def cotizacion_create(request):
    """Crear nueva cotización"""
    logger = logging.getLogger(__name__)
    
    if request.method == 'POST':
        logger.info('📝 POST recibido para crear cotización')
        logger.info(f'📝 Datos POST: {request.POST}')
        form = CotizacionForm(request.POST, request.FILES)
        if form.is_valid():
            logger.info('✅ Formulario válido, guardando cotización')
            cotizacion = form.save(commit=False)
            cotizacion.creado_por = request.user
            cotizacion.save()
            logger.info(f'✅ Cotización {cotizacion.numero_cotizacion} guardada')
            
            # Guardar items de la cotización
            items = []
            for i in range(1000):  # Loop para buscar todos los items
                if f'items[{i}][descripcion]' in request.POST:
                    items.append({
                        'descripcion': request.POST[f'items[{i}][descripcion]'],
                        'cantidad': Decimal(request.POST.get(f'items[{i}][cantidad]', '1')),
                        'precio_unitario': Decimal(request.POST.get(f'items[{i}][precio_unitario]', '0')),
                        'precio_costo': Decimal(request.POST.get(f'items[{i}][precio_costo]', '0')),
                        'orden': int(request.POST.get(f'items[{i}][orden]', i)),
                    })
                else:
                    break  # No hay más items
            
            logger.info(f'📦 Items encontrados en POST: {len(items)}')
            
            # Guardar items en la base de datos
            for item_data in items:
                ItemCotizacion.objects.create(
                    cotizacion=cotizacion,
                    descripcion=item_data['descripcion'],
                    cantidad=item_data['cantidad'],
                    precio_unitario=item_data['precio_unitario'],
                    precio_costo=item_data['precio_costo'],
                    orden=item_data['orden']
                )
                
                # Crear o actualizar item reutilizable si no existe
                if item_data['descripcion']:
                    try:
                        ItemReutilizable.objects.get(descripcion=item_data['descripcion'])
                    except ItemReutilizable.DoesNotExist:
                        # Crear nuevo item reutilizable
                        ItemReutilizable.objects.create(
                            descripcion=item_data['descripcion'],
                            precio_unitario=item_data['precio_unitario'],
                            precio_costo=item_data['precio_costo'],
                            creado_por=request.user,
                            activo=True
                        )
            
            logger.info('✅ Items guardados exitosamente')
            messages.success(request, f'Cotización {cotizacion.numero_cotizacion} creada exitosamente.')
            return redirect('cotizaciones_list')
        else:
            logger.error(f'❌ Errores en formulario: {form.errors}')
            for field, errors in form.errors.items():
                for error in errors:
                    logger.error(f'❌ Campo {field}: {error}')
            messages.error(request, f'Por favor corrige los errores en el formulario. Errores: {form.errors}')
    else:
        form = CotizacionForm()
        
        # Generar número de cotización automáticamente
        ultima_cotizacion = Cotizacion.objects.all().order_by('-id').first()
        if ultima_cotizacion and ultima_cotizacion.numero_cotizacion:
            try:
                ultimo_numero = int(ultima_cotizacion.numero_cotizacion.split('-')[-1])
                nuevo_numero = ultimo_numero + 1
            except (ValueError, IndexError):
                nuevo_numero = 1
        else:
            nuevo_numero = 1
        
        año_actual = timezone.now().year
        numero_generado = f"COT-{año_actual}-{nuevo_numero:04d}"
        form.initial['numero_cotizacion'] = numero_generado
    
    # Datos para el formulario
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
    
    context = {
        'form': form,
        'proyectos': proyectos,
        'clientes': clientes,
    }
    
    return render(request, 'core/cotizaciones/create.html', context)


@login_required
def cotizacion_detail(request, cotizacion_id):
    """Detalle de una cotización"""
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    
    context = {
        'cotizacion': cotizacion,
    }
    
    return render(request, 'core/cotizaciones/detail.html', context)


@login_required
def cotizacion_pdf(request, cotizacion_id):
    """Generar PDF de una cotización"""
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    
    # Obtener items de la cotización
    items = ItemCotizacion.objects.filter(cotizacion=cotizacion).order_by('orden')
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Header con logo a la izquierda y número a la derecha
    header_data = []
    
    # Logo
    try:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'LOGO-TELECOM-small.png')
        if os.path.exists(logo_path):
            logo_cell = Image(logo_path, width=100, height=50)
        else:
            logo_cell = Paragraph("TELECOM<br/>Technology Panama INC.", styles['Heading3'])
    except:
        logo_cell = Paragraph("TELECOM<br/>Technology Panama INC.", styles['Heading3'])
    
    # Número de cotización
    number_cell = Paragraph(
        f'<b>COTIZACIÓN NO.</b><br/><font color="red">{cotizacion.numero_cotizacion}</font>',
        ParagraphStyle('Number', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT)
    )
    
    header_table = Table([[logo_cell, number_cell]], colWidths=[4*inch, 2*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (0, 0), colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))
    
    # Título principal
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=32,
        spaceAfter=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#000000'),
        fontName='Helvetica-Bold',
        leading=36
    )
    elements.append(Paragraph("COTIZACIÓN", title_style))
    elements.append(Spacer(1, 5))
    
    # Fecha
    fecha_style = ParagraphStyle(
        'FechaStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.black,
        spaceAfter=20
    )
    elements.append(Paragraph(f'<b>FECHA:</b> {cotizacion.fecha_emision.strftime("%d/%m/%Y")}', fecha_style))
    elements.append(Spacer(1, 15))
    
    # Sección "COTIZACIÓN PARA" con barra azul
    cliente_bar_data = [
        [Paragraph(f'<font color="white"><b>COTIZACIÓN PARA</b></font>', ParagraphStyle('Bar', fontSize=12, textColor=colors.white, alignment=TA_CENTER, fontName='Helvetica-Bold'))]
    ]
    cliente_bar_table = Table(cliente_bar_data, colWidths=[6*inch])
    cliente_bar_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#007bff')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(cliente_bar_table)
    elements.append(Spacer(1, 10))
    
    # Información del cliente
    info_data = [
        ['CLIENTE:', cotizacion.cliente.razon_social],
        ['PROYECTO:', cotizacion.proyecto.nombre],
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 4.5*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (0, -1), 10),
        ('FONTSIZE', (1, 0), (1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 10))
    
    # Título en negrita sin label
    title_paragraph = Paragraph(
        f'<b>{cotizacion.titulo}</b>',
        ParagraphStyle('TitleOnly', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', spaceAfter=15)
    )
    elements.append(title_paragraph)
    elements.append(Spacer(1, 5))
    
    # Tabla de items
    if items:
        headers = ['DESCRIPCIÓN', 'CANTIDAD', 'PRECIO', 'TOTAL']
        data = [headers]
        
        for item in items:
            data.append([
                item.descripcion,
                str(item.cantidad),
                f"${item.precio_unitario:,.2f}",
                f"${item.total:,.2f}"
            ])
        
        # Totales
        data.append(['', '', 'Subtotal:', f"${cotizacion.monto_subtotal:,.2f}"])
        data.append(['', '', 'ITBMS (7%):', f"${cotizacion.monto_iva:,.2f}"])
        data.append(['', '', 'TOTAL:', f"${cotizacion.monto_total:,.2f}"])
        
        items_table = Table(data, colWidths=[3.5*inch, 0.8*inch, 0.9*inch, 0.8*inch])
        items_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (-1, 0), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Body
            ('BACKGROUND', (0, 1), (-1, -4), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -4), colors.black),
            ('FONTSIZE', (0, 1), (-1, -4), 10),
            ('ALIGN', (0, 1), (0, -4), 'LEFT'),
            ('ALIGN', (1, 1), (1, -4), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -4), 'RIGHT'),
            ('GRID', (0, 0), (-1, -4), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -4), [colors.white, colors.HexColor('#f8f9fa')]),
            # Totales
            ('BACKGROUND', (2, -3), (-1, -2), colors.white),
            ('FONTNAME', (2, -3), (-1, -2), 'Helvetica-Bold'),
            ('FONTSIZE', (2, -3), (-1, -2), 10),
            ('ALIGN', (2, -3), (-1, -2), 'RIGHT'),
            ('GRID', (2, -3), (-1, -2), 0.5, colors.black),
            # TOTAL final
            ('BACKGROUND', (2, -1), (-1, -1), colors.black),
            ('TEXTCOLOR', (2, -1), (-1, -1), colors.white),
            ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (2, -1), (-1, -1), 14),
            ('ALIGN', (2, -1), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (2, -1), (-1, -1), 10),
            ('TOPPADDING', (2, -1), (-1, -1), 10),
        ]))
        
        elements.append(items_table)
        elements.append(Spacer(1, 20))
    
    # Términos y condiciones
    if cotizacion.terminos_condiciones:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("<b>TÉRMINOS Y CONDICIONES</b>", ParagraphStyle('TermsHeader', parent=styles['Normal'], fontSize=11, fontName='Helvetica-Bold', spaceAfter=5)))
        elements.append(Paragraph(cotizacion.terminos_condiciones, ParagraphStyle('TermsBody', parent=styles['Normal'], fontSize=9, spaceAfter=15)))
    
    # Información de la empresa
    config = ConfiguracionSistema.get_config()
    company_info = Paragraph(
        f'<b>{config.nombre_empresa}</b><br/>'
        'Technology Panama INC.<br/>'
        'Correo: info@telecompanama.com<br/>'
        'Tel: +507 206-3456',
        ParagraphStyle('CompanyInfo', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#6c757d'))
    )
    elements.append(Spacer(1, 20))
    elements.append(company_info)
    
    # Pie de página
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=7,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#9ca3af')
    )
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("Documento generado electrónicamente", footer_style))
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Cotizacion_{cotizacion.numero_cotizacion}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def cotizacion_edit(request, cotizacion_id):
    """Editar cotización"""
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    
    if request.method == 'POST':
        form = CotizacionForm(request.POST, request.FILES, instance=cotizacion)
        if form.is_valid():
            cotizacion = form.save(commit=False)
            cotizacion.modificado_por = request.user
            cotizacion.save()
            messages.success(request, f'Cotización {cotizacion.numero_cotizacion} actualizada exitosamente.')
            return redirect('cotizacion_detail', cotizacion_id=cotizacion.id)
    else:
        form = CotizacionForm(instance=cotizacion)
    
    context = {
        'form': form,
        'cotizacion': cotizacion,
    }
    
    return render(request, 'core/cotizaciones/edit.html', context)


@login_required
def cotizacion_delete(request, cotizacion_id):
    """Eliminar cotización"""
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    
    if request.method == 'POST':
        numero_cotizacion = cotizacion.numero_cotizacion
        cotizacion.delete()
        messages.success(request, f'Cotización {numero_cotizacion} eliminada exitosamente.')
        return redirect('cotizaciones_list')
    
    context = {
        'cotizacion': cotizacion,
    }
    
    return render(request, 'core/cotizaciones/delete.html', context)


@login_required
def cotizacion_aprobar(request, cotizacion_id):
    """Aprobar cotización"""
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    
    if request.method == 'POST':
        cotizacion.estado = 'aceptada'
        cotizacion.fecha_aceptacion = timezone.now().date()
        cotizacion.modificado_por = request.user
        cotizacion.save()
        
        messages.success(request, f'✅ Cotización {cotizacion.numero_cotizacion} aprobada exitosamente.')
        return redirect('cotizacion_detail', cotizacion_id=cotizacion.id)
    
    context = {
        'cotizacion': cotizacion,
    }
    
    return render(request, 'core/cotizaciones/aprobar.html', context)


@login_required
def cotizacion_rechazar(request, cotizacion_id):
    """Rechazar cotización"""
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    
    if request.method == 'POST':
        motivo_rechazo = request.POST.get('motivo_rechazo', '')
        
        cotizacion.estado = 'rechazada'
        if motivo_rechazo:
            cotizacion.observaciones = f"Motivo de rechazo: {motivo_rechazo}\n\n{cotizacion.observaciones}"
        cotizacion.modificado_por = request.user
        cotizacion.save()
        
        messages.warning(request, f'❌ Cotización {cotizacion.numero_cotizacion} rechazada.')
        return redirect('cotizacion_detail', cotizacion_id=cotizacion.id)
    
    context = {
        'cotizacion': cotizacion,
    }
    
    return render(request, 'core/cotizaciones/rechazar.html', context)


@login_required
def cotizaciones_dashboard(request):
    """Dashboard de cotizaciones"""
    # Estadísticas generales
    total_cotizaciones = Cotizacion.objects.count()
    cotizaciones_aceptadas = Cotizacion.objects.filter(estado='aceptada').count()
    cotizaciones_pendientes = Cotizacion.objects.filter(estado__in=['borrador', 'enviada']).count()
    cotizaciones_vencidas = Cotizacion.objects.filter(estado__in=['borrador', 'enviada']).filter(fecha_vencimiento__lt=timezone.now().date()).count()
    
    # Montos
    monto_total_cotizado = Cotizacion.objects.aggregate(total=Sum('monto_total'))['total'] or 0
    monto_aceptado = Cotizacion.objects.filter(estado='aceptada').aggregate(total=Sum('monto_total'))['total'] or 0
    monto_pendiente = Cotizacion.objects.filter(estado__in=['borrador', 'enviada']).aggregate(total=Sum('monto_total'))['total'] or 0
    
    # Cotizaciones recientes
    cotizaciones_recientes = Cotizacion.objects.select_related('proyecto', 'cliente').order_by('-fecha_emision')[:10]
    
    # Cotizaciones por estado
    cotizaciones_por_estado = Cotizacion.objects.values('estado').annotate(
        cantidad=Count('id'),
        monto_total=Sum('monto_total')
    ).order_by('estado')
    
    # Cotizaciones por proyecto
    cotizaciones_por_proyecto = Cotizacion.objects.values('proyecto__nombre').annotate(
        cantidad=Count('id'),
        monto_total=Sum('monto_total')
    ).order_by('-monto_total')[:10]
    
    context = {
        'total_cotizaciones': total_cotizaciones,
        'cotizaciones_aceptadas': cotizaciones_aceptadas,
        'cotizaciones_pendientes': cotizaciones_pendientes,
        'cotizaciones_vencidas': cotizaciones_vencidas,
        'monto_total_cotizado': monto_total_cotizado,
        'monto_aceptado': monto_aceptado,
        'monto_pendiente': monto_pendiente,
        'cotizaciones_recientes': cotizaciones_recientes,
        'cotizaciones_por_estado': cotizaciones_por_estado,
        'cotizaciones_por_proyecto': cotizaciones_por_proyecto,
    }
    
    return render(request, 'core/cotizaciones/dashboard.html', context)


@login_required
def cotizaciones_proyecto(request, proyecto_id):
    """Cotizaciones de un proyecto específico"""
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    cotizaciones = Cotizacion.objects.filter(proyecto=proyecto).select_related('cliente', 'creado_por').order_by('-fecha_emision')
    
    # Estadísticas del proyecto
    total_cotizaciones = cotizaciones.count()
    cotizaciones_aceptadas = cotizaciones.filter(estado='aceptada').count()
    monto_total_cotizado = cotizaciones.aggregate(total=Sum('monto_total'))['total'] or 0
    monto_aceptado = cotizaciones.filter(estado='aceptada').aggregate(total=Sum('monto_total'))['total'] or 0
    
    context = {
        'proyecto': proyecto,
        'cotizaciones': cotizaciones,
        'total_cotizaciones': total_cotizaciones,
        'cotizaciones_aceptadas': cotizaciones_aceptadas,
        'monto_total_cotizado': monto_total_cotizado,
        'monto_aceptado': monto_aceptado,
    }
    
    return render(request, 'core/cotizaciones/proyecto.html', context)


@login_required
def items_reutilizables_list(request):
    """Listar items reutilizables para AJAX"""
    logger = logging.getLogger(__name__)
    logger.info(f"📦 Petición para listar items reutilizables - Método: {request.method}")
    
    if request.method == 'GET':
        try:
            items = ItemReutilizable.objects.filter(activo=True).order_by('categoria', 'descripcion')
            logger.info(f"📦 Items encontrados: {items.count()}")
            items_data = []
            for item in items:
                items_data.append({
                    'id': item.id,
                    'descripcion': item.descripcion,
                    'categoria': item.categoria,
                    'precio_unitario': str(item.precio_unitario),
                    'precio_costo': str(item.precio_costo),
                })
            logger.info(f"✅ Retornando {len(items_data)} items")
            return JsonResponse({'items': items_data})
        except Exception as e:
            logger.error(f"❌ Error al obtener items: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    logger.warning(f"⚠️ Método no permitido: {request.method}")
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def item_reutilizable_create(request):
    """Crear un item reutilizable desde AJAX"""
    if request.method == 'POST':
        try:
            descripcion = request.POST.get('descripcion')
            categoria = request.POST.get('categoria', '')
            precio_unitario = Decimal(request.POST.get('precio_unitario', '0'))
            precio_costo = Decimal(request.POST.get('precio_costo', '0'))
            
            item = ItemReutilizable.objects.create(
                descripcion=descripcion,
                categoria=categoria,
                precio_unitario=precio_unitario,
                precio_costo=precio_costo,
                creado_por=request.user
            )
            
            return JsonResponse({
                'success': True,
                'item': {
                    'id': item.id,
                    'descripcion': item.descripcion,
                    'categoria': item.categoria,
                    'precio_unitario': str(item.precio_unitario),
                    'precio_costo': str(item.precio_costo),
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ============================================================================
# VISTAS DE CAJA MENUDA
# ============================================================================

@login_required
def caja_menuda_dashboard(request):
    """Dashboard principal de Caja Menuda"""
    try:
        # Estadísticas básicas
        total_movimientos = CajaMenuda.objects.filter(activo=True).count()
        total_monto = CajaMenuda.objects.filter(activo=True).aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')
        
        # Movimientos recientes
        movimientos_recientes = CajaMenuda.objects.filter(activo=True).order_by('-fecha', '-creado_en')[:10]
        
        context = {
            'total_movimientos': total_movimientos,
            'total_monto': total_monto,
            'movimientos_recientes': movimientos_recientes,
        }
        
        return render(request, 'core/caja-menuda/dashboard.html', context)
    except Exception as e:
        logger.error(f"Error en caja_menuda_dashboard: {str(e)}")
        messages.error(request, f'Error al cargar el dashboard: {str(e)}')
        return redirect('dashboard')


@login_required
def caja_menuda_list(request):
    """Lista todos los movimientos de caja menuda"""
    movimientos = CajaMenuda.objects.filter(activo=True).order_by('-fecha', '-creado_en')
    
    context = {
        'movimientos': movimientos,
    }
    
    return render(request, 'core/caja-menuda/list.html', context)


@login_required
def caja_menuda_create(request):
    """Crear nuevo movimiento de caja menuda"""
    if request.method == 'POST':
        form = CajaMenudaForm(request.POST)
        if form.is_valid():
            try:
                movimiento = form.save(commit=False)
                movimiento.creado_por = request.user
                movimiento.save()
                
                messages.success(request, '✅ Movimiento de caja menuda registrado exitosamente')
                return redirect('caja_menuda_list')
            except Exception as e:
                logger.error(f"Error al crear movimiento de caja menuda: {str(e)}")
                messages.error(request, f'❌ Error al registrar movimiento: {str(e)}')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        form = CajaMenudaForm()
    
    context = {
        'form': form,
        'accion': 'Crear'
    }
    
    return render(request, 'core/caja-menuda/create.html', context)


@login_required
def caja_menuda_edit(request, pk):
    """Editar movimiento de caja menuda"""
    movimiento = get_object_or_404(CajaMenuda, pk=pk)
    
    if request.method == 'POST':
        form = CajaMenudaForm(request.POST, instance=movimiento)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, '✅ Movimiento actualizado exitosamente')
                return redirect('caja_menuda_list')
            except Exception as e:
                logger.error(f"Error al actualizar movimiento: {str(e)}")
                messages.error(request, f'❌ Error al actualizar: {str(e)}')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        form = CajaMenudaForm(instance=movimiento)
    
    context = {
        'form': form,
        'movimiento': movimiento,
        'accion': 'Editar'
    }
    
    return render(request, 'core/caja-menuda/edit.html', context)


@login_required
def caja_menuda_delete(request, pk):
    """Eliminar (desactivar) movimiento de caja menuda"""
    movimiento = get_object_or_404(CajaMenuda, pk=pk)
    
    if request.method == 'POST':
        try:
            movimiento.activo = False
            movimiento.save()
            messages.success(request, '✅ Movimiento eliminado exitosamente')
            return redirect('caja_menuda_list')
        except Exception as e:
            logger.error(f"Error al eliminar movimiento: {str(e)}")
            messages.error(request, f'❌ Error al eliminar: {str(e)}')
    
    context = {
        'movimiento': movimiento
    }
    
    return render(request, 'core/caja-menuda/delete.html', context)


# ============================================================================
# MÓDULO DE TORREROS
# ============================================================================

@login_required
def torreros_dashboard(request):
    """Dashboard principal del módulo de torreros"""
    try:
        # Servicios activos - forzar recálculo de días trabajados para asegurar datos actualizados
        servicios_activos = ServicioTorrero.objects.filter(
            estado='activo',
            activo=True
        ).select_related('cliente', 'proyecto').order_by('-fecha_inicio')
        
        # Asegurar que todos los servicios tengan los días trabajados correctos
        # (solo recalcular si es necesario, para no afectar el rendimiento)
        for servicio in servicios_activos[:10]:  # Solo los primeros 10 que se mostrarán
            # Verificar si necesita recálculo comparando con la suma real
            from django.db.models import Sum
            total_real = servicio.registros_dias.filter(
                aprobado=True
            ).aggregate(total=Sum('dias_trabajados'))['total'] or Decimal('0.00')
            
            if isinstance(total_real, Decimal):
                total_real = total_real
            else:
                total_real = Decimal(str(total_real))
            
            if servicio.dias_trabajados != total_real:
                servicio.recalcular_dias_trabajados()
                servicio.refresh_from_db()
        
        # Estadísticas generales
        total_servicios_activos = servicios_activos.count()
        total_servicios = ServicioTorrero.objects.filter(activo=True).count()
        
        # Ingresos y pagos
        ingresos_totales = ServicioTorrero.objects.filter(
            activo=True
        ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0.00')
        
        pagos_recibidos = ServicioTorrero.objects.filter(
            activo=True
        ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0.00')
        
        saldo_pendiente = ingresos_totales - pagos_recibidos
        
        # Servicios por completar pronto (menos del 30% de días restantes)
        servicios_por_completar = []
        for servicio in servicios_activos:
            if servicio.porcentaje_completado >= 70:
                servicios_por_completar.append(servicio)
        
        # Servicios con pagos pendientes
        servicios_pago_pendiente = servicios_activos.filter(
            monto_pagado__lt=models.F('monto_total')
        ).order_by('cliente__nombre')[:5]
        
        # Últimos registros de días trabajados
        ultimos_registros = RegistroDiasTrabajados.objects.select_related(
            'servicio__cliente', 'registrado_por'
        ).order_by('-fecha_registro')[:10]
        
        # Últimos pagos
        ultimos_pagos = PagoServicioTorrero.objects.select_related(
            'servicio__cliente', 'registrado_por'
        ).order_by('-fecha_pago')[:10]
        
        # Torreros activos para el modal
        torreros_activos = Torrero.objects.filter(activo=True).order_by('nombre')
        
        context = {
            'servicios_activos': servicios_activos[:10],
            'total_servicios_activos': total_servicios_activos,
            'total_servicios': total_servicios,
            'ingresos_totales': ingresos_totales,
            'pagos_recibidos': pagos_recibidos,
            'saldo_pendiente': saldo_pendiente,
            'servicios_por_completar': servicios_por_completar[:5],
            'servicios_pago_pendiente': servicios_pago_pendiente,
            'ultimos_registros': ultimos_registros,
            'ultimos_pagos': ultimos_pagos,
            'torreros_activos': torreros_activos,
        }
        
        return render(request, 'core/torreros/dashboard.html', context)
        
    except Exception as e:
        logger.error(f'Error en torreros_dashboard: {e}')
        messages.error(request, 'Error al cargar el dashboard de torreros')
        return redirect('dashboard')


@login_required
def servicio_torrero_list(request):
    """Lista completa de servicios de torreros"""
    try:
        servicios = ServicioTorrero.objects.filter(
            activo=True
        ).select_related('cliente', 'proyecto', 'creado_por').order_by('-fecha_inicio')
        
        # Filtros
        filtro_estado = request.GET.get('estado', '')
        filtro_cliente = request.GET.get('cliente', '')
        
        if filtro_estado:
            servicios = servicios.filter(estado=filtro_estado)
        
        if filtro_cliente:
            servicios = servicios.filter(cliente_id=filtro_cliente)
        
        # Paginación
        from django.core.paginator import Paginator
        paginator = Paginator(servicios, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Opciones para filtros
        clientes = Cliente.objects.filter(activo=True).order_by('razon_social')
        
        context = {
            'page_obj': page_obj,
            'servicios': page_obj,
            'clientes': clientes,
            'filtro_estado': filtro_estado,
            'filtro_cliente': filtro_cliente,
        }
        
        return render(request, 'core/torreros/servicio_list.html', context)
        
    except Exception as e:
        logger.error(f'Error en servicio_torrero_list: {e}')
        messages.error(request, 'Error al cargar la lista de servicios')
        return redirect('torreros_dashboard')


@login_required
def servicio_torrero_create(request):
    """Crear nuevo servicio de torrero"""
    if request.method == 'POST':
        form = ServicioTorreroForm(request.POST)
        if form.is_valid():
            try:
                servicio = form.save(commit=False)
                servicio.creado_por = request.user
                
                # Asegurar que los valores numéricos estén correctos
                if not servicio.monto_pagado:
                    servicio.monto_pagado = Decimal('0.00')
                if not servicio.dias_trabajados:
                    servicio.dias_trabajados = Decimal('0.00')
                
                # Calcular monto total antes de guardar
                servicio.calcular_monto_total()
                
                servicio.save()
                
                # Los torreros se asignarán cuando se registren días trabajados
                # (no se asignan en la creación del servicio)
                
                # Log de actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='crear',
                    modulo='Servicios Torreros',
                    descripcion=f'Creó servicio de torrero para {servicio.cliente.razon_social} (ID: {servicio.id})'
                )
                
                messages.success(request, '✅ Servicio de torrero creado exitosamente')
                return redirect('servicio_torrero_detail', pk=servicio.id)
                
            except Exception as e:
                logger.error(f"Error al crear servicio: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                messages.error(request, f'❌ Error al crear servicio: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    field_label = form.fields[field].label if field in form.fields else field
                    error_messages.append(f'{field_label}: {error}')
            if error_messages:
                messages.error(request, '❌ Por favor corrige los errores del formulario:<br>' + '<br>'.join(error_messages), extra_tags='html')
            else:
                messages.error(request, '❌ Por favor corrige los errores del formulario')
    else:
        form = ServicioTorreroForm()
    
    context = {
        'form': form,
        'titulo': 'Nuevo Servicio de Torrero'
    }
    
    return render(request, 'core/torreros/servicio_form.html', context)


@login_required
def servicio_torrero_detail(request, pk):
    """Detalle de un servicio de torrero"""
    try:
        servicio = get_object_or_404(
            ServicioTorrero.objects.select_related('cliente', 'proyecto', 'creado_por'),
            pk=pk
        )
        
        # Registros de días trabajados
        registros = servicio.registros_dias.select_related(
            'registrado_por', 'aprobado_por'
        ).order_by('-fecha_registro')
        
        # Pagos
        pagos = servicio.pagos.select_related('registrado_por').order_by('-fecha_pago')
        
        # Formularios para agregar registros y pagos
        form_registro = RegistroDiasTrabajarForm()
        form_pago = PagoServicioTorreroForm()
        
        context = {
            'servicio': servicio,
            'registros': registros,
            'pagos': pagos,
            'form_registro': form_registro,
            'form_pago': form_pago,
        }
        
        return render(request, 'core/torreros/servicio_detail.html', context)
        
    except Exception as e:
        logger.error(f'Error en servicio_torrero_detail: {e}')
        messages.error(request, 'Error al cargar el detalle del servicio')
        return redirect('torreros_dashboard')


@login_required
def servicio_torrero_edit(request, pk):
    """Editar servicio de torrero"""
    servicio = get_object_or_404(ServicioTorrero, pk=pk)
    
    if request.method == 'POST':
        form = ServicioTorreroForm(request.POST, instance=servicio)
        if form.is_valid():
            try:
                servicio = form.save(commit=False)
                servicio.calcular_monto_total()
                servicio.save()
                
                # Log de actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='editar',
                    modelo='ServicioTorrero',
                    objeto_id=servicio.id,
                    descripcion=f'Editó servicio de torrero para {servicio.cliente.razon_social}'
                )
                
                messages.success(request, '✅ Servicio actualizado exitosamente')
                return redirect('servicio_torrero_detail', pk=servicio.id)
                
            except Exception as e:
                logger.error(f"Error al editar servicio: {str(e)}")
                messages.error(request, f'❌ Error al editar servicio: {str(e)}')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario')
    else:
        form = ServicioTorreroForm(instance=servicio)
    
    context = {
        'form': form,
        'servicio': servicio,
        'titulo': f'Editar Servicio - {servicio.cliente.razon_social}'
    }
    
    return render(request, 'core/torreros/servicio_form.html', context)


@login_required
def servicio_torrero_delete(request, pk):
    """Eliminar (desactivar) servicio de torrero"""
    servicio = get_object_or_404(ServicioTorrero, pk=pk)
    
    if request.method == 'POST':
        try:
            servicio.activo = False
            servicio.estado = 'cancelado'
            servicio.save()
            
            # Log de actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='eliminar',
                modulo='Servicios Torreros',
                descripcion=f'Eliminó servicio de torrero para {servicio.cliente.razon_social}'
            )
            
            messages.success(request, '✅ Servicio eliminado exitosamente')
            return redirect('servicio_torrero_list')
            
        except Exception as e:
            logger.error(f"Error al eliminar servicio: {str(e)}")
            messages.error(request, f'❌ Error al eliminar: {str(e)}')
    
    context = {
        'servicio': servicio
    }
    
    return render(request, 'core/torreros/servicio_delete.html', context)


@login_required
def registro_dias_create(request, servicio_id):
    """Crear registro de días trabajados"""
    servicio = get_object_or_404(ServicioTorrero, pk=servicio_id)
    
    if request.method == 'POST':
        form = RegistroDiasTrabajarForm(request.POST, servicio=servicio)
        if form.is_valid():
            try:
                registro = form.save(commit=False)
                registro.servicio = servicio
                registro.registrado_por = request.user
                
                # Actualizar torreros_presentes basado en los torreros seleccionados
                torreros_seleccionados = form.cleaned_data.get('torreros', [])
                registro.torreros_presentes = len(torreros_seleccionados)
                registro.save()
                
                # Asignar torreros seleccionados al servicio (si no están ya asignados)
                for torrero in torreros_seleccionados:
                    AsignacionTorrero.objects.get_or_create(
                        servicio=servicio,
                        torrero=torrero,
                        defaults={
                            'tarifa_acordada': torrero.tarifa_diaria,
                            'asignado_por': request.user,
                            'activo': True
                        }
                    )
                
                # Log de actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='crear',
                    modulo='Servicios Torreros',
                    descripcion=f'Registró {registro.dias_trabajados} día(s) con {len(torreros_seleccionados)} torrero(s) para {servicio.cliente.razon_social}'
                )
                
                messages.success(request, f'✅ {registro.dias_trabajados} día(s) registrado(s) con {len(torreros_seleccionados)} torrero(s) exitosamente')
                return redirect('servicio_torrero_detail', pk=servicio.id)
                
            except Exception as e:
                logger.error(f"Error al registrar días: {str(e)}")
                messages.error(request, f'❌ Error al registrar días: {str(e)}')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario')
    else:
        form = RegistroDiasTrabajarForm(servicio=servicio)
    
    context = {
        'form': form,
        'servicio': servicio,
        'titulo': f'Registrar Días Trabajados - {servicio.cliente.razon_social}'
    }
    
    return render(request, 'core/torreros/registro_dias_form.html', context)


@login_required
def registro_dias_quick(request, servicio_id):
    """Registro rápido de días trabajados (AJAX)"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            servicio = get_object_or_404(ServicioTorrero, pk=servicio_id)
            
            # Obtener torreros seleccionados
            torreros_ids = data.get('torreros', [])
            if not torreros_ids:
                return JsonResponse({
                    'success': False,
                    'message': 'Debes seleccionar al menos un torrero'
                }, status=400)
            
            torreros_seleccionados = Torrero.objects.filter(
                id__in=torreros_ids,
                activo=True
            )
            
            if not torreros_seleccionados.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'No se encontraron torreros válidos'
                }, status=400)
            
            # Crear registro
            from decimal import Decimal
            registro = RegistroDiasTrabajados.objects.create(
                servicio=servicio,
                fecha_registro=data.get('fecha_registro'),
                dias_trabajados=Decimal(str(data.get('dias_trabajados', 1))),
                torreros_presentes=len(torreros_seleccionados),
                observaciones=data.get('observaciones', ''),
                registrado_por=request.user,
                aprobado=True,  # Auto-aprobar registros rápidos
                aprobado_por=request.user
            )
            
            # Asignar torreros seleccionados al servicio (si no están ya asignados)
            for torrero in torreros_seleccionados:
                AsignacionTorrero.objects.get_or_create(
                    servicio=servicio,
                    torrero=torrero,
                    defaults={
                        'tarifa_acordada': torrero.tarifa_diaria,
                        'asignado_por': request.user,
                        'activo': True
                    }
                )
            
            # Log de actividad
            torreros_nombres = ', '.join([t.nombre for t in torreros_seleccionados])
            LogActividad.objects.create(
                usuario=request.user,
                accion='crear',
                modulo='Servicios Torreros',
                descripcion=f'Registró {registro.dias_trabajados} día(s) con {len(torreros_seleccionados)} torrero(s) ({torreros_nombres}) para {servicio.cliente.razon_social}'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'✅ {registro.dias_trabajados} día(s) registrado(s) con {len(torreros_seleccionados)} torrero(s) exitosamente'
            })
            
        except Exception as e:
            logger.error(f"Error en registro rápido: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required
def registro_dias_aprobar(request, pk):
    """Aprobar/desaprobar registro de días trabajados"""
    if request.method == 'POST':
        try:
            registro = get_object_or_404(RegistroDiasTrabajados, pk=pk)
            
            # Guardar el estado anterior para el log
            estado_anterior = registro.aprobado
            
            # Toggle del estado
            registro.aprobado = not registro.aprobado
            registro.aprobado_por = request.user if registro.aprobado else None
            
            # Guardar el registro (esto activará el método save() que recalculará los días del servicio)
            registro.save()
            
            # Forzar recálculo explícito usando el método del modelo
            # Importar el modelo para asegurar que obtenemos la instancia fresca
            from core.models import ServicioTorrero
            servicio = ServicioTorrero.objects.get(pk=registro.servicio.pk)
            
            # Usar el método del modelo para recalcular
            servicio.recalcular_dias_trabajados()
            
            # Refrescar para obtener los valores actualizados
            servicio.refresh_from_db()
            
            # Log de actividad
            accion = 'aprobó' if registro.aprobado else 'desaprobó'
            LogActividad.objects.create(
                usuario=request.user,
                accion='aprobar' if registro.aprobado else 'desaprobar',
                modulo='Servicios Torreros',
                descripcion=f'{accion.capitalize()} registro de {registro.dias_trabajados} día(s) para {servicio.cliente.razon_social}. Días trabajados actualizados a: {servicio.dias_trabajados} (ID: {registro.id})'
            )
            
            estado = 'aprobado' if registro.aprobado else 'desaprobado'
            
            # Recalcular días restantes y porcentaje
            dias_restantes = max(Decimal('0.00'), servicio.dias_solicitados - servicio.dias_trabajados)
            
            return JsonResponse({
                'success': True,
                'aprobado': registro.aprobado,
                'message': f'Registro {estado} exitosamente. Días trabajados actualizados: {servicio.dias_trabajados}',
                'dias_trabajados': float(servicio.dias_trabajados),
                'dias_solicitados': float(servicio.dias_solicitados),
                'dias_restantes': float(dias_restantes),
                'porcentaje_completado': float(servicio.porcentaje_completado),
                'monto_total': float(servicio.monto_total),
                'monto_pagado': float(servicio.monto_pagado),
                'saldo_pendiente': float(servicio.saldo_pendiente)
            })
            
        except Exception as e:
            logger.error(f"Error al aprobar/desaprobar registro: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required
def registro_dias_delete(request, pk):
    """Eliminar registro de días trabajados"""
    if request.method == 'POST':
        try:
            registro = get_object_or_404(RegistroDiasTrabajados, pk=pk)
            servicio = registro.servicio
            dias_trabajados = registro.dias_trabajados
            cliente_nombre = servicio.cliente.razon_social
            
            # Guardar información para el log antes de eliminar
            registro_id = registro.id
            
            # Eliminar el registro
            registro.delete()
            
            # Recalcular los días trabajados del servicio
            from core.models import ServicioTorrero
            servicio = ServicioTorrero.objects.get(pk=servicio.pk)
            servicio.recalcular_dias_trabajados()
            servicio.refresh_from_db()
            
            # Log de actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='eliminar',
                modulo='Servicios Torreros',
                descripcion=f'Eliminó registro de {dias_trabajados} día(s) para {cliente_nombre}. Días trabajados actualizados a: {servicio.dias_trabajados} (ID eliminado: {registro_id})'
            )
            
            # Retornar datos actualizados del servicio para actualizar la UI
            return JsonResponse({
                'success': True,
                'message': f'✅ Registro eliminado exitosamente. Días trabajados actualizados a: {servicio.dias_trabajados}',
                'dias_trabajados': float(servicio.dias_trabajados),
                'dias_solicitados': float(servicio.dias_solicitados),
                'dias_restantes': float(servicio.dias_restantes),
                'porcentaje_completado': float(servicio.porcentaje_completado),
                'monto_total': float(servicio.monto_total),
                'monto_pagado': float(servicio.monto_pagado),
                'saldo_pendiente': float(servicio.saldo_pendiente)
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar registro: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required
def servicio_torrero_toggle_pago(request, servicio_id):
    """Toggle estado de pago del servicio (AJAX)"""
    if request.method == 'POST':
        try:
            servicio = get_object_or_404(ServicioTorrero, pk=servicio_id)
            
            # Si está pagado, marcar como pendiente (monto_pagado = 0)
            # Si está pendiente, marcar como pagado (monto_pagado = monto_total)
            if servicio.esta_pagado:
                servicio.monto_pagado = Decimal('0.00')
                estado = 'pendiente'
                mensaje = 'Servicio marcado como pendiente'
            else:
                servicio.monto_pagado = servicio.monto_total
                estado = 'pagado'
                mensaje = 'Servicio marcado como pagado'
            
            servicio.save()
            
            # Log de actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='editar',
                modulo='Servicios Torreros',
                descripcion=f'Marcó servicio de {servicio.cliente.razon_social} como {estado} (ID: {servicio.id})'
            )
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'estado': estado,
                'monto_pagado': float(servicio.monto_pagado),
                'saldo_pendiente': float(servicio.saldo_pendiente),
                'esta_pagado': servicio.esta_pagado
            })
            
        except Exception as e:
            logger.error(f"Error al cambiar estado de pago: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required
def pago_servicio_create(request, servicio_id):
    """Crear pago de servicio de torrero"""
    servicio = get_object_or_404(ServicioTorrero, pk=servicio_id)
    
    if request.method == 'POST':
        form = PagoServicioTorreroForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                pago = form.save(commit=False)
                pago.servicio = servicio
                pago.registrado_por = request.user
                pago.save()
                
                # Log de actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='crear',
                    modelo='PagoServicioTorrero',
                    objeto_id=pago.id,
                    descripcion=f'Registró pago de ${pago.monto} para {servicio.cliente.razon_social}'
                )
                
                messages.success(request, '✅ Pago registrado exitosamente')
                return redirect('servicio_torrero_detail', pk=servicio.id)
                
            except Exception as e:
                logger.error(f"Error al registrar pago: {str(e)}")
                messages.error(request, f'❌ Error al registrar pago: {str(e)}')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario')
    
    return redirect('servicio_torrero_detail', pk=servicio.id)


# ============================================================================
# CRUD DE TORREROS (CATÁLOGO)
# ============================================================================

@login_required
def torreros_list(request):
    """Lista de torreros registrados"""
    try:
        torreros = Torrero.objects.filter(activo=True).order_by('nombre')
        
        # Filtros
        busqueda = request.GET.get('q', '')
        if busqueda:
            torreros = torreros.filter(
                Q(nombre__icontains=busqueda) |
                Q(cedula__icontains=busqueda) |
                Q(especialidad__icontains=busqueda)
            )
        
        context = {
            'torreros': torreros,
            'busqueda': busqueda,
        }
        
        return render(request, 'core/torreros/torreros_list.html', context)
        
    except Exception as e:
        logger.error(f'Error en torreros_list: {e}')
        messages.error(request, 'Error al cargar la lista de torreros')
        return redirect('torreros_dashboard')


@login_required
def torrero_create(request):
    """Crear nuevo torrero"""
    if request.method == 'POST':
        form = TorreroForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                torrero = form.save(commit=False)
                torrero.creado_por = request.user
                torrero.save()
                
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='crear',
                    modulo='Torreros',
                    descripcion=f'Creó torrero: {torrero.nombre} (ID: {torrero.id})'
                )
                
                messages.success(request, f'✅ Torrero {torrero.nombre} creado exitosamente')
                return redirect('torreros_list')
                
            except Exception as e:
                logger.error(f"Error al crear torrero: {str(e)}")
                messages.error(request, f'❌ Error al crear torrero: {str(e)}')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario')
    else:
        form = TorreroForm()
    
    context = {
        'form': form,
        'titulo': 'Nuevo Torrero'
    }
    
    return render(request, 'core/torreros/torrero_form.html', context)


@login_required
def torrero_edit(request, pk):
    """Editar torrero"""
    torrero = get_object_or_404(Torrero, pk=pk)
    
    if request.method == 'POST':
        form = TorreroForm(request.POST, request.FILES, instance=torrero)
        if form.is_valid():
            try:
                torrero = form.save()
                
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='editar',
                    modulo='Torreros',
                    descripcion=f'Editó torrero: {torrero.nombre} (ID: {torrero.id})'
                )
                
                messages.success(request, f'✅ Torrero {torrero.nombre} actualizado exitosamente')
                return redirect('torreros_list')
                
            except Exception as e:
                logger.error(f"Error al editar torrero: {str(e)}")
                messages.error(request, f'❌ Error al editar torrero: {str(e)}')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario')
    else:
        form = TorreroForm(instance=torrero)
    
    context = {
        'form': form,
        'torrero': torrero,
        'titulo': f'Editar Torrero - {torrero.nombre}'
    }
    
    return render(request, 'core/torreros/torrero_form.html', context)


@login_required
def torrero_delete(request, pk):
    """Eliminar (desactivar) torrero"""
    torrero = get_object_or_404(Torrero, pk=pk)
    
    if request.method == 'POST':
        try:
            torrero.activo = False
            torrero.save()
            
            LogActividad.objects.create(
                usuario=request.user,
                accion='eliminar',
                modulo='Torreros',
                descripcion=f'Eliminó torrero: {torrero.nombre} (ID: {torrero.id})'
            )
            
            messages.success(request, f'✅ Torrero {torrero.nombre} eliminado exitosamente')
            return redirect('torreros_list')
            
        except Exception as e:
            logger.error(f"Error al eliminar torrero: {str(e)}")
            messages.error(request, f'❌ Error al eliminar: {str(e)}')
    
    context = {
        'torrero': torrero
    }
    
    return render(request, 'core/torreros/torrero_delete.html', context)


@login_required
def torrero_registro_rapido(request):
    """Registro rápido de torrero (solo datos esenciales)"""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            tarifa_diaria = request.POST.get('tarifa_diaria', '0')
            
            if not nombre:
                return JsonResponse({
                    'success': False,
                    'message': 'El nombre es obligatorio'
                }, status=400)
            
            # Generar cédula automática única (basada en timestamp)
            import time
            cedula = f"T-{int(time.time())}"
            
            # Verificar que no exista (muy improbable)
            while Torrero.objects.filter(cedula=cedula).exists():
                cedula = f"T-{int(time.time())}-{Torrero.objects.count()}"
            
            # Crear torrero
            torrero = Torrero.objects.create(
                nombre=nombre,
                cedula=cedula,
                tarifa_diaria=Decimal(tarifa_diaria) if tarifa_diaria else Decimal('0.00'),
                fecha_ingreso=timezone.now().date(),
                creado_por=request.user
            )
            
            # Log de actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='crear',
                modulo='Torreros',
                descripcion=f'Registro rápido de torrero: {torrero.nombre} (ID: {torrero.id})'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Torrero {nombre} registrado exitosamente',
                'torrero_id': torrero.id
            })
            
        except Exception as e:
            logger.error(f"Error en registro rápido de torrero: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error al registrar: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método no permitido'
    }, status=405)


@login_required
def servicio_torrero_pdf(request, pk):
    """Generar PDF de respaldo del servicio de torrero"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from io import BytesIO
    
    servicio = get_object_or_404(ServicioTorrero, pk=pk)
    registros = RegistroDiasTrabajados.objects.filter(servicio=servicio).order_by('fecha_registro')
    pagos = PagoServicioTorrero.objects.filter(servicio=servicio).order_by('fecha_pago')
    # Obtener torreros asignados a través de AsignacionTorrero
    asignaciones = AsignacionTorrero.objects.filter(servicio=servicio, activo=True).select_related('torrero')
    torreros = [asignacion.torrero for asignacion in asignaciones]
    
    # Crear el PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados mejorados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=26,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=32
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=25,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold',
        borderWidth=0,
        borderPadding=0,
        backColor=colors.HexColor('#f1f5f9'),
        leftIndent=10,
        rightIndent=10,
        topPadding=8,
        bottomPadding=8
    )
    
    # Encabezado profesional con diseño moderno
    from django.utils import timezone
    fecha_actual = timezone.localtime(timezone.now())
    
    # Título principal
    header_text = f"""
    <para align="center" spaceAfter="12">
        <font size="26" name="Helvetica-Bold" color="#0f172a">REPORTE DE SERVICIO</font><br/>
        <font size="26" name="Helvetica-Bold" color="#6366f1">TORREROS</font>
    </para>
    """
    elements.append(Paragraph(header_text, styles['Normal']))
    
    # Subtítulo
    elements.append(Paragraph(f"<para align='center'><font size='11' color='#64748b'>Sistema ARCA - Generado el {fecha_actual.strftime('%d/%m/%Y a las %I:%M %p')}</font></para>", styles['Normal']))
    
    # Línea decorativa
    elements.append(Spacer(1, 0.2*inch))
    divider_data = [['']]
    divider_table = Table(divider_data, colWidths=[7*inch])
    divider_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (0, 0), 2, colors.HexColor('#6366f1')),
    ]))
    elements.append(divider_table)
    elements.append(Spacer(1, 0.25*inch))
    
    # Información del Servicio - Diseño mejorado
    info_heading = Paragraph("INFORMACIÓN DEL SERVICIO", heading_style)
    elements.append(info_heading)
    
    info_data = [
        ['Cliente:', servicio.cliente.razon_social],
        ['Proyecto:', servicio.proyecto.nombre if servicio.proyecto else 'No asignado'],
        ['Descripción:', servicio.descripcion or 'Sin descripción'],
        ['Estado del Servicio:', Paragraph(f"{'PAGADO' if servicio.esta_pagado else 'PENDIENTE'}", ParagraphStyle('EstadoStyle', parent=styles['Normal'], textColor=colors.HexColor('#10b981') if servicio.esta_pagado else colors.HexColor('#f59e0b'), fontName='Helvetica-Bold'))],
        ['', ''],  # Separador
        ['Días Solicitados:', f"{servicio.dias_solicitados} día(s)"],
        ['Días Trabajados:', f"{servicio.dias_trabajados} día(s)"],
        ['Días Restantes:', f"{servicio.dias_restantes} día(s)"],
        ['Progreso:', f"{servicio.porcentaje_completado}%"],
        ['', ''],  # Separador
        ['Tarifa Diaria:', f"${servicio.tarifa_por_dia:,.2f}"],
        ['Monto Total:', Paragraph(f"${servicio.monto_total:,.2f}", ParagraphStyle('MontoStyle', parent=styles['Normal'], textColor=colors.HexColor('#1e40af'), fontName='Helvetica-Bold'))],
        ['Total Pagado:', Paragraph(f"${servicio.monto_pagado:,.2f}", ParagraphStyle('MontoStyle', parent=styles['Normal'], textColor=colors.HexColor('#10b981'), fontName='Helvetica-Bold'))],
        ['Saldo Pendiente:', Paragraph(f"${servicio.saldo_pendiente:,.2f}", ParagraphStyle('MontoStyle', parent=styles['Normal'], textColor=colors.HexColor('#dc2626'), fontName='Helvetica-Bold'))],
    ]
    
    info_table = Table(info_data, colWidths=[2.2*inch, 4.3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#ffffff')),  # Separador
        ('BACKGROUND', (0, 9), (-1, 9), colors.HexColor('#ffffff')),  # Separador
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#6366f1')),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # No mostrar tabla de torreros asignados aquí, se mostrarán en el detalle de registros
    
    # Registro de Días Trabajados con Detalle Completo - Diseño Mejorado
    if registros.exists():
        elements.append(Paragraph("REGISTRO DETALLADO DE DÍAS TRABAJADOS", heading_style))
        
        # Obtener todas las asignaciones activas para el servicio
        asignaciones_activas = AsignacionTorrero.objects.filter(servicio=servicio, activo=True).select_related('torrero')
        lista_torreros_asignados = [a.torrero.nombre for a in asignaciones_activas]
        
        # Nueva tabla mejorada: Fecha, Torreros que Trabajaron, Días, Estado
        registros_data = [['FECHA TRABAJADA', 'TORREROS', 'DÍAS', 'ESTADO']]
        
        for registro in registros:
            # Obtener nombres de torreros (usar los asignados al servicio)
            torreros_nombres = ", ".join(lista_torreros_asignados) if lista_torreros_asignados else "No asignado"
            
            # Formatear fecha
            fecha_str = registro.fecha_registro.strftime('%d/%m/%Y')
            
            # Estado con ícono
            estado_str = '✓ Aprobado' if registro.aprobado else '⏳ Pendiente'
            
            registros_data.append([
                fecha_str,
                torreros_nombres,
                f"{registro.dias_trabajados}",
                estado_str
            ])
        
        # Tabla más ancha y profesional
        registros_table = Table(registros_data, colWidths=[1.5*inch, 3.5*inch, 0.8*inch, 1.2*inch])
        registros_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            # Body
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Fecha centrada
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Torreros izquierda
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Días centrada
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Estado centrada
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 1), (-1, -1), 8),
            ('RIGHTPADDING', (0, 1), (-1, -1), 8),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#6366f1')),
            # Row backgrounds alternados
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(registros_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Resumen mejorado de días trabajados
        elements.append(Paragraph("RESUMEN DE DÍAS", heading_style))
        resumen_dias_data = [
            ['Total Días Solicitados:', f"{servicio.dias_solicitados} día(s)"],
            ['Total Días Trabajados:', f"{servicio.dias_trabajados} día(s)"],
            ['Días Restantes:', f"{servicio.dias_restantes} día(s)"],
            ['Progreso:', Paragraph(f"{servicio.porcentaje_completado}%", ParagraphStyle('ProgressStyle', parent=styles['Normal'], textColor=colors.HexColor('#10b981'), fontName='Helvetica-Bold'))],
        ]
        
        resumen_dias_table = Table(resumen_dias_data, colWidths=[3*inch, 3.5*inch])
        resumen_dias_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#6366f1')),
        ]))
        elements.append(resumen_dias_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Historial de Pagos
    if pagos.exists():
        elements.append(Paragraph("HISTORIAL DE PAGOS", heading_style))
        
        pagos_data = [['Fecha', 'Monto', 'Método', 'Referencia', 'Registrado Por']]
        for pago in pagos:
            pagos_data.append([
                pago.fecha_pago.strftime('%d/%m/%Y'),
                f"${pago.monto:,.2f}",
                pago.metodo_pago,
                pago.numero_referencia or '-',
                pago.registrado_por.get_full_name() or pago.registrado_por.username
            ])
        
        pagos_table = Table(pagos_data, colWidths=[1.2*inch, 1.2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        pagos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        elements.append(pagos_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Pie de página mejorado
    elements.append(Spacer(1, 0.6*inch))
    
    # Línea decorativa superior
    footer_divider = Table([['']], colWidths=[7*inch])
    footer_divider.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (0, 0), 1, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(footer_divider)
    elements.append(Spacer(1, 0.2*inch))
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER,
        leading=12
    )
    
    footer_text = f"""
    <para align="center" spaceAfter="4">
        <font size="9" color="#64748b">
            <b>Sistema ARCA</b> - Technology Panama INC.<br/>
            Documento generado automáticamente por: <b>{request.user.get_full_name() or request.user.username}</b><br/>
            <font size="8">Fecha de generación: {fecha_actual.strftime('%d/%m/%Y a las %I:%M:%S %p')}</font>
        </font>
    </para>
    """
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"Servicio_Torrero_{servicio.cliente.razon_social.replace(' ', '_')}_{servicio.id}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    # Log de actividad
    LogActividad.objects.create(
        usuario=request.user,
        accion='generar_pdf',
        modulo='Servicios Torreros',
        descripcion=f'Generó PDF de servicio para {servicio.cliente.razon_social}'
    )
    
    return response


# ============================================
# VISTAS DE SUBPROYECTOS
# ============================================

@login_required
def subproyecto_create(request, proyecto_id):
    """Crear subproyecto dentro de un proyecto"""
    from .forms_simple import SubproyectoForm
    
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    
    if request.method == 'POST':
        form = SubproyectoForm(request.POST, proyecto=proyecto)
        if form.is_valid():
            try:
                subproyecto = form.save(commit=False)
                subproyecto.proyecto = proyecto
                subproyecto.creado_por = request.user
                subproyecto.save()
                
                # Log de actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='crear',
                    modulo='Subproyectos',
                    descripcion=f'Creó subproyecto {subproyecto.codigo} en {proyecto.nombre}'
                )
                
                messages.success(request, '✅ Subproyecto creado exitosamente')
                return redirect('proyectos_list')  # Cambiar a proyecto_detail cuando exista
                
            except Exception as e:
                logger.error(f"Error al crear subproyecto: {str(e)}")
                messages.error(request, f'❌ Error al crear: {str(e)}')
    else:
        form = SubproyectoForm(proyecto=proyecto)
    
    context = {
        'form': form,
        'proyecto': proyecto,
        'titulo': f'Nuevo Subproyecto - {proyecto.nombre}'
    }
    
    return render(request, 'core/subproyectos/form.html', context)


@login_required
def subproyecto_edit(request, pk):
    """Editar subproyecto"""
    from .forms_simple import SubproyectoForm
    
    subproyecto = get_object_or_404(Subproyecto, pk=pk)
    proyecto = subproyecto.proyecto
    
    if request.method == 'POST':
        form = SubproyectoForm(request.POST, instance=subproyecto, proyecto=proyecto)
        if form.is_valid():
            try:
                subproyecto = form.save()
                
                # Log de actividad
                LogActividad.objects.create(
                    usuario=request.user,
                    accion='editar',
                    modulo='Subproyectos',
                    descripcion=f'Editó subproyecto {subproyecto.codigo}'
                )
                
                messages.success(request, '✅ Subproyecto actualizado exitosamente')
                return redirect('proyectos_list')  # Cambiar a proyecto_detail cuando exista
                
            except Exception as e:
                logger.error(f"Error al editar subproyecto: {str(e)}")
                messages.error(request, f'❌ Error al editar: {str(e)}')
    else:
        form = SubproyectoForm(instance=subproyecto, proyecto=proyecto)
    
    context = {
        'form': form,
        'subproyecto': subproyecto,
        'proyecto': proyecto,
        'titulo': f'Editar Subproyecto - {subproyecto.nombre}'
    }
    
    return render(request, 'core/subproyectos/form.html', context)


@login_required
def subproyecto_delete(request, pk):
    """Eliminar (desactivar) subproyecto"""
    subproyecto = get_object_or_404(Subproyecto, pk=pk)
    proyecto = subproyecto.proyecto
    
    if request.method == 'POST':
        try:
            subproyecto.activo = False
            subproyecto.estado = 'cancelado'
            subproyecto.save()
            
            # Log de actividad
            LogActividad.objects.create(
                usuario=request.user,
                accion='eliminar',
                modulo='Subproyectos',
                descripcion=f'Eliminó subproyecto {subproyecto.codigo}'
            )
            
            messages.success(request, '✅ Subproyecto eliminado exitosamente')
            return redirect('proyectos_list')  # Cambiar a proyecto_detail cuando exista
            
        except Exception as e:
            logger.error(f"Error al eliminar subproyecto: {str(e)}")
            messages.error(request, f'❌ Error al eliminar: {str(e)}')
    
    context = {
        'subproyecto': subproyecto,
        'proyecto': proyecto
    }
    
    return render(request, 'core/subproyectos/delete.html', context)


@login_required
def subproyectos_dashboard(request, proyecto_id):
    """Dashboard de rentabilidad de subproyectos"""
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    
    # Obtener subproyectos activos
    subproyectos = Subproyecto.objects.filter(
        proyecto=proyecto,
        activo=True
    ).select_related('cotizacion', 'creado_por').order_by('-creado_en')
    
    # Calcular totales
    total_cotizado = sum(sub.monto_cotizado for sub in subproyectos)
    total_ingresos = sum(sub.total_ingresos for sub in subproyectos)
    total_gastos = sum(sub.total_gastos for sub in subproyectos)
    rentabilidad_total = total_ingresos - total_gastos
    
    # Preparar datos para gráficos
    subproyectos_data = []
    for sub in subproyectos:
        subproyectos_data.append({
            'id': sub.id,
            'codigo': sub.codigo,
            'nombre': sub.nombre,
            'monto_cotizado': float(sub.monto_cotizado),
            'total_ingresos': float(sub.total_ingresos),
            'total_gastos': float(sub.total_gastos),
            'rentabilidad': float(sub.rentabilidad),
            'margen_rentabilidad': float(sub.margen_rentabilidad),
            'estado': sub.estado,
            'porcentaje_avance': float(sub.porcentaje_avance),
        })
    
    import json
    subproyectos_json = json.dumps(subproyectos_data)
    
    context = {
        'proyecto': proyecto,
        'subproyectos': subproyectos,
        'total_cotizado': total_cotizado,
        'total_ingresos': total_ingresos,
        'total_gastos': total_gastos,
        'rentabilidad_total': rentabilidad_total,
        'subproyectos_json': subproyectos_json,
        'titulo': f'Dashboard de Subproyectos - {proyecto.nombre}'
    }
    
    return render(request, 'core/subproyectos/dashboard.html', context)


@login_required
def planillas_liquidadas_historial(request):
    """Vista completa para consultar todas las planillas liquidadas pasadas"""
    from django.db.models import Q, Sum
    from django.core.paginator import Paginator
    
    # Obtener parámetros de búsqueda
    proyecto_id = request.GET.get('proyecto')
    año = request.GET.get('año')
    mes = request.GET.get('mes')
    tipo_planilla = request.GET.get('tipo_planilla', 'todas')  # 'todas', 'personal', 'trabajadores_diarios'
    search = request.GET.get('search', '')
    
    # Query inicial
    planillas = PlanillaLiquidada.objects.select_related(
        'proyecto', 'liquidada_por'
    ).order_by('-año', '-mes', '-quincena', '-fecha_liquidacion')
    
    # Filtro por tipo de planilla
    if tipo_planilla == 'personal':
        # Solo planillas de personal (sin observaciones de trabajadores diarios)
        planillas = planillas.exclude(observaciones__icontains='trabajadores diarios')
    elif tipo_planilla == 'trabajadores_diarios':
        # Solo planillas de trabajadores diarios
        planillas = planillas.filter(observaciones__icontains='trabajadores diarios')
    # Si es 'todas', no aplicamos filtro
    
    # Filtros adicionales
    if proyecto_id:
        planillas = planillas.filter(proyecto_id=proyecto_id)
    
    if año:
        planillas = planillas.filter(año=año)
    
    if mes:
        planillas = planillas.filter(mes=mes)
    
    if search:
        planillas = planillas.filter(
            Q(proyecto__nombre__icontains=search) |
            Q(observaciones__icontains=search)
        )
    
    # Paginación
    paginator = Paginator(planillas, 25)  # 25 por página
    page = request.GET.get('page')
    planillas_page = paginator.get_page(page)
    
    # Calcular totales de las planillas filtradas
    total_planillas = planillas.count()
    total_general = planillas.aggregate(
        total=Sum('total_planilla')
    )['total'] or Decimal('0.00')
    
    total_salarios_general = planillas.aggregate(
        total=Sum('total_salarios')
    )['total'] or Decimal('0.00')
    
    total_anticipos_general = planillas.aggregate(
        total=Sum('total_anticipos')
    )['total'] or Decimal('0.00')
    
    # Calcular estadísticas generales (todas las planillas, sin filtro de tipo)
    todas_planillas = PlanillaLiquidada.objects.all()
    total_planillas_personal = todas_planillas.exclude(
        observaciones__icontains='trabajadores diarios'
    ).count()
    total_planillas_trabajadores_diarios = todas_planillas.filter(
        observaciones__icontains='trabajadores diarios'
    ).count()
    total_planillas_todas = todas_planillas.count()
    
    total_general_personal = todas_planillas.exclude(
        observaciones__icontains='trabajadores diarios'
    ).aggregate(total=Sum('total_planilla'))['total'] or Decimal('0.00')
    
    total_general_trabajadores_diarios = todas_planillas.filter(
        observaciones__icontains='trabajadores diarios'
    ).aggregate(total=Sum('total_planilla'))['total'] or Decimal('0.00')
    
    # Obtener proyectos para el filtro
    proyectos = Proyecto.objects.filter(
        planillas_liquidadas__isnull=False
    ).distinct().order_by('nombre')
    
    # Obtener años disponibles
    años = PlanillaLiquidada.objects.values_list('año', flat=True).distinct().order_by('-año')
    
    context = {
        'planillas': planillas_page,
        'total_planillas': total_planillas,
        'total_general': total_general,
        'total_salarios_general': total_salarios_general,
        'total_anticipos_general': total_anticipos_general,
        'proyectos': proyectos,
        'años': años,
        'proyecto_selected': proyecto_id,
        'año_selected': año,
        'mes_selected': mes,
        'tipo_planilla_selected': tipo_planilla,
        'search': search,
        'meses_choices': PlanillaLiquidada.MESES_CHOICES,
        # Estadísticas generales por tipo
        'total_planillas_todas': total_planillas_todas,
        'total_planillas_personal': total_planillas_personal,
        'total_planillas_trabajadores_diarios': total_planillas_trabajadores_diarios,
        'total_general_personal': total_general_personal,
        'total_general_trabajadores_diarios': total_general_trabajadores_diarios,
    }
    
    return render(request, 'core/planillas/historial.html', context)


@login_required
def planilla_liquidada_delete(request, planilla_id):
    """Eliminar una planilla liquidada"""
    planilla = get_object_or_404(PlanillaLiquidada, id=planilla_id)
    proyecto = planilla.proyecto
    
    if request.method == 'POST':
        # Guardar información antes de eliminar para el log
        proyecto_nombre = proyecto.nombre
        mes_nombre = planilla.get_mes_display()
        año = planilla.año
        quincena_nombre = planilla.get_quincena_display()
        total_planilla = planilla.total_planilla
        
        # Eliminar la planilla
        planilla.delete()
        
        # Registrar actividad
        LogActividad.objects.create(
            usuario=request.user,
            accion='Eliminar Planilla Liquidada',
            modulo='Planillas Liquidadas',
            descripcion=f'Planilla liquidada eliminada: {proyecto_nombre} - {mes_nombre} {año} ({quincena_nombre}) - Total: ${total_planilla:,.2f}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(
            request,
            f'✅ Planilla liquidada eliminada exitosamente.<br>'
            f'<strong>Proyecto:</strong> {proyecto_nombre}<br>'
            f'<strong>Período:</strong> {mes_nombre} {año} ({quincena_nombre})<br>'
            f'<strong>Total eliminado:</strong> ${total_planilla:,.2f}',
            extra_tags='html'
        )
        
        # Redirigir de vuelta al historial manteniendo los filtros
        redirect_url = reverse('planillas_liquidadas_historial')
        params = request.GET.copy()
        if params:
            redirect_url += '?' + params.urlencode()
        return redirect(redirect_url)
    
    # Si es GET, mostrar confirmación
    return render(request, 'core/planillas/delete.html', {
        'planilla': planilla,
        'proyecto': proyecto
    })


@login_required
def trabajadores_diarios_dashboard(request):
    """Dashboard principal del módulo de Trabajadores Diarios"""
    from django.db.models import Sum, Count, Q
    
    # Obtener todos los trabajadores diarios activos
    trabajadores = TrabajadorDiario.objects.select_related(
        'proyecto', 'planilla', 'creado_por'
    ).filter(
        activo=True
    ).order_by('-fecha_registro')
    
    # Estadísticas generales
    total_trabajadores = trabajadores.count()
    
    # Obtener proyectos con trabajadores diarios
    proyectos_con_trabajadores = Proyecto.objects.filter(
        trabajadores_diarios__activo=True
    ).distinct().order_by('nombre')
    
    # Calcular totales (usando propiedades del modelo)
    total_a_pagar = sum(t.total_a_pagar for t in trabajadores)
    total_dias_trabajados = sum(t.total_dias_trabajados for t in trabajadores)
    total_anticipos = sum(t.total_anticipos_aplicados for t in trabajadores)
    
    # Obtener trabajadores recientes (últimos 10)
    trabajadores_recientes = trabajadores[:10]
    
    # Obtener proyectos más activos (con más trabajadores diarios)
    proyectos_mas_activos = Proyecto.objects.filter(
        trabajadores_diarios__activo=True
    ).annotate(
        num_trabajadores=Count('trabajadores_diarios', filter=Q(trabajadores_diarios__activo=True))
    ).order_by('-num_trabajadores')[:5]
    
    # Obtener TODOS los proyectos activos para el selector
    todos_proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'total_trabajadores': total_trabajadores,
        'total_a_pagar': total_a_pagar,
        'total_dias_trabajados': total_dias_trabajados,
        'total_anticipos': total_anticipos,
        'trabajadores_recientes': trabajadores_recientes,
        'proyectos_con_trabajadores': proyectos_con_trabajadores,
        'proyectos_mas_activos': proyectos_mas_activos,
        'todos_proyectos': todos_proyectos,
        'titulo': 'Dashboard de Trabajadores Diarios'
    }
    
    return render(request, 'core/trabajadores-diarios/dashboard.html', context)


@login_required
def planillas_trabajadores_diarios_gestor(request):
    """Gestor de planillas de trabajadores diarios - Vista general de todas las planillas"""
    from django.db.models import Count, Q
    
    # Filtros
    proyecto_id = request.GET.get('proyecto')
    estado_filter = request.GET.get('estado')
    
    # Obtener todas las planillas
    planillas = PlanillaTrabajadoresDiarios.objects.select_related(
        'proyecto', 'creada_por', 'finalizada_por'
    ).annotate(
        num_trabajadores=Count('trabajadores', filter=Q(trabajadores__activo=True))
    ).order_by('-fecha_creacion')
    
    # Obtener proyectos para el filtro (todos los proyectos activos, no solo los que tienen planillas)
    proyectos_filtro = Proyecto.objects.filter(
        planillas_trabajadores_diarios__isnull=False
    ).distinct().order_by('nombre')
    
    # Obtener TODOS los proyectos para poder crear planillas (mostrar todos, no solo activos)
    todos_proyectos = Proyecto.objects.all().order_by('nombre')
    
    # Estadísticas sobre TODAS las planillas (no filtradas)
    todas_planillas = PlanillaTrabajadoresDiarios.objects.all()
    total_planillas = todas_planillas.count()
    planillas_activas = todas_planillas.filter(estado='activa').count()
    planillas_finalizadas = todas_planillas.filter(estado='finalizada').count()
    
    # Aplicar filtros
    if proyecto_id:
        planillas = planillas.filter(proyecto_id=proyecto_id)
    
    if estado_filter:
        planillas = planillas.filter(estado=estado_filter)
    
    context = {
        'planillas': planillas,
        'proyectos': proyectos_filtro,
        'todos_proyectos': todos_proyectos,
        'total_planillas': total_planillas,
        'planillas_activas': planillas_activas,
        'planillas_finalizadas': planillas_finalizadas,
        'proyecto_selected': proyecto_id,
        'estado_selected': estado_filter,
    }
    
    return render(request, 'core/trabajadores_diarios/gestor_planillas.html', context)
